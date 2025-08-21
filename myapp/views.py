from django.shortcuts import render, redirect 
from django.http import HttpResponse, JsonResponse, HttpResponseBadRequest
from django.db import connection, DatabaseError, OperationalError
from django.contrib.admin.views.decorators import staff_member_required 
from datetime import datetime, date, timedelta, time
import seaborn as sns  # type: ignore
import pandas as pd
import os
from django.conf import settings
import matplotlib
matplotlib.use('Agg') 
import matplotlib.pyplot as plt
import plotly.graph_objects as go
import json
import plotly.express as px
from decimal import Decimal
import random
from plotly.subplots import make_subplots
from django.contrib.auth.decorators import user_passes_test
import dash
from dash import Dash, dcc, html, dash_table
import re
from collections import defaultdict
import requests
import base64
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from decouple import config
import hmac
import hashlib
from geopy.geocoders import Nominatim
import pycountry
from fpdf import FPDF
import csv
import json
import urllib.parse
from urllib.parse import urlparse
from plotly.offline import plot
from plotly.graph_objects import Figure, Table
from django.contrib import messages
import jwt
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from collections import OrderedDict
from calendar import month_name
import boto3
import io
import contextlib
import traceback
from io import StringIO
import numpy as np
import gender_guesser.detector as gender

## Done

JWT_SECRET = config('SECRET_KEY')
JWT_ALGORITHM = config('JWT_ALGORITHM')
JWT_EXP_DELTA_SECONDS = config('JWT_EXP_DELTA_SECONDS')


def generate_token(username):
    payload = {
        'username': username,
        'exp': datetime.utcnow() + timedelta(seconds=JWT_EXP_DELTA_SECONDS)  
    }
    token = jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)
    print(token)
    return token




def first(request):
    return HttpResponse("Successfull Connected")


@staff_member_required
def code_terminal(request):
    output = ''
    if request.method == 'POST':
        code = request.POST.get('code', '')
        stdout = io.StringIO()
        local_vars = {}

        try:
            with contextlib.redirect_stdout(stdout):
                exec(code, globals(), local_vars)

            response_obj = local_vars.get("response") or globals().get("response")
            if isinstance(response_obj, HttpResponse):
                return response_obj  # 🟢 Trigger file download


            output = stdout.getvalue()
            
        except Exception:
            output = traceback.format_exc()
        finally:
            stdout.close()

    return render(request, 'code_terminal.html', {'output': output})


## "For Downloading the File in code_terminal"
# csv_buffer = StringIO()
# data.to_csv(csv_buffer, index=False)  ## Changed 'data' with Dataframe
# csv_data = csv_buffer.getvalue()

# # Create response for direct download
# response = HttpResponse(csv_data, content_type='text/csv')
# response['Content-Disposition'] = 'attachment; filename="downloaded_file.csv"'

# globals()['response'] = response

# print("Download Complete")   


class Parent:
    def first(request):
        error_message = ""
        monthly_data = []
        chart_filename = ""
        year = ''
        month_most_package = []
        data_for_selected_date = []
        data_headers = []
        avg_date = []
        current_month_year = datetime.now().strftime("%Y-%m")

        month_for_info = request.GET.get('month_for_info',current_month_year)
        selected_date = request.POST.get('selected_date')
        
        details_for = request.POST.get('details_for')
        title = request.POST.get('title')
        
        try:
            if month_for_info:
                year, month = month_for_info.split('-')
                with connection.cursor() as cursor:
                    cursor.execute(f"""EXEC usp_Django_MostSearchDate_MostLoginDate {year}
                                """)
                    columns = [col[0] for col in cursor.description]
                    avg_date  = [dict(zip(columns, row)) for row in cursor.fetchall()]
                
                with connection.cursor() as cursor:
                    #For Same Day Registration and Verification 
                    cursor.execute(f"""EXEC usp_Django_Agents_Registration_Verification_by_Month {year}, {month}
                                """)
                    
                    monthly_data = cursor.fetchall()

                with connection.cursor() as cursor:
                    cursor.execute(f"""EXEC usp_Django_Most_Searched_Package_By_Month {year}, {month}
                                    """)
                    month_most_package = cursor.fetchall()

                with connection.cursor() as cursor:
                    cursor.execute(f"""EXEC usp_Django_Agents_Registration_Verification_by_Year {year}
                                """)
                    data_to_show=cursor.fetchall()
                df = pd.DataFrame(data_to_show, columns=['Month', 'Count_of_Agents_Registered', 'Count_of_Verified_Agents'])

                # Mapping month numbers to names
                month_map = {
                    1: "January", 2: "February", 3: "March", 4: "April", 5: "May", 6: "June", 
                    7: "July", 8: "August", 9: "September", 10: "October", 11: "November", 12: "December"
                }
                df["Month"] = df["Month"].map(month_map)

                # Ensure Month is categorical and ordered properly
                month_order = list(month_map.values())  # Ensure correct order
                df["Month"] = pd.Categorical(df["Month"], categories=month_order, ordered=True)
                df = df.sort_values(by="Month")

                fig, ax = plt.subplots(figsize=(12, 6))

                bar_width = 0.4  
                x = range(len(df))  

                ax.bar(x, df["Count_of_Verified_Agents"], width=bar_width, label="Verified Agents", color="blue", align="center")
                ax.bar([p + bar_width for p in x], df["Count_of_Agents_Registered"], width=bar_width, label="Registered Agents", color="green", align="center")

                # Set labels and title
                ax.set_xlabel("Month")
                ax.set_ylabel("Count")
                ax.set_title(f"Registered vs Verified Agents {year}")
                ax.set_xticks([p + bar_width / 2 for p in x]) 
                ax.set_xticklabels(df["Month"], rotation=45)  

                # Add legend
                ax.legend()

                # Save plot as image in Django static folder
                chart_filename = f"bar_chart_{year}.png"
                chart_path = os.path.join(settings.BASE_DIR, 'staticfiles', 'Charts', chart_filename)
                os.makedirs(os.path.dirname(chart_path), exist_ok=True)
                plt.savefig(chart_path, bbox_inches='tight')
                plt.close()

            if selected_date:

                if details_for == "search":
                    
                    with connection.cursor() as cursor:
                        cursor.execute(f"""EXEC usp_Django_Most_Search_Tour_By_Date '{selected_date}'
                                    """) 
                        data_headers = [col[0] for col in cursor.description]
                        data_for_selected_date = cursor.fetchall()
                    

                elif details_for == "30days":
                    with connection.cursor() as cursor:
                        cursor.execute(f"""EXEC usp_Django_Agent_LoginAndRegis_Last30days '{selected_date}'
                                    """)
                        data_headers = [col[0] for col in cursor.description]
                        data_for_selected_date = cursor.fetchall()


                elif details_for == "15days":
                    with connection.cursor() as cursor:
                        cursor.execute(f"""EXEC usp_Django_Agent_LoginAndRegis_Last15days '{selected_date}'
                                    """)
                        data_headers = [col[0] for col in cursor.description]
                        data_for_selected_date = cursor.fetchall()


                elif details_for == "05days":
                    with connection.cursor() as cursor:
                        cursor.execute(f"""EXEC usp_Django_Agent_LoginAndRegis_Last05days '{selected_date}'
                                    """)
                        data_headers = [col[0] for col in cursor.description]
                        data_for_selected_date = cursor.fetchall()
                
        except OperationalError: # For Database Connection Error
            return render(request, 'no_database.html', {})
        
        except DatabaseError as e:  # For Database Error or Query Error
            error_message = f"Database Error: {str(e)}"
            return render(request, 'no_database.html', {})
        
        except Exception as e: # For Network Connection Error
            if "Internet" in str(e):  
                return render(request, 'no_internet.html', {})
            else:
                error_message = f"Unknown Error: {str(e)}"
                #return render(request, 'general_error.html', {})

        executed_query = connection.queries[-1]['sql'] if connection.queries else "No query executed"

        return render(request, 'first.html', {
            'executed_query' : executed_query,
            'error_message' : error_message,
            'monthly_data' : monthly_data,
            'month_for_info' : month_for_info,
            "chart_path": f"Charts/{chart_filename}",
            'year' : year,
            'month_most_package' : month_most_package,
            'data_for_selected_date' : data_for_selected_date,
            'selected_date' : selected_date,
            'data_headers' : data_headers,
            'title' : title,
            'details_for' : details_for,
            'avg_date' : avg_date,
        })
     

class CRMGraphCopy:
    def summary(request):
        error_message = "Test"
        current_year = datetime.now().year
        chart_show = request.GET.get('chart_show','Agent Registration Comparison')
        select_package = request.GET.get('select_package',8)
        chart_json = ''


        try:
            if chart_show == "Agent Registration Comparison":
                with connection.cursor() as cursor:
                    cursor.execute(f"""EXEC usp_Django_Agent_Registration_TillLast3Years
                                """)
                    columns = [col[0] for col in cursor.description]
                    data_to_show=cursor.fetchall()

                    df = pd.DataFrame(data_to_show, columns=['Month', (current_year-3), (current_year-2), (current_year-1), (current_year)])

                    x= []
                    y_current_year_3 = []
                    y_current_year_2 = []
                    y_current_year_1 = [] 
                    y_current_year = []

                    for i in df["Month"]:
                        x.append(i)

                    for i in df[(current_year-3)]:
                        y_current_year_3.append(i)

                    for i in df[(current_year-2)]:
                        y_current_year_2.append(i)

                    for i in df[(current_year-1)]:
                        y_current_year_1.append(i)

                    for i in df[(current_year)]:
                        y_current_year.append(i)

                    trace1 = go.Bar(
                        x=x, 
                        y=y_current_year_3, 
                        marker=dict(color='#0099ff'), 
                        name=str(current_year-3),
                        text=[f"{val:.0f}" for val in y_current_year_3], 
                        # textposition='outside',
                        textfont=dict(size=12),
                        hovertemplate="Year: " + str(current_year - 3) + "<br> %{y}",
                        )
                
                    trace2 = go.Bar(
                        x=x, 
                        y=y_current_year_2, 
                        marker=dict(color="#E257A4"), 
                        name=str(current_year-2),
                        text=[f"{val:.0f}" for val in y_current_year_2], 
                        # textposition='outside',
                        textfont=dict(size=12),
                        hovertemplate="Year: " + str(current_year - 2) + "<br> %{y}",
                        )
            
                    trace3 = go.Bar(
                        x=x, 
                        y=y_current_year_1, 
                        marker=dict(color='#ff9900'), 
                        name=str(current_year-1),
                        text=[f"{val:.0f}" for val in y_current_year_1], 
                        # textposition='outside',
                        textfont=dict(size=12),
                        hovertemplate="Year: " + str(current_year - 1) + "<br> %{y}",
                        )
                    
                    trace4 = go.Bar(
                        x=x, 
                        y=y_current_year, 
                        marker=dict(color='#28a745'), 
                        name=str(current_year),
                        text=[f"{val:.0f}" for val in y_current_year], 
                        # textposition='outside',
                        textfont=dict(size=12),
                        hovertemplate="Year: " + str(current_year) + "<br> %{y}",
                        )

                    fig = go.Figure(data=[trace1, trace2, trace3, trace4])

                    fig.update_layout(
                        title='Agents Registration Comparison',
                        xaxis=dict(title='Months'),
                        yaxis=dict(title='Agents Registration'),
                        barmode='group',  # Grouped bar chart
                        margin=dict(t=25, l=25, r=25, b=25),  # Adjust margins
                        height=600,  # Set height
                        autosize=True,   # Set width
                    )

                    chart_json = fig.to_json()

            elif chart_show == "Deposit Comparison":
                with connection.cursor() as cursor:
                    cursor.execute("""EXEC usp_Django_DepositAmount_TillLast3Years
                                """)
                    columns = [col[0] for col in cursor.description]
                    data_to_show=cursor.fetchall()

                    df = pd.DataFrame(data_to_show, columns=['Month', (current_year-3), (current_year-2), (current_year-1), (current_year)])

                    x = df["Month"].tolist()
                    y_current_year_3 = df[current_year-3].tolist()
                    y_current_year_2 = df[current_year-2].tolist()
                    y_current_year_1 = df[current_year-1].tolist()
                    y_current_year = df[current_year].tolist()

                    trace1 = go.Bar(
                        x=x, 
                        y=y_current_year_3, 
                        marker=dict(color='#0099ff'), 
                        name=str(current_year-3),
                        text=[f"{val:.0f}" for val in y_current_year_3], 
                        # textposition='outside',
                        textfont=dict(size=12),
                        hovertemplate="Year: " + str(current_year - 3) + "<br>Month: %{x} <br> Amount: $ %{y:.2f}",
                        )
                    
                    trace2 = go.Bar(
                        x=x, 
                        y=y_current_year_2, 
                        marker=dict(color='#E257A4'), 
                        name=str(current_year-2),
                        text=[f"{val:.0f}" for val in y_current_year_2], 
                        # textposition='outside',
                        textfont=dict(size=12),
                        hovertemplate="Year: " + str(current_year - 2) + "<br>Month: %{x} <br> Amount: $ %{y:.2f}",
                        )
                

                    trace3 = go.Bar(
                        x=x, 
                        y=y_current_year_1, 
                        marker=dict(color='#ff9900'), 
                        name=str(current_year-1),
                        text=[f"{val:.0f}" for val in y_current_year_1], 
                        # textposition='outside',
                        textfont=dict(size=12),
                        hovertemplate="Year: " + str(current_year - 1) + "<br>Month: %{x} <br> Amount: $ %{y:.2f}",
                        )
                    
                    trace4 = go.Bar(
                        x=x, 
                        y=y_current_year, 
                        marker=dict(color='#28a745'), 
                        name=str(current_year),
                        text=[f"{val:.0f}" for val in y_current_year], 
                        # textposition='outside',
                        textfont=dict(size=12),
                        hovertemplate="Year: " + str(current_year) + "<br>Month: %{x} <br> Amount: $ %{y:.2f}",
                        )

                    fig = go.Figure(data=[trace1, trace2, trace3, trace4])

                    fig.update_layout(
                        title=f'Monthly Deposit Amount',
                        xaxis=dict(title='Months'),
                        yaxis=dict(title='Deposit Amount'),
                        barmode='group',  # Grouped bar chart
                        margin=dict(t=25, l=25, r=25, b=25),
                        height=600,  # Set height
                        autosize=True,   # Set width
                    )

                    chart_json = fig.to_json()


            elif chart_show == "Package Comparison":
                
                with connection.cursor() as cursor:
                    cursor.execute(f"""EXEC usp_Django_DepositAmountOfPackage_TillLast3Years {select_package}
                                """)
                    columns = [col[0] for col in cursor.description]
                    data_to_show=cursor.fetchall()
                    
                    df = pd.DataFrame(data_to_show, columns=['Month', (current_year-3), (current_year-2), (current_year-1), (current_year)])

                    x = df["Month"]
                    y_current_year_3 = df[current_year-3].tolist()
                    y_current_year_2 = df[current_year-2].tolist()
                    y_current_year_1 = df[current_year-1].tolist()
                    y_current_year = df[current_year].tolist()

                    
                    trace1 = go.Bar(
                        x=x, 
                        y=y_current_year_3, 
                        marker=dict(color='#0099ff'), 
                        name=str(current_year-3),
                        text=[f"{val:.0f}" for val in y_current_year_3], 
                        # textposition='outside',
                        textfont=dict(size=12),
                        hovertemplate="Year: " + str(current_year - 3) + "<br>Month: %{x} <br> Amount: $ %{y:.2f}",
                        )
                    
                    trace2 = go.Bar(
                        x=x, 
                        y=y_current_year_2, 
                        marker=dict(color='#E257A4'), 
                        name=str(current_year-2),
                        text=[f"{val:.0f}" for val in y_current_year_2], 
                        # textposition='outside',
                        textfont=dict(size=12),
                        hovertemplate="Year: " + str(current_year - 2) + "<br>Month: %{x} <br> Amount: $ %{y:.2f}",
                        )
                

                    trace3 = go.Bar(
                        x=x, 
                        y=y_current_year_1, 
                        marker=dict(color='#ff9900'), 
                        name=str(current_year-1),
                        text=[f"{val:.0f}" for val in y_current_year_1], 
                        # textposition='outside',
                        textfont=dict(size=12),
                        hovertemplate="Year: " + str(current_year - 1) + "<br>Month: %{x} <br> Amount: $ %{y:.2f}",
                        )
                    
                    trace4 = go.Bar(
                        x=x, 
                        y=y_current_year, 
                        marker=dict(color='#28a745'), 
                        name=str(current_year),
                        text=[f"{val:.0f}" for val in y_current_year], 
                        # textposition='outside',
                        textfont=dict(size=12),
                        hovertemplate="Year: " + str(current_year) + "<br>Month: %{x} <br> Amount: $ %{y:.2f}",
                        )

                    fig = go.Figure(data=[trace1, trace2, trace3, trace4])

                    fig.update_layout(
                        title='Monthly Deposit Amount',
                        xaxis=dict(title='Months'),
                        yaxis=dict(title='Deposit Amount'),
                        barmode='group',  # Grouped bar chart
                        margin=dict(t=25, l=25, r=25, b=25),
                        height=600,  # Set height
                        autosize=True,   # Set width
                    )

                    chart_json = fig.to_json()
                    

        except:
            pass
        return render(request, 'summary.html', {
            'error_message' : error_message,
            "chart_json": chart_json,
            'chart_show' : chart_show,
            'select_package' : select_package,

        })

    def login_details(request):
        error_message = "Test"
        current_year = datetime.now().year
        data_details = request.POST.get('data_details','')
        view_type = request.GET.get('view_type','Login Summary')
        chart_json = '' 
        

        try:
            if view_type == "Login Summary":
                with connection.cursor() as cursor:
                    if data_details == "DISTINCT":
                        query = """EXEC usp_Django_UniqueAgentLogin_TillLast3Years
                                """
                    else:
                        query = """EXEC usp_Django_TotalAgentLogin_TillLast3Years
                        """  
                    cursor.execute(query)
                    data_to_show = cursor.fetchall()
                df = pd.DataFrame(data_to_show, columns=['Month', (current_year-3), (current_year-2), (current_year-1), (current_year)])
                x = df["Month"]
                y_current_year_3 = df[current_year-3].tolist()
                y_current_year_2 = df[current_year-2].tolist()
                y_current_year_1 = df[current_year-1].tolist()
                y_current_year = df[current_year].tolist()

                
                trace1 = go.Bar(
                    x=x, 
                    y=y_current_year_3, 
                    marker=dict(color='#0099ff'), 
                    name=str(current_year-3),
                    text=[f"{val:.0f}" for val in y_current_year_3], 
                    #  textposition='outside',
                    textfont=dict(size=12),
                    hovertemplate="Year: " + str(current_year-3) + "<br>Month: %{x} <br> Agents: %{y:.0f}",
                    )
                
                trace2 = go.Bar(
                    x=x, 
                    y=y_current_year_2, 
                    marker=dict(color='#E257A4'), 
                    name=str(current_year-2),
                    text=[f"{val:.0f}" for val in y_current_year_2], 
                    # textposition='outside',
                    textfont=dict(size=12),
                    hovertemplate="Year: " + str(current_year-2) + "<br>Month: %{x} <br> Agents: %{y:.0f}",
                    )
                

                trace3 = go.Bar(
                    x=x, 
                    y=y_current_year_1, 
                    marker=dict(color='#ff9900'), 
                    name=str(current_year-1),
                    text=[f"{val:.0f}" for val in y_current_year_1], 
                    # textposition='outside',
                    textfont=dict(size=12),
                    hovertemplate="Year: " + str(current_year-1) + "<br>Month: %{x} <br> Agents: %{y:.0f}",
                    )
                
                trace4 = go.Bar(
                    x=x, 
                    y=y_current_year, 
                    marker=dict(color='#28a745'), 
                    name=str(current_year),
                    text=[f"{val:.0f}" for val in y_current_year], 
                    # textposition='outside',
                    textfont=dict(size=12),
                    hovertemplate="Year: " + str(current_year) + "<br>Month: %{x} <br> Agents: %{y:.0f}",
                    )

                fig = go.Figure(data=[trace1, trace2, trace3, trace4])

                if data_details == "DISTINCT":
                    chart_title='Monthly Unique Agents Login'
                else:
                    chart_title='Monthly Overall Agents Login'

                fig.update_layout(
                    title=chart_title,
                    xaxis=dict(title='Months'),
                    yaxis=dict(title='No. of Agents'),
                    barmode='group',  # Grouped bar chart
                    margin=dict(t=25, l=25, r=25, b=25),
                    height=600,  # Set height
                    autosize=True,   # Set width
                )

                chart_json = fig.to_json()
            

            elif view_type == "Payment Summary":
                with connection.cursor() as cursor:
                    if data_details == "DISTINCT":
                        query = """EXEC usp_Django_NoOfDepositByUniqueAgents_TillLast3Years
                                """

                    else:
                        query = """EXEC usp_Django_TotalNoOfDepositByAgents_TillLast3Years
                                """
                        
                        
                    cursor.execute(query)
                    data_to_show = cursor.fetchall()
                    
                df = pd.DataFrame(data_to_show, columns=['Month', (current_year-3), (current_year-2), (current_year-1), (current_year)])
                x = df["Month"]
                y_current_year_3 = df[current_year-3].tolist()
                y_current_year_2 = df[current_year-2].tolist()
                y_current_year_1 = df[current_year-1].tolist()
                y_current_year = df[current_year].tolist()


                trace1 = go.Bar(
                    x=x, 
                    y=y_current_year_3, 
                    marker=dict(color='#0099ff'), 
                    name=str(current_year-3),
                    text=[f"{val:.0f}" for val in y_current_year_3], 
                    # textposition='outside',
                    textfont=dict(size=12),
                    hovertemplate="Year: " + str(current_year-3) + "<br>Month: %{x} <br> Agents: %{y:.0f}",
                    )
                
                trace2 = go.Bar(
                    x=x, 
                    y=y_current_year_2, 
                    marker=dict(color='#E257A4'), 
                    name=str(current_year-2),
                    text=[f"{val:.0f}" for val in y_current_year_2], 
                    # textposition='outside',
                    textfont=dict(size=12),
                    hovertemplate="Year: " + str(current_year-2) + "<br>Month: %{x} <br> Agents: %{y:.0f}",
                    )
                

                trace3 = go.Bar(
                    x=x, 
                    y=y_current_year_1, 
                    marker=dict(color='#ff9900'), 
                    name=str(current_year-1),
                    text=[f"{val:.0f}" for val in y_current_year_1], 
                    # textposition='outside',
                    textfont=dict(size=12),
                    hovertemplate="Year: " + str(current_year-1) + "<br>Month: %{x} <br> Agents: %{y:.0f}",
                    )
                
                trace4 = go.Bar(
                    x=x, 
                    y=y_current_year, 
                    marker=dict(color='#28a745'), 
                    name=str(current_year),
                    text=[f"{val:.0f}" for val in y_current_year], 
                    # textposition='outside',
                    textfont=dict(size=12),
                    hovertemplate="Year: " + str(current_year) + "<br>Month: %{x} <br> Agents: %{y:.0f}",
                    )

                fig = go.Figure(data=[trace1, trace2, trace3, trace4])

                if data_details == "DISTINCT":
                    chart_title='Monthly No of Unique Agents Payments'
                else:
                    chart_title='Monthly No of Overall Agents Payments'

                fig.update_layout(
                    title=chart_title,
                    xaxis=dict(title='Months'),
                    yaxis=dict(title='No. of Payments'),
                    barmode='group',  # Grouped bar chart
                    margin=dict(t=25, l=25, r=25, b=25),
                    height=600,  # Set height
                    autosize=True,   # Set width
                
                )

                chart_json = fig.to_json()

        except:
            pass

        return render(request, 'login_details.html', {
            'error_message' : error_message,
            'data_details' : data_details,
            'view_type' : view_type,
            'chart_json' : chart_json,
        })

    def sales_report(request):
        print("Start")
        chart_json_1 = ''
        chart_json_2 = ''
        chart_json_3 = ''
        table_json = ''
        package_title_id= ''
        selected_pkg_title = request.GET.get("selected_pkg_title")
        selected_pkg_id = request.GET.get("selected_pkg_id")
        #convert_rate =  get_usd_to_inr()

        if selected_pkg_id:
            add_query = f"AND b.PackgID = '{selected_pkg_id}'"
        else:
            add_query = ''

        try:
            with connection.cursor() as cursor:
                cursor.execute(f"""EXEC usp_Django_get_NotCustomize_titles
                        """)
                package_title_id = cursor.fetchall()
            
            with connection.cursor() as cursor:
                cursor.execute(f"""
                        WITH DateSeries AS (
                            SELECT CAST(GETDATE() AS DATE) AS Date
                            UNION ALL
                            SELECT DATEADD(DAY, -1, Date) FROM DateSeries WHERE Date > DATEADD(DAY, -29, CAST(GETDATE() AS DATE))
                        )
                        SELECT 
                            d.Date, 
                            COALESCE(SUM(CAST(b.USDamt AS MONEY)), 0)*87 AS Total_Amount
                        FROM DateSeries d
                        LEFT JOIN tbl_booking b 
                            ON d.Date = TRY_CONVERT(DATE, b.CreatedDate, 103) 
                            AND b.txn_msg = 'success'
                            AND b.Is_cancelled <> 1
                            {add_query}
                        GROUP BY d.Date
                        ORDER BY d.Date DESC;
                            """)
                data_to_show_1 = cursor.fetchall()

            df_1 = pd.DataFrame(data_to_show_1, columns=['Date', 'Total Amount'])

            x_1 = [f"{date.strftime('%Y-%m-%d')}, ({amount:.0f})" for date, amount in zip(df_1['Date'], df_1['Total Amount'])]
            y_1 = df_1['Total Amount']

            colors = ['red', 'blue', 'green', 'purple', 'orange', 'cyan', 'magenta'] * (len(df_1) // 7 + 1)
            
            fig = go.Figure()

            for num_bars in range(5, 31, 5):
                fig.add_trace(go.Bar(
                    x=x_1[:num_bars], 
                    y=y_1[:num_bars], 
                    text=[f"₹ {val:.0f}" for val in y_1], 
                    textposition='outside',
                    textfont=dict(size=12, color='black'),
                    hovertemplate="Amount: ₹ %{y:.2f}<extra></extra>",
                    marker=dict(color=colors[:num_bars], line=dict(width=1.5, color='black')),
                    name=f"Last {num_bars} Days",  # Legend label
                    visible=(num_bars == 5)  # Default to 5 bars
                    
                ))

            fig.update_layout(
                    title=dict(
                            text="<b>Last 30 Days Sales</b>",
                            font=dict(size=20, family="Times New Roman, Times, serif", color="black", weight="bold")
                        ),

                    xaxis=dict(
                        title='Dates',
                        type='category',  # Treat x-axis as categorical so all dates appear
                        tickangle=-45,
                        showgrid=False,
                        zeroline=False,
                        tickfont=dict(size=12, color='black')
                        ),

                    yaxis=dict(
                        title='Total Amount',
                        showgrid=True,
                        gridcolor='lightgrey',
                        zeroline=True,
                        zerolinecolor='black',
                        tickfont=dict(size=12, color='black')
                        ),

                    margin=dict(t=25, l=25, r=25, b=25),
                    height=600,  # Set height
                    autosize=True,   # Set width
                    paper_bgcolor="#ECF0F1",  # Change the outer background color
                    plot_bgcolor="#FAFAFA", # Change the inner plot area background color
                    updatemenus=[{
                        "buttons": [
                            {"label": f"Last {num_bars} Days", "method": "update", "args": [{"visible": [i == j for j in range(len(range(5, 31, 5)))]}]}  
                            for i, num_bars in enumerate(range(5, 31, 5))
                        ],
                        "direction": "right",
                        "pad": {"r": 10, "t": 10},
                        "showactive": True,
                        "x": 0.17,
                        "xanchor": "left",
                        "y": 1.2,
                        "yanchor": "top",
                        "bgcolor": "white",  # Background color of the dropdown
                        "bordercolor": "black",  # Border color of the dropdown
                        "borderwidth": 1  # Border width
                    }]
                )

            chart_json_1 = fig.to_json()

            with connection.cursor() as cursor:
                cursor.execute(f"""
                            WITH DateSeries AS (
                                SELECT CAST(GETDATE() AS DATE) AS Date
                                UNION ALL
                                SELECT DATEADD(DAY, -1, Date) FROM DateSeries WHERE Date > DATEADD(DAY, -29, CAST(GETDATE() AS DATE))
                            )
                            SELECT 
                                d.Date, 
                                COALESCE(SUM(CAST(b.USDamt AS MONEY)), 0)*87 AS Total_Amount
                            FROM DateSeries d
                            LEFT JOIN tbl_booking b 
                                ON d.Date = TRY_CONVERT(DATE, b.CreatedDate, 103) 
                                AND b.txn_msg = 'success'
                                AND b.Is_cancelled <> 1
                                {add_query}
                                AND b.Paymode in ('Deposit', 'MinimumDeposit')
                            GROUP BY d.Date
                            ORDER BY d.Date DESC;
                                """)
                data_to_show_2 = cursor.fetchall()

            df_2 = pd.DataFrame(data_to_show_2, columns=['Date', 'Total Amount'])

            x_2 = [f"{date.strftime('%Y-%m-%d')}, ({amount:.0f})" for date, amount in zip(df_2['Date'], df_2['Total Amount'])]
            y_2 = df_2['Total Amount']

            colors = ['red', 'blue', 'green', 'purple', 'orange', 'cyan', 'magenta'] * (len(df_1) // 7 + 1)

            fig = go.Figure()

            for num_bars in range(5, 31, 5):
                fig.add_trace(go.Bar(
                    x=x_2[:num_bars], 
                    y=y_2[:num_bars], 
                    text=[f"₹ {val:.0f}" for val in y_2], 
                    textposition='outside',
                    textfont=dict(size=12, color='black'),
                    hovertemplate="Amount: ₹ %{y:.2f}<extra></extra>",
                    marker=dict(color=colors[:num_bars], line=dict(width=1.5, color='black')),
                    name=f"Last {num_bars} Days",  # Legend label
                    visible=(num_bars == 5)  # Default to 10 bars
                ))

            fig.update_layout(
                    title=dict(
                        text="<b>Last 30 Days Deposit</b>",
                        font=dict(size=20, family="Times New Roman, Times, serif", color="black", weight="bold")
                    ),
                    xaxis=dict(
                        title='Dates',
                        type='category',  # Treat x-axis as categorical so all dates appear
                        tickangle=-45,
                        showgrid=False,
                        zeroline=False,
                        tickfont=dict(size=12, color='black'),
                        ),

                    yaxis=dict(
                        title='Total Amount',
                        showgrid=True,
                        gridcolor='lightgrey',
                        zeroline=True,
                        zerolinecolor='black',
                        tickfont=dict(size=12, color='black'),
                        ),

                    margin=dict(t=25, l=25, r=25, b=25),
                    height=600,  # Set height
                    autosize=True,   # Set width
                    paper_bgcolor="#ECF0F1",  # Change the outer background color
                    plot_bgcolor="#FAFAFA", # Change the inner plot area background color
                    updatemenus=[{
                        "buttons": [
                            {"label": f"Last {num_bars} Days", "method": "update", "args": [{"visible": [i == j for j in range(len(range(5, 31, 5)))]}]}  
                            for i, num_bars in enumerate(range(5, 31, 5))
                        ],
                        "direction": "right",
                        "pad": {"r": 10, "t": 10},
                        "showactive": True,
                        "x": 0.17,
                        "xanchor": "left",
                        "y": 1.2,
                        "yanchor": "top",
                        "bgcolor": "white",  # Background color of the dropdown
                        "bordercolor": "black",  # Border color of the dropdown
                        "borderwidth": 1  # Border width
                    }]
                )

            chart_json_2 = fig.to_json()

            with connection.cursor() as cursor:
                cursor.execute(f"""EXEC usp_Django_NoOfDeposits_CurrentYear_For_PKGID_8
                            """)
                data_to_show_3 = cursor.fetchall()
            df_3 = pd.DataFrame(data_to_show_3, columns=['Month', 'No of Tours'])
        
            x_3 = df_3["Month"].tolist()
            y_3 = df_3["No of Tours"].tolist()


            with connection.cursor() as cursor:
                cursor.execute(f"""EXEC usp_Django_NoOfDeposits_CurrentYear_For_PKGID_57
                            """)
                data_to_show_4 = cursor.fetchall()

            df_4 = pd.DataFrame(data_to_show_4, columns=['Month', 'No of Tours'])
        
            x_4 = df_4["Month"].tolist()
            y_4 = df_4["No of Tours"].tolist()

            fig = make_subplots(
                rows=1, 
                cols=2, 
                subplot_titles=[
                    "<b>Dashing Dubai Trip (Deposit)</b>", 
                    "<b>Facinating Bali Trip (Deposit)</b>"])
        
            # Package 8 (Dubai) Bar Chart (Left Side)
            fig.add_trace(go.Bar(
                x=x_3,
                y=y_3,
                marker=dict(color="Blue", opacity=0.80, line=dict(width=1.5, color='black')),
                text=[f"{val:.0f}" for val in y_3], 
                textposition='outside',
                textfont=dict(size=12, family='Arial Black', color='#2C3E50'),
                name="",
                hovertemplate="%{x}<br>No. of Bookings: %{y}",
                width=[0.5] * len(x_3),
            ), row=1, col=1)

            # Package 57 (Bali) Bar Chart (Right Side)
            fig.add_trace(go.Bar(
                x=x_4,
                y=y_4,
                marker=dict(color="red", opacity=0.80,line=dict(width=1.5, color='black')),
                text=[f"{val:.0f}" for val in y_4], 
                textposition='outside',
                textfont=dict(size=12, family='Arial Black', color='#2C3E50'),
                name="",
                hovertemplate="%{x}<br>No. of Bookings: %{y}",
                width=[0.5] * len(x_3),
            ), row=1, col=2)

            # Layout settings
            fig.update_layout(
                title=dict(
                        text="<b>Monthly Deposit Comparison</b>",
                        font=dict(size=20, family="Times New Roman, Times, serif", color="black", weight="bold"),
                        x=0.5,
                    ),
                #xaxis_title="",
                yaxis_title="<b>No of Bookings</b>",
                
                showlegend=False,  # Hide duplicate legends
                height=600,
                autosize=True,  # Ensure enough space
                paper_bgcolor="#ECF0F1",  # Change the outer background color
                plot_bgcolor="#FAFAFA", # Change the inner plot area background color
                margin=dict(l=50, r=50, t=80, b=50),
                font=dict(size=14, family="Arial", color="#2C3E50"),
                hoverlabel=dict(font_size=12)
            )

            chart_json_3 = fig.to_json()

        except:
            pass

        return render(request, 'sales_report.html', {
            'chart_json_1' : chart_json_1,
            'chart_json_2' : chart_json_2,
            'chart_json_3' : chart_json_3,
            'table_json' : table_json,
            'package_title_id':package_title_id,
            "selected_pkg_title": selected_pkg_title,
            "selected_pkg_id": selected_pkg_id,

        })
        

class test:
    def query_report(request):
        msg_chart = request.GET.get('msg_chart','Package_Query')
        current_year = datetime.now().year
        data_to_show = ''
        chart_json =''
        selected_country = request.GET.get('selected_country','all')
        countries =[]
        query = ''

        
        if msg_chart == "Package_Query":
            with connection.cursor() as cursor:
                cursor.execute(f"""EXEC usp_Django_CountriesWithQueryCount_ForPackagesRelatedQuery
                            """)
                countries = [row[0] for row in cursor.fetchall()]


            if selected_country == "all":
                query = """EXEC usp_Django_MonthlyQueryCount_ForPackageQuery_TillLast3years
                        """
            
            elif selected_country == "Others":
                query = """EXEC usp_Django_MonthlyCountofNullPKGID_ForPackageQuery_TillLast3years
                            """
                
            elif selected_country != "Others" or selected_country != "all":
                query = f"""WITH Months AS (
                                SELECT 1 AS MonthNumber UNION ALL
                                SELECT 2 UNION ALL
                                SELECT 3 UNION ALL
                                SELECT 4 UNION ALL
                                SELECT 5 UNION ALL
                                SELECT 6 UNION ALL
                                SELECT 7 UNION ALL
                                SELECT 8 UNION ALL
                                SELECT 9 UNION ALL
                                SELECT 10 UNION ALL
                                SELECT 11 UNION ALL
                                SELECT 12
                            ),
                            MonthlyData AS (
                                SELECT
                                    MONTH(TRY_CONVERT(DATE, m.CREATED_DATE, 103)) AS MonthNumber,
                                    SUM(CASE WHEN YEAR(TRY_CONVERT(DATE, m.CREATED_DATE, 103)) = YEAR(GETDATE()) - 3 THEN 1 ELSE 0 END) AS Previous_3_Year,
                                    SUM(CASE WHEN YEAR(TRY_CONVERT(DATE, m.CREATED_DATE, 103)) = YEAR(GETDATE()) - 2 THEN 1 ELSE 0 END) AS Previous_2_Year,
                                    SUM(CASE WHEN YEAR(TRY_CONVERT(DATE, m.CREATED_DATE, 103)) = YEAR(GETDATE()) - 1 THEN 1 ELSE 0 END) AS Previous_Year,
                                    SUM(CASE WHEN YEAR(TRY_CONVERT(DATE, m.CREATED_DATE, 103)) = YEAR(GETDATE()) THEN 1 ELSE 0 END) AS Current_Year
                                FROM TBL_MESSAGE m
                                JOIN TBL_PKG_DETAILS d
                                    ON m.PKG_ID = d.PKG_ID
                                WHERE
                                    (m.MSG_TYPE = 'holidays' OR m.MSG_TYPE = 'package')
                                    AND YEAR(TRY_CONVERT(DATE, m.CREATED_DATE, 103)) BETWEEN YEAR(GETDATE()) - 3 AND YEAR(GETDATE())
                                    AND d.country = '{selected_country}'
                                GROUP BY
                                    MONTH(TRY_CONVERT(DATE, m.CREATED_DATE, 103))
                            )
                            SELECT
                                DATENAME(MONTH, DATEFROMPARTS(2000, m.MonthNumber, 1)) AS Month,
                                COALESCE(md.Previous_3_Year, 0) AS Previous_3_Year,
                                COALESCE(md.Previous_2_Year, 0) AS Previous_2_Year,
                                COALESCE(md.Previous_Year, 0) AS Previous_Year,
                                COALESCE(md.Current_Year, 0) AS Current_Year
                            FROM Months m
                            LEFT JOIN MonthlyData md
                                ON m.MonthNumber = md.MonthNumber
                            ORDER BY
                                m.MonthNumber;
                            """
                

        elif msg_chart == "Customized_Trip_Query":
            with connection.cursor() as cursor:
                cursor.execute(f"""EXEC usp_Django_CountriesWithQueryCount_ForCustomizedTrips
                            """)
                countries = [row[0] for row in cursor.fetchall()]

            if selected_country == "all":
                query = """EXEC usp_Django_MonthlyQueryCount_ForCustomizedTrips_TillLast3years
                            """

            elif selected_country == "Others":
                query = """EXEC usp_Django_MonthlyCountofNullPKGID_ForCustomizedTrips_TillLast3years
                        """

            elif selected_country != "Others" or selected_country != "all":
                query = f"""WITH Months AS (
                                SELECT 1 AS MonthNumber UNION ALL
                                SELECT 2 UNION ALL
                                SELECT 3 UNION ALL
                                SELECT 4 UNION ALL
                                SELECT 5 UNION ALL
                                SELECT 6 UNION ALL
                                SELECT 7 UNION ALL
                                SELECT 8 UNION ALL
                                SELECT 9 UNION ALL
                                SELECT 10 UNION ALL
                                SELECT 11 UNION ALL
                                SELECT 12
                            ),
                            MonthlyData AS (
                                SELECT
                                    MONTH(TRY_CONVERT(DATE, m.CREATED_DATE, 103)) AS MonthNumber,
                                    SUM(CASE WHEN YEAR(TRY_CONVERT(DATE, m.CREATED_DATE, 103)) = YEAR(GETDATE()) - 3 THEN 1 ELSE 0 END) AS Previous_3_Year,
                                    SUM(CASE WHEN YEAR(TRY_CONVERT(DATE, m.CREATED_DATE, 103)) = YEAR(GETDATE()) - 2 THEN 1 ELSE 0 END) AS Previous_2_Year,
                                    SUM(CASE WHEN YEAR(TRY_CONVERT(DATE, m.CREATED_DATE, 103)) = YEAR(GETDATE()) - 1 THEN 1 ELSE 0 END) AS Previous_Year,
                                    SUM(CASE WHEN YEAR(TRY_CONVERT(DATE, m.CREATED_DATE, 103)) = YEAR(GETDATE()) THEN 1 ELSE 0 END) AS Current_Year
                                FROM TBL_MESSAGE m
                                JOIN TBL_PKG_DETAILS d
                                    ON m.PKG_ID = d.PKG_ID
                                WHERE
                                    m.MSG_TYPE = 'CustomizedTrip'
                                    AND YEAR(TRY_CONVERT(DATE, m.CREATED_DATE, 103)) BETWEEN YEAR(GETDATE()) - 3 AND YEAR(GETDATE())
                                    AND d.country = '{selected_country}'
                                GROUP BY
                                    MONTH(TRY_CONVERT(DATE, m.CREATED_DATE, 103))
                            )
                            SELECT
                                DATENAME(MONTH, DATEFROMPARTS(2000, m.MonthNumber, 1)) AS Month,
                                COALESCE(md.Previous_3_Year, 0) AS Previous_3_Year,
                                COALESCE(md.Previous_2_Year, 0) AS Previous_2_Year,
                                COALESCE(md.Previous_Year, 0) AS Previous_Year,
                                COALESCE(md.Current_Year, 0) AS Current_Year
                            FROM Months m
                            LEFT JOIN MonthlyData md
                                ON m.MonthNumber = md.MonthNumber
                            ORDER BY
                                m.MonthNumber;
                """
      
        elif msg_chart == "Offline_Query":
            with connection.cursor() as cursor:
                cursor.execute(f"""EXEC usp_Django_CountriesWithQueryCount_ForOfflineRequest
                            """)
                countries = [row[0] for row in cursor.fetchall()]


            if selected_country == "all":
                query = """EXEC usp_Django_MonthlyQueryCount_ForOfflineRequest_TillLast3years
                        """
                
            elif selected_country == "Others":
                query = """EXEC usp_Django_MonthlyCountofNullPKGID_ForOfflineRequest_TillLast3years
                        """
                
            elif selected_country != "Others" or selected_country != "all":
                query = f"""WITH Months AS (
                                SELECT 1 AS MonthNumber UNION ALL
                                SELECT 2 UNION ALL
                                SELECT 3 UNION ALL
                                SELECT 4 UNION ALL
                                SELECT 5 UNION ALL
                                SELECT 6 UNION ALL
                                SELECT 7 UNION ALL
                                SELECT 8 UNION ALL
                                SELECT 9 UNION ALL
                                SELECT 10 UNION ALL
                                SELECT 11 UNION ALL
                                SELECT 12
                            ),
                            MonthlyData AS (
                                SELECT
                                    MONTH(TRY_CONVERT(DATE, m.CREATED_DATE, 103)) AS MonthNumber,
                                    SUM(CASE WHEN YEAR(TRY_CONVERT(DATE, m.CREATED_DATE, 103)) = YEAR(GETDATE()) - 3 THEN 1 ELSE 0 END) AS Previous_3_Year,
                                    SUM(CASE WHEN YEAR(TRY_CONVERT(DATE, m.CREATED_DATE, 103)) = YEAR(GETDATE()) - 2 THEN 1 ELSE 0 END) AS Previous_2_Year,
                                    SUM(CASE WHEN YEAR(TRY_CONVERT(DATE, m.CREATED_DATE, 103)) = YEAR(GETDATE()) - 1 THEN 1 ELSE 0 END) AS Previous_Year,
                                    SUM(CASE WHEN YEAR(TRY_CONVERT(DATE, m.CREATED_DATE, 103)) = YEAR(GETDATE()) THEN 1 ELSE 0 END) AS Current_Year
                                FROM TBL_MESSAGE m
                                JOIN TBL_PKG_DETAILS d
                                    ON m.PKG_ID = d.PKG_ID
                                WHERE
                                    m.MSG_TYPE = 'Offline Request'
                                    AND YEAR(TRY_CONVERT(DATE, m.CREATED_DATE, 103)) BETWEEN YEAR(GETDATE()) - 3 AND YEAR(GETDATE())
                                    AND d.country = '{selected_country}'
                                GROUP BY
                                    MONTH(TRY_CONVERT(DATE, m.CREATED_DATE, 103))
                            )
                            SELECT
                                DATENAME(MONTH, DATEFROMPARTS(2000, m.MonthNumber, 1)) AS Month,
                                COALESCE(md.Previous_3_Year, 0) AS Previous_3_Year,
                                COALESCE(md.Previous_2_Year, 0) AS Previous_2_Year,
                                COALESCE(md.Previous_Year, 0) AS Previous_Year,
                                COALESCE(md.Current_Year, 0) AS Current_Year
                            FROM Months m
                            LEFT JOIN MonthlyData md
                                ON m.MonthNumber = md.MonthNumber
                            ORDER BY
                                m.MonthNumber;
                            """


        elif msg_chart == "General_Query":
            with connection.cursor() as cursor:
                cursor.execute(f"""EXEC usp_Django_CountriesWithQueryCount_ForGeneralQuery
                            """)
                countries = [row[0] for row in cursor.fetchall()]


            if selected_country == "all":
                query = f"""EXEC usp_Django_MonthlyQueryCount_ForGeneralQuery_TillLast3years
                    """
                
            elif selected_country == "Others":
                query = """EXEC usp_Django_MonthlyCountofNullPKGID_ForGeneralQuery_TillLast3years
                        """

            elif selected_country != "Others" or selected_country != "all":
                query = f"""WITH Months AS (
                                SELECT 1 AS MonthNumber UNION ALL
                                SELECT 2 UNION ALL
                                SELECT 3 UNION ALL
                                SELECT 4 UNION ALL
                                SELECT 5 UNION ALL
                                SELECT 6 UNION ALL
                                SELECT 7 UNION ALL
                                SELECT 8 UNION ALL
                                SELECT 9 UNION ALL
                                SELECT 10 UNION ALL
                                SELECT 11 UNION ALL
                                SELECT 12
                            ),
                            MonthlyData AS (
                                SELECT
                                    MONTH(TRY_CONVERT(DATE, m.CREATED_DATE, 103)) AS MonthNumber,
                                    SUM(CASE WHEN YEAR(TRY_CONVERT(DATE, m.CREATED_DATE, 103)) = YEAR(GETDATE()) - 3 THEN 1 ELSE 0 END) AS Previous_3_Year,
                                    SUM(CASE WHEN YEAR(TRY_CONVERT(DATE, m.CREATED_DATE, 103)) = YEAR(GETDATE()) - 2 THEN 1 ELSE 0 END) AS Previous_2_Year,
                                    SUM(CASE WHEN YEAR(TRY_CONVERT(DATE, m.CREATED_DATE, 103)) = YEAR(GETDATE()) - 1 THEN 1 ELSE 0 END) AS Previous_Year,
                                    SUM(CASE WHEN YEAR(TRY_CONVERT(DATE, m.CREATED_DATE, 103)) = YEAR(GETDATE()) THEN 1 ELSE 0 END) AS Current_Year
                                FROM TBL_MESSAGE m
                                JOIN TBL_PKG_DETAILS d
                                    ON m.PKG_ID = d.PKG_ID
                                WHERE
                                    m.MSG_TYPE NOT IN ('holidays', 'package', 'CustomizedTrip', 'Offline Request')
                                    AND YEAR(TRY_CONVERT(DATE, m.CREATED_DATE, 103)) BETWEEN YEAR(GETDATE()) - 3 AND YEAR(GETDATE())
                                    AND d.country = '{selected_country}'
                                GROUP BY
                                    MONTH(TRY_CONVERT(DATE, m.CREATED_DATE, 103))
                            )
                            SELECT
                                DATENAME(MONTH, DATEFROMPARTS(2000, m.MonthNumber, 1)) AS Month,
                                COALESCE(md.Previous_3_Year, 0) AS Previous_3_Year,
                                COALESCE(md.Previous_2_Year, 0) AS Previous_2_Year,
                                COALESCE(md.Previous_Year, 0) AS Previous_Year,
                                COALESCE(md.Current_Year, 0) AS Current_Year
                            FROM Months m
                            LEFT JOIN MonthlyData md
                                ON m.MonthNumber = md.MonthNumber
                            ORDER BY
                                m.MonthNumber;
                        """


        elif msg_chart == "Total_Query":
            query = """EXEC usp_Django_Total_No_Of_Queries
                    """

        if query:
            with connection.cursor() as cursor:
                cursor.execute(query)
                data_to_show = cursor.fetchall()


        if data_to_show:
            df = pd.DataFrame(data_to_show, columns=['Month', (current_year-3), (current_year-2), (current_year-1), (current_year)])

            x = df["Month"]
            y_current_year_3 = df[current_year-3].tolist()
            y_current_year_2 = df[current_year-2].tolist()
            y_current_year_1 = df[current_year-1].tolist()
            y_current_year = df[current_year].tolist()

            
            trace1 = go.Bar(
                x=x, 
                y=y_current_year_3, 
                marker=dict(color='#0099ff'), 
                name=str(current_year-3),
                text=[f"{val:.0f}" for val in y_current_year_3], 
                # textposition='outside',
                textfont=dict(size=12, color='black'),
                hovertemplate="Year: " + str(current_year-3) + "<br>Month: %{x} <br> Queries: %{y:.0f}",
                )
            
            trace2 = go.Bar(
                x=x, 
                y=y_current_year_2, 
                marker=dict(color='#E257A4'), 
                name=str(current_year-2),
                text=[f"{val:.0f}" for val in y_current_year_2], 
                # textposition='outside',
                textfont=dict(size=12),
                hovertemplate="Year: " + str(current_year-2) + "<br>Month: %{x} <br> Queries: %{y:.0f}",
                )
            

            trace3 = go.Bar(
                x=x, 
                y=y_current_year_1, 
                marker=dict(color='#ff9900'), 
                name=str(current_year-1),
                text=[f"{val:.0f}" for val in y_current_year_1], 
                textfont=dict(size=12),
                hovertemplate="Year: " + str(current_year-1) + "<br>Month: %{x} <br> Queries: %{y:.0f}",
                )
            
            trace4 = go.Bar(
                x=x, 
                y=y_current_year, 
                marker=dict(color='#28a745'), 
                name=str(current_year),
                text=[f"{val:.0f}" for val in y_current_year], 
                textfont=dict(size=12),
                hovertemplate="Year: " + str(current_year) + "<br>Month: %{x} <br> Queries: %{y:.0f}",
                )

            fig = go.Figure(data=[trace1, trace2, trace3, trace4])

            fig.update_layout(
                title=dict(
                        text="<b>Monthly Queries</b>",
                        font=dict(size=18, family="Times New Roman, Times, serif", color="black", weight="bold")
                        ),
                
                xaxis=dict(
                        title="<b>Months</b>",
                        type='category',  # Treat x-axis as categorical so all dates appear
                        tickangle=0,
                        showgrid=False,
                        zeroline=False,
                        tickfont=dict(size=12, color='black')
                    ),

                yaxis=dict(
                        title="<b>No. of Query</b>",
                        ),

                barmode='group',  # Grouped bar chart
                margin=dict(t=25, l=25, r=25, b=25),
                height=600,  # Set height
                autosize=True,   # Set width
            )

            chart_json = fig.to_json()

        return render(request, 'query_report.html', {
            'msg_chart' : msg_chart,
            "chart_json": chart_json,
            'countries' : countries,
            'selected_country' : selected_country,
        })
        

def get_usd_to_inr():
    """Fetch and return the USD to INR exchange rate, stored for one day."""
    API_KEY = "7866c4e21be4d3a804cab3ea03634f30"
    URL = f"http://data.fixer.io/api/latest?access_key={API_KEY}"

    # Static variables inside the function to store values across calls
    if not hasattr(get_usd_to_inr, "last_updated_price") or get_usd_to_inr.last_updated_price != datetime.today().date():
        response = requests.get(URL).json()
        if response.get("success"):
            get_usd_to_inr.rate = round(response["rates"]["INR"] / response["rates"]["USD"], 2)
            get_usd_to_inr.last_updated_price = datetime.today().date()

    return get_usd_to_inr.rate


SECRET_KEY = config('SECRET_KEY')


def get_client_ip(request):
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip


def decrypt_params(encrypted_data):
    try:
        # Decode from base64 and then decode as UTF-16 little-endian
        json_string = base64.b64decode(encrypted_data).decode("utf-16le")
        return json.loads(json_string)
    except Exception as e:
        print("Decryption Error:", str(e))
        return None
    

def upload_file_to_s3(uploaded_file, bucket_name, s3_key):
    region_name = config('AWS_REGION') 
    print(region_name)
    
    s3 = boto3.client(
        's3',
        aws_access_key_id=config('AWS_ACCESS_KEY_ID'),
        aws_secret_access_key=config('AWS_SECRET_ACCESS_KEY'),
        region_name=region_name
    )

    print(s3)

    uploaded_file.seek(0)
    # Upload the fileobj directly to S3
    s3.upload_fileobj(
        Fileobj=uploaded_file,
        Bucket=bucket_name,
        Key=s3_key
       # ExtraArgs={'ACL': 'public-read'}  # Optional: make the file public
    )


    
    # Return the S3 file URL or key
    return f"https://{bucket_name}.s3.{region_name}.amazonaws.com/{s3_key}"

# ewAiAHMAZQBsAGUAYwB0AGUAZABfAHAAawBnAF8AaQBkACIAOgA4ACwAIgBzAGUAbABlAGMAdABlAGQAXwBhAGcAZQBuAHQASQBEACIAOgAiAEMASABBAEcAVAAwADAAMAAwADAAMwA2ADkAMgAiACwAIgBzAGUAbABlAGMAdABlAGQAXwBkAGEAdABlACIAOgAiADIANgAvADAANAAvADIAMAAyADUAIgB9AA==




class TravInfo:

    def null_if_blank(value):
        return value if value.strip() else None


    @csrf_exempt
    def gender_guess_view(request):
        if request.method == 'POST':
            try:
                data = json.loads(request.body)
                name = data.get('first_name', '').strip()
                
                first_name = name.split()[0].capitalize() if name else ''

                d = gender.Detector()
                gender_type = d.get_gender(first_name)

                if gender_type in ['male', 'mostly_male']:
                    return JsonResponse({'gender': 'Male'})
                elif gender_type in ['female', 'mostly_female']:
                    return JsonResponse({'gender': 'Female'})
                elif gender_type == 'andy':
                    return JsonResponse({'gender': ''})
                else:
                    return JsonResponse({'gender': ''})
            except Exception as e:
                return JsonResponse({'error': str(e)}, status=400)

        return JsonResponse({'error': 'Invalid request'}, status=400)




    

    


    @csrf_exempt
    def Trav_details(request):

        selected_pkg_id = ''
        selected_date = ''
        selected_agentID = ''

        encrypted_data = request.GET.get("data")

        if encrypted_data:
            params = decrypt_params(encrypted_data)
            if params:
                selected_pkg_id = params.get("selected_pkg_id")
                selected_agentID = params.get("selected_agentID")
                selected_date = params.get("selected_date")

            else:
                return HttpResponse("Invalid Data", status=400)

        executed_query = ''
        error_message= ''
        year_list = []
        data_to_show = []
        #data_headers = []
        raw_data=[]
        agent_details = []
        uploaded_file = ""
        flight_list= []
        passport_list= []
        show_message = ''
        selected_pkg_title = ''
        formatted_date= ''
        extra_details= ''
        date_to_show = ''
        no_data_message= ''
        results = []
        upload_success = False
        flight_traveller_name = []
        file_uploaded_type = ''
        no_details_found = ''
        guest_flight_data = []
        guest_flight_details = []
        # number_of_guest_left= ''  
        data_count_show = []  

        current_year = datetime.today().year
        for i in range(3):
            year_list.append(current_year+i)
            
        selected_year = request.GET.get('selected_year') 
        changed_data = json.loads(request.POST.get('changed_data', '[]'))
        selected_trav_id = request.POST.getlist("trav_id[]")
        remarks = request.POST.get("remarks",'')
        user_ip = get_client_ip(request)
        get_details = request.GET.get('get_details','guest_details')

        optional_tour = request.POST.get("optional_tour")
        extension_tour = request.POST.get("extension_tour")
        kids_info = request.POST.get("kids_info")
        special_request = request.POST.get("special_request")

        arrival_flight = request.POST.get("arrival_flight")
        arrival_date = request.POST.get("arrival_date")
        arrival_time = request.POST.get("arrival_time")
        arrival_airport = request.POST.get("arrival_airport")

        departure_flight = request.POST.get("departure_flight")
        departure_date = request.POST.get("departure_date")
        departure_time = request.POST.get("departure_time")
        departure_airport = request.POST.get("departure_airport")


        arrival_flight_list = request.POST.getlist("arrival_flight[]")
        arrival_airport_list = request.POST.getlist("arrival_airport[]")
        arrival_date_list = request.POST.getlist("arrival_date[]")
        arrival_time_list = request.POST.getlist("arrival_time[]") 
        departure_flight_list = request.POST.getlist("departure_flight[]")
        departure_airport_list = request.POST.getlist("departure_airport[]")
        departure_date_list = request.POST.getlist("departure_date[]")
        departure_time_list = request.POST.getlist("departure_time[]")
        guest_trav_id = request.POST.get('trav_id')
        manual_edit = request.POST.get('manual_edit','')

        print(guest_trav_id)
        print(manual_edit)


        submit_type = request.POST.get("submit_type")

        if selected_pkg_id:
            request.session['selected_pkg_id'] = selected_pkg_id

        if selected_year:
            request.session['selected_year'] = selected_year

        if selected_date:
            request.session['selected_date'] = selected_date

        if selected_agentID:
            request.session['selected_agentID'] = selected_agentID

        selected_pkg_id = request.session.get('selected_pkg_id')
        selected_year = request.session.get('selected_year')
        selected_date = request.session.get('selected_date')
        selected_agentID = request.session.get('selected_agentID')

        if selected_date:
            formatted_date = datetime.strptime(selected_date, "%d/%m/%Y").strftime("%Y-%m-%d")
            date_to_show = datetime.strptime(selected_date, "%d/%m/%Y").strftime("%b %d, %Y")

        print(f"Package ID: {selected_pkg_id}, Year: {selected_year}, Date: {selected_date}, Agent ID: {selected_agentID}, Formatted Date: {formatted_date}" )  # Debugging  

        if selected_pkg_id:
            with connection.cursor() as cursor:
                cursor.execute(f"""SELECT PKG_TITLE
                                from TBL_PKG_DETAILS
                                where pkg_id = {selected_pkg_id}
                                """)
                selected_pkg_title = cursor.fetchone()[0]
                
        if selected_agentID and selected_pkg_id and formatted_date:
            with connection.cursor() as cursor: 
                cursor.execute(f"""select Name, Comp_Name, trim(lower(Emailid)) as Emailid, Contact
                                    from TBL_Agent 
                                    where AgentID = '{selected_agentID}'
                            """)
                agent_details = cursor.fetchall()

 
                cursor.execute(f"""SELECT trim(upper(T_FNAME)) as T_FNAME, trim(upper(T_LNAME)) as T_LNAME, TRAV_ID
                                FROM TBL_TRAVELLER_NAME_DRAFT
                                WHERE Status = 'Active'
                                    AND PKG_ID = {selected_pkg_id}
                                    AND Convert(date,TourDate,103) = '{formatted_date}'
                                    AND AgentID = '{selected_agentID}'
                                    -- and TRAV_ID not in (select TRAV_ID from TBL_TRAVELLER_PASSPORT_DETAILS)
                                    order by T_FNAME
                                """)
                data_to_show = cursor.fetchall()

                if len(data_to_show) == 0:
                    no_details_found = "No"


            
            ## Data Count for showing Information related to Flight and Passport
            

            with connection.cursor() as cursor: 
                # Total Number of Guests
                cursor.execute(f"SELECT COUNT(*) FROM TBL_TRAVELLER_NAME_DRAFT WHERE PKG_ID = {selected_pkg_id} AND Tourdate = '{formatted_date}' AND AgentID = '{selected_agentID}'")
                check_draft_data = cursor.fetchone()[0]
                
                if check_draft_data > 0:
                    cursor.execute(f"""SELECT count(*)
                                    FROM TBL_TRAVELLER_NAME_DRAFT
                                    WHERE Status = 'Active'
                                        AND PKG_ID = {selected_pkg_id}
                                        AND Tourdate = '{formatted_date}'
                                        AND AgentID = '{selected_agentID}'
                                    """)
                    total_guest = cursor.fetchone()[0]


                    cursor.execute(f"""SELECT count(*)
                        FROM TBL_TRAVELLER_NAME_DRAFT
                        WHERE Status = 'Active'
                            AND PKG_ID = {selected_pkg_id}
                            AND Tourdate = '{formatted_date}'
                            AND AgentID = '{selected_agentID}'
                            AND TRAV_ID NOT in (Select TRAV_ID from TBL_GUEST_FLIGHT_DETAILS)
                               """)
                    
                    flight_ticket_left = cursor.fetchone()[0]

                    


                else:
                    cursor.execute(f"""SELECT count(*)
                                    FROM TBL_TRAVELLER_NAME 
                                    WHERE Status1 = 'Active'
                                        AND PaxDepositAmount > 0
                                        AND Pkgid = {selected_pkg_id}
                                        AND try_convert(date,TourDate,103) = '{formatted_date}'
                                        AND CREATED_BY = '{selected_agentID}'
                                    """)
                    total_guest = cursor.fetchone()[0]
                    flight_ticket_left = total_guest


                print("Total Guest : ", total_guest)
                data_count_show.append(total_guest)


                print("flight_ticket_left : ", flight_ticket_left)
                data_count_show.append(flight_ticket_left)




                # Number of Guest whose Flight Titcket is not Uploaded
                # cursor.execute(f"""SELECT count(*)
                #         FROM TBL_TRAVELLER_NAME_DRAFT
                #         WHERE Status = 'Active'
                #             AND PKG_ID = {selected_pkg_id}
                #             AND Tourdate = '{formatted_date}'
                #             AND AgentID = '{selected_agentID}'
                #             AND TRAV_ID NOT in (Select TRAV_ID from TBL_GUEST_FLIGHT_DETAILS)
                #                """)
                # flight_ticket_left = cursor.fetchone()[0]

                # print("flight_ticket_left : ", flight_ticket_left)
                # data_count_show.append(flight_ticket_left)

                ## Number of Guest whose Flight Ticket Uploaded

                # flight_ticket_upload = total_guest - flight_ticket_left

                # print("flight_ticket_upload : ", flight_ticket_upload)
                # data_count_show.append(flight_ticket_upload)

               


                ## Arrivial Details left
                # arrivial_details_missing = 0
                # departure_details_missing = 0

                # cursor.execute(f"""
                #             select TRAV_ID, Arrival_Flight_Details, Arrival_Airport, Arrival_Date, Arrival_Time, Departure_Flight_Details, Departure_Airport, Departure_Date, Departure_Time
                #             from TBL_GUEST_FLIGHT_DETAILS
                #             where status = 1
                #                 AND PKG_ID = {selected_pkg_id}
                #                 AND try_convert(date,TourDate,103) = '{formatted_date}'
                #                 AND Agent_ID = '{selected_agentID}'
                #             """)

                # guest_flight_data = cursor.fetchall()

                # merged = []
                # for i in guest_flight_data:
                #     trav_id = i[0]

                    ## --- Arrival check ---
                    # if (i[1] not in (None, '', 'NULL') and
                    #     i[2] not in (None, '', 'NULL') and
                    #     i[3] not in (None, '', 'NULL') and
                    #     i[4] not in (None, '', 'NULL')):
                    #     arrival_status = "Yes"
                    # else:
                    #     arrival_status = "No"

                    ## --- Departure check ---
                    # if (i[5] not in (None, '', 'NULL') and
                    #     i[6] not in (None, '', 'NULL') and
                    #     i[7] not in (None, '', 'NULL') and
                    #     i[8] not in (None, '', 'NULL')):
                    #     departure_status = "Yes"
                    # else:
                    #     departure_status = "No"

                    ## Append row result into list
                    # merged.append([trav_id, arrival_status, departure_status])

                # for i in merged:
                #     if i[1] == 'No' and i[2] == 'No':
                #         arrivial_details_missing = arrivial_details_missing + 1
                #         departure_details_missing = departure_details_missing + 1

                #     elif i[1] == 'No' and i[2] == 'Yes':
                #         arrivial_details_missing = arrivial_details_missing + 1

                #     elif i[1] == 'Yes' and i[2] == 'No':
                #         departure_details_missing = departure_details_missing + 1

                # print(arrivial_details_missing)
                # print(departure_details_missing)
                

                # data_count_show.append(arrivial_details_missing)
                # data_count_show.append(departure_details_missing)



                ## Passport Details Not Submitted
                cursor.execute(f"""select count(*) 
                                from TBL_TRAVELLER_NAME_DRAFT
                               WHERE Status = 'Active'
                                    AND PKG_ID = {selected_pkg_id}
                                    AND Tourdate = '{formatted_date}'
                                    AND AgentID = '{selected_agentID}'
                                    AND TRAV_ID NOT IN (select TRAV_ID from TBL_TRAVELLER_PASSPORT_DETAILS)
                               """)
                passport_left = cursor.fetchone()[0]

                print("Passport Left : ", passport_left)

                if check_draft_data == 0:
                    passport_left = total_guest

                data_count_show.append(passport_left)

                
                
                    


                ## Passport Details Submit
                # passport_submit = total_guest - passport_left

                # print("passport_submit : ", passport_submit)
                # data_count_show.append(passport_submit)


                print(data_count_show)



                    

        

                
                
        if get_details == "guest_details":
            try:
                if selected_agentID and selected_pkg_id and formatted_date:
                    check_draft_data = ''
                    with connection.cursor() as cursor:
                    
                        cursor.execute(f"SELECT COUNT(*) FROM TBL_TRAVELLER_NAME_DRAFT WHERE PKG_ID = {selected_pkg_id} AND Tourdate = '{formatted_date}' AND AgentID = '{selected_agentID}'")
                        check_draft_data = cursor.fetchone()[0]
                        
                        if check_draft_data > 0:
                            cursor.execute(f"""SELECT RoomNo, T_FNAME, T_LNAME, TRAV_ID, GENDER, RoomType, RoomPref, Noofprenight, noopostnight, Status
                                            FROM TBL_TRAVELLER_NAME_DRAFT
                                            WHERE PKG_ID = {selected_pkg_id}
                                                AND Tourdate = '{formatted_date}'
                                                AND AgentID = '{selected_agentID}'
                                                ORDER BY RoomNo
                                           """)
                            raw_data = cursor.fetchall()

                        else:
                            cursor.execute(f"""SELECT RoomNo, T_FNAME, T_LNAME, TRAV_ID, GENDER, RoomType, RoomPref, Noofprenight, noopostnight, Status1
                                            FROM TBL_TRAVELLER_NAME 
                                            WHERE Status1 = 'Active'
                                                AND PaxDepositAmount > 0
                                                AND Pkgid = {selected_pkg_id}
                                                AND try_convert(date,TourDate,103) = '{formatted_date}'
                                                AND CREATED_BY = '{selected_agentID}'
                                                ORDER BY RoomNo
                                            """)
                            raw_data = cursor.fetchall()

                data_to_show = []
                for row in raw_data:
                    data_to_show.append(row)

                field_to_column_map = {
                    'room_number': 'RoomNo',
                    'first_name': 'T_FNAME',
                    'last_name': 'T_LNAME',
                    'trav_id': 'TRAV_ID',
                    'gender': 'GENDER',
                    'room_type': 'RoomType',
                    'bed_type': 'RoomPref',
                    'pre_nights': 'Noofprenight',
                    'post_nights': 'noopostnight',
                    'guest_travelling': 'Status',
                }


                extra_fields = {
                    'AgentID': selected_agentID,
                    'Tourdate': formatted_date,
                    'PKG_ID': selected_pkg_id,
                }
                
                if changed_data:
                    try:
                        with connection.cursor() as cursor:
                            for row in changed_data:
                                trav_id = row.get('trav_id')
                                column_names = []
                                parameters = []

                                for key, value in row.items():
                                    column_name = field_to_column_map.get(key)
                                    if column_name:
                                        column_names.append(column_name)
                                        parameters.append(value)
 
                                for col, val in extra_fields.items():
                                    column_names.append(col)
                                    parameters.append(val)

                                if column_names:
                                    if check_draft_data > 0:
                                        update_query = f"""
                                            UPDATE TBL_TRAVELLER_NAME_DRAFT
                                            SET {', '.join([f"{col} = %s" for col in column_names])}
                                            ,Updated_Date = CURRENT_TIMESTAMP
                                            WHERE TRAV_ID = %s
                                            """
                                        parameters.append(trav_id)  # Add TRAV_ID to parameters for the WHERE clause    
                                        cursor.execute(update_query, parameters)

                                    else:
                                        placeholders = ', '.join(['%s'] * len(parameters))
                                        insert_query = f"""
                                            INSERT INTO TBL_TRAVELLER_NAME_DRAFT
                                            ({', '.join(column_names)})
                                            VALUES ({placeholders})
                                        """
                                        print(insert_query)
                                        print(parameters)

                                        cursor.execute(insert_query, parameters)
                                        
                                executed_query = connection.queries[-1]['sql'] if connection.queries else "No query executed"

                        connection.commit()
                        messages.success(request, "Guest Details Submitted Successfully!")
                        return redirect(f'/trav-info?get_details={get_details}')

                    except DatabaseError as e:
                        error_message = f"Database Error during update: {str(e)}"
                        connection.rollback()
                    except Exception as e:
                        error_message = f"Unexpected Error during update: {str(e)}"
                        connection.rollback()

                executed_query = connection.queries[-1]['sql'] if connection.queries else "No query executed"
            
            except DatabaseError as e:  
                error_message = f"Database Error: {str(e)}"
                return render(request, 'no_database.html', {})

            except Exception as e:
                error_message = f"Error: {str(e)}"
                return render(request, 'no_database.html', {})
            

        elif get_details == "flight_details":
            try:
                if selected_agentID and selected_pkg_id and formatted_date:
                    with connection.cursor() as cursor: 
                        cursor.execute(f"""SELECT trim(upper(T_FNAME)) as T_FNAME, trim(upper(T_LNAME)) as T_LNAME, TRAV_ID
                                        FROM TBL_TRAVELLER_NAME_DRAFT
                                        WHERE Status = 'Active'
                                            AND PKG_ID = {selected_pkg_id}
                                            AND try_convert(date,TourDate,103) = '{formatted_date}'
                                            AND AgentID = '{selected_agentID}'
                                            -- and TRAV_ID not in (select TRAV_ID from TBL_GUEST_FLIGHT_DETAILS)
                                            order by T_FNAME
                                        """)
                        data_to_show = cursor.fetchall()
                        print("Data for Showing :", data_to_show)



                        # cursor.execute(f"""select count(*) from TBL_GUEST_FLIGHT_DETAILS
                        #                     where Status = 1
                        #                         AND PKG_ID = {selected_pkg_id}
                        #                         AND try_convert(date,TourDate,103) = '{formatted_date}'
                        #                         AND Agent_ID = '{selected_agentID}'
                        #                """)
                        # total_count = cursor.fetchone()
                        # count_of_guest = total_count[0] if total_count else 0  
                        # print("Total Count : ", count_of_guest)

                        # number_of_guest_left = len(data_to_show) - count_of_guest
                        # print(number_of_guest_left)
                        

                         



                        cursor.execute(f"""
                            select TRAV_ID, Arrival_Flight_Details, Arrival_Airport, Arrival_Date, Arrival_Time, Departure_Flight_Details, Departure_Airport, Departure_Date, Departure_Time
                            from TBL_GUEST_FLIGHT_DETAILS
                            where status = 1
                                AND PKG_ID = {selected_pkg_id}
                                AND try_convert(date,TourDate,103) = '{formatted_date}'
                                AND Agent_ID = '{selected_agentID}'
                                       
                        """)
                        guest_flight_data = cursor.fetchall()

                        # print(guest_flight_data)

                merged = []
                for i in guest_flight_data:
                    trav_id = i[0]

                    # --- Arrival check ---
                    if (i[1] not in (None, '', 'NULL') and
                        i[2] not in (None, '', 'NULL') and
                        i[3] not in (None, '', 'NULL') and
                        i[4] not in (None, '', 'NULL')):
                        arrival_status = "Yes"
                    else:
                        arrival_status = "No"

                    # --- Departure check ---
                    if (i[5] not in (None, '', 'NULL') and
                        i[6] not in (None, '', 'NULL') and
                        i[7] not in (None, '', 'NULL') and
                        i[8] not in (None, '', 'NULL')):
                        departure_status = "Yes"
                    else:
                        departure_status = "No"

                    # Append row result into list
                    merged.append([trav_id, arrival_status, departure_status])

                # print(merged)

                lookup = {b[0]: b[1:] for b in merged}
                for first, last, trav_id in data_to_show:
                    if trav_id in lookup:
                        guest_flight_details.append((trav_id, first, last, *lookup[trav_id]))
                    else:
                        # if no match, fill with None (or 'N/A')
                        guest_flight_details.append((trav_id, first, last, 'No', 'No'))

                # print(guest_flight_details)

                


                if len(data_to_show) == 0:
                    with connection.cursor() as cursor: 
                        cursor.execute(f"""Select count(*) as Total from TBL_GUEST_FLIGHT_DETAILS
                                    WHERE PKG_ID = {selected_pkg_id}
                                        AND try_convert(date,TourDate,103) = '{formatted_date}'
                                        AND Agent_ID = '{selected_agentID}'                                               
                                """)
                        flight_details_guests = cursor.fetchone()
                    total_guests = flight_details_guests[0]

                    if total_guests > 0:
                        no_data_message = f'You have filled in the data of all guests.'
                    else:
                        no_data_message = f'No data found. Please fill the <a href="/trav-info?get_details=guest_details">Guest Details</a> first.'


                uploaded_file = request.FILES.get("file")

                if uploaded_file:
                    upload_success = True
                    file_mime_type = uploaded_file.content_type

                    files = {"file": (uploaded_file.name, uploaded_file.read(), file_mime_type)}

                    print("File Type is ", file_mime_type)
                    
                    if  'image' in file_mime_type or 'pdf' in file_mime_type:
                        file_uploaded_type = 'image'

                        # Send the POST request with form-data
                        url = "https://prodzodiac.cultureholidays.com/api/flight/flight-details"
                        
                        response = requests.post(url, files=files)

                        if response.status_code == 200:
                            
                            flight_details = response.json()

                            extracted = flight_details.get("extractedDetails", {})

                            # Extract traveller names
                            flight_traveller_name = [
                                traveller["name"].upper()
                                for traveller in extracted.get("travellers", [])
                            ]

                            print(extracted)

                            # Extract flight details
                            flight_list = [
                                (
                                    flight["departure_flight_details"],  # 0
                                    flight["departure_city"],            # 1
                                    flight["departure_date"],            # 2
                                    flight["departure_time"],            # 3
                                    flight["arrival_flight_details"],    # 4
                                    flight["arrival_city"],              # 5
                                    flight["arrival_date"],              # 6
                                    flight["arrival_time"],              # 7
                                    flight["departure_airport"],         # 8
                                    flight["arrival_airport"]            # 9
                                )
                                for flight in extracted.get("flight", [])
                            ]

                            print("Traveler Names:", flight_traveller_name)
                            print("Flight List:", flight_list)
                            
                            # print("File Type : ", file_uploaded_type)

                            if not flight_list:
                                error_message = f"Error: No Data Found in the Ticket ❌. Please Re-Upload the Ticket or Try a Different File Format or Fill the Form manually."
                                print(error_message)



                    elif 'spreadsheetml.sheet' in file_mime_type:
                    
                        file_uploaded_type = 'sheet'
                        url = "https://prodzodiac.cultureholidays.com/api/flight/flight-excel"

                        print("Excel Sheet Data")
                        response = requests.post(url, files=files)
                        if response.status_code == 200:
                            flight_details = response.json()

                            flight_traveller_name = [
                                traveller["name"].upper()
                                for traveller in flight_details['data']
                            ]

                            flight_list = [
                                (   
                                    flight['name'],                   #0  
                                    flight["arrival_flight"],         #1  
                                    flight["arrival_date"],           #2 
                                    flight["arrival_time"],           #3  
                                    flight["departure_flight"],       #4 
                                    flight["departure_date"],         #5        
                                    flight["departure_time"],         #6 
                                    flight["arrival_airport"],        #7 
                                    flight["departure_airport"],      #8              
                                )
                                for flight in flight_details.get("data")
                            ]
        
                            print("Traveler Names:", flight_traveller_name)
                            print("Flight List:", flight_list)

                            print("Data_to_show", data_to_show)

                            full_name_to_id = {f"{t[0]} {t[1]}".upper(): t[2] for t in data_to_show}

                            flight_list = [
                                row + (
                                    "Yes" if row[0].upper() in full_name_to_id else "No",
                                    full_name_to_id.get(row[0].upper(), "")
                                )
                                for row in flight_list
                            ]

                            # Filter only rows where status is 'Yes'
                            flight_list = [row for row in flight_list if row[9] == 'Yes']

                            # Debug
                            # print("Final Flight List : ", flight_list)





                if manual_edit == 'manual_edit':
                    print("Manual Edit Mode Activated")
                    print(guest_trav_id)

                    file_uploaded_type = 'manual'

                    with connection.cursor() as cursor:
                        cursor.execute(f"""
                                SELECT 
                                    ISNULL(Arrival_Flight_Details, '') AS Arrival_Flight_Details,
                                    ISNULL(Arrival_Airport, '') AS Arrival_Airport,
                                    CONVERT(varchar(10), Arrival_Date, 23) AS Arrival_Date,   -- dd/MM/yyyy
                                    CONVERT(varchar(5), Arrival_Time, 108) AS Arrival_Time,    -- HH:mm

                                    ISNULL(Departure_Flight_Details, '') AS Departure_Flight_Details,
                                    ISNULL(Departure_Airport, '') AS Departure_Airport,
                                    CONVERT(varchar(10), Departure_Date, 23) AS Departure_Date,
                                    CONVERT(varchar(5), Departure_Time, 108) AS Departure_Time
                                FROM TBL_GUEST_FLIGHT_DETAILS
                                WHERE TRAV_ID = {guest_trav_id}
                        """)

                        flight_list = cursor.fetchall()

                        print("Flight List:", flight_list)

                        cursor.execute(f"""select concat(trim(upper(T_FNAME)),' ', trim(upper(T_LNAME))) as Name  from TBL_TRAVELLER_NAME_DRAFT
                                    where TRAV_ID = {guest_trav_id}
                                    """)
                        
                        flight_traveller_name = cursor.fetchone()
                        print("Flight Traveller Names:", flight_traveller_name)




                if submit_type == 'image':
                    print("Submitting Data in SQL")
                    if selected_trav_id: 
                        formatted_date = datetime.strptime(selected_date, "%d/%m/%Y").strftime("%Y-%m-%d")
                        for trav_id in selected_trav_id:
                            
                            with connection.cursor() as cursor:
                                sql_query = """
                                    MERGE INTO TBL_GUEST_FLIGHT_DETAILS AS target
                                        USING (SELECT %s AS Agent_ID, %s AS TourDate, %s AS PKG_ID, %s AS TRAV_ID) AS source
                                        ON target.Agent_ID = source.Agent_ID
                                        AND target.TourDate = source.TourDate
                                        AND target.PKG_ID = source.PKG_ID
                                        AND target.TRAV_ID = source.TRAV_ID
                                        WHEN MATCHED THEN
                                            UPDATE SET
                                                target.Arrival_Flight_Details = %s,
                                                target.Arrival_Airport = %s,
                                                target.Arrival_Date = %s,
                                                target.Arrival_Time = %s,
                                                target.Departure_Flight_Details = %s,
                                                target.Departure_Airport = %s,
                                                target.Departure_Date = %s,
                                                target.Departure_Time = %s,
                                                target.Remarks = %s,
                                                target.Updated_Date = CURRENT_TIMESTAMP,
                                                target.Updated_By = %s,
                                                target.IpAddress = %s,
                                                target.status = 1
                                        WHEN NOT MATCHED THEN
                                            INSERT (
                                                TRAV_ID, PKG_ID, TourDate, Agent_ID, Arrival_Flight_Details, Arrival_Airport, Arrival_Date, Arrival_Time,
                                                Departure_Flight_Details, Departure_Airport, Departure_Date, Departure_Time, Remarks, Created_By, IpAddress, Created_Date
                                            )
                                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, CURRENT_TIMESTAMP);
                                        """
                                
                                params = [
                                        selected_agentID, 
                                        formatted_date, 
                                        selected_pkg_id, 
                                        trav_id,
                                        arrival_flight,
                                        arrival_airport,
                                        TravInfo.null_if_blank(arrival_date),
                                        TravInfo.null_if_blank(arrival_time),
                                        departure_flight,
                                        departure_airport,
                                        TravInfo.null_if_blank(departure_date),
                                        TravInfo.null_if_blank(departure_time),
                                        remarks, 
                                        selected_agentID, 
                                        user_ip,
                                        trav_id, 
                                        selected_pkg_id, 
                                        formatted_date, 
                                        selected_agentID,
                                        arrival_flight, 
                                        arrival_airport,
                                        TravInfo.null_if_blank(arrival_date),
                                        TravInfo.null_if_blank(arrival_time),
                                        departure_flight, 
                                        departure_airport,
                                        TravInfo.null_if_blank(departure_date),
                                        TravInfo.null_if_blank(departure_time),
                                        remarks, 
                                        selected_agentID, 
                                        user_ip
                                    ]
                                
                                cursor.execute(sql_query, params)
                                executed_query = connection.queries[-1]['sql'] if connection.queries else "No query executed"
                                print(executed_query)


                        messages.success(request, "Flight Details Submitted Successfully!")
                        return redirect(f'/trav-info?get_details={get_details}')
                    
                elif submit_type == 'sheet':
                    print("Inseting Sheet Data")
                    if selected_trav_id:
                        formatted_date = datetime.strptime(selected_date, "%d/%m/%Y").strftime("%Y-%m-%d")


                        with connection.cursor() as cursor:
                            for i in range(len(selected_trav_id)):
                                sql_query = """
                                    MERGE INTO TBL_GUEST_FLIGHT_DETAILS AS target
                                        USING (SELECT %s AS Agent_ID, %s AS TourDate, %s AS PKG_ID, %s AS TRAV_ID) AS source
                                        ON target.Agent_ID = source.Agent_ID
                                        AND target.TourDate = source.TourDate
                                        AND target.PKG_ID = source.PKG_ID
                                        AND target.TRAV_ID = source.TRAV_ID
                                        WHEN MATCHED THEN
                                            UPDATE SET
                                                target.Arrival_Flight_Details = %s,
                                                target.Arrival_Airport = %s,
                                                target.Arrival_Date = %s,
                                                target.Arrival_Time = %s,
                                                target.Departure_Flight_Details = %s,
                                                target.Departure_Airport = %s,
                                                target.Departure_Date = %s,
                                                target.Departure_Time = %s,
                                                target.Remarks = %s,
                                                target.Updated_Date = CURRENT_TIMESTAMP,
                                                target.Updated_By = %s,
                                                target.IpAddress = %s,
                                                target.status = 1
                                        WHEN NOT MATCHED THEN
                                            INSERT (
                                                TRAV_ID, PKG_ID, TourDate, Agent_ID, Arrival_Flight_Details, Arrival_Airport, Arrival_Date, Arrival_Time,
                                                Departure_Flight_Details, Departure_Airport, Departure_Date, Departure_Time, Remarks, Created_By, IpAddress, Created_Date
                                            )
                                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, CURRENT_TIMESTAMP);
                                        """
                                
                                params = [
                                        selected_agentID, 
                                        formatted_date, 
                                        selected_pkg_id, 
                                        selected_trav_id[i],
                                        arrival_flight_list[i], 
                                        arrival_airport_list[i],
                                        TravInfo.null_if_blank(arrival_date_list[i]),
                                        TravInfo.null_if_blank(arrival_time_list[i]),
                                        departure_flight_list[i], 
                                        departure_airport_list[i],
                                        TravInfo.null_if_blank(departure_date_list[i]),
                                        TravInfo.null_if_blank(departure_time_list[i]),
                                        remarks, 
                                        selected_agentID, 
                                        user_ip,
                                        selected_trav_id[i], 
                                        selected_pkg_id, 
                                        formatted_date, 
                                        selected_agentID,
                                        arrival_flight_list[i], 
                                        arrival_airport_list[i],
                                        TravInfo.null_if_blank(arrival_date_list[i]),
                                        TravInfo.null_if_blank(arrival_time_list[i]),
                                        departure_flight_list[i], 
                                        departure_airport_list[i],
                                        TravInfo.null_if_blank(departure_date_list[i]),
                                        TravInfo.null_if_blank(departure_time_list[i]),
                                        remarks, 
                                        selected_agentID, 
                                        user_ip
                                    ]
                                
                                print(params)
                                cursor.execute(sql_query, params)


                            messages.success(request, "Flight Details Submitted Successfully!")
                            return redirect(f'/trav-info?get_details={get_details}')

                # elif submit_type == 'manual':
                #     print("This is Manual Editing")

                #     if guest_trav_id:
                #         formatted_date = datetime.strptime(selected_date, "%d/%m/%Y").strftime("%Y-%m-%d")
                #         with connection.cursor() as cursor:
                #             sql_query = """
                #                 MERGE INTO TBL_GUEST_FLIGHT_DETAILS AS target
                #                     USING (SELECT %s AS Agent_ID, %s AS TourDate, %s AS PKG_ID, %s AS TRAV_ID) AS source
                #                     ON target.Agent_ID = source.Agent_ID
                #                     AND target.TourDate = source.TourDate
                #                     AND target.PKG_ID = source.PKG_ID
                #                     AND target.TRAV_ID = source.TRAV_ID
                #                     WHEN MATCHED THEN
                #                         UPDATE SET
                #                             target.Arrival_Flight_Details = %s,
                #                             target.Arrival_Airport = %s,
                #                             target.Arrival_Date = %s,
                #                             target.Arrival_Time = %s,
                #                             target.Departure_Flight_Details = %s,
                #                             target.Departure_Airport = %s,
                #                             target.Departure_Date = %s,
                #                             target.Departure_Time = %s,
                #                             target.Remarks = %s,
                #                             target.Updated_Date = CURRENT_TIMESTAMP,
                #                             target.Updated_By = %s,
                #                             target.IpAddress = %s
                #                     WHEN NOT MATCHED THEN
                #                         INSERT (
                #                             TRAV_ID, PKG_ID, TourDate, Agent_ID, Arrival_Flight_Details, Arrival_Airport, Arrival_Date, Arrival_Time,
                #                             Departure_Flight_Details, Departure_Airport, Departure_Date, Departure_Time, Remarks, Created_By, IpAddress, Created_Date
                #                         )
                #                         VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, CURRENT_TIMESTAMP);
                #                     """
                            
                #             params = [
                #                     selected_agentID, 
                #                     formatted_date, 
                #                     selected_pkg_id, 
                #                     guest_trav_id,
                #                     arrival_flight,
                #                     arrival_airport,
                #                     TravInfo.null_if_blank(arrival_date),
                #                     TravInfo.null_if_blank(arrival_time),
                #                     departure_flight,
                #                     departure_airport,
                #                     TravInfo.null_if_blank(departure_date),
                #                     TravInfo.null_if_blank(departure_time),
                #                     remarks, 
                #                     selected_agentID, 
                #                     user_ip,
                #                     guest_trav_id, 
                #                     selected_pkg_id, 
                #                     formatted_date, 
                #                     selected_agentID,
                #                     arrival_flight, 
                #                     arrival_airport,
                #                     TravInfo.null_if_blank(arrival_date),
                #                     TravInfo.null_if_blank(arrival_time),
                #                     departure_flight, 
                #                     departure_airport,
                #                     TravInfo.null_if_blank(departure_date),
                #                     TravInfo.null_if_blank(departure_time),
                #                     remarks, 
                #                     selected_agentID, 
                #                     user_ip
                #                 ]
                            
                #             cursor.execute(sql_query, params)
                #             executed_query = connection.queries[-1]['sql'] if connection.queries else "No query executed"
                #             print(executed_query)


                #         messages.success(request, "Flight Details Updated Successfully!")
                #         return redirect(f'/trav-info?get_details={get_details}')
                        



            except Exception as e:
                print(e)
                error_message = "Error: Something went wrong. Please try again later."
                


        elif get_details == "passport_details":
            try:
                if selected_agentID:
                    with connection.cursor() as cursor: 
                        cursor.execute(f"""SELECT trim(upper(T_FNAME)) as T_FNAME, trim(upper(T_LNAME)) as T_LNAME, TRAV_ID
                                        FROM TBL_TRAVELLER_NAME_DRAFT
                                        WHERE Status = 'Active'
                                            AND PKG_ID = {selected_pkg_id}
                                            AND Convert(date,TourDate,103) = '{formatted_date}'
                                            AND AgentID = '{selected_agentID}'
                                            -- and TRAV_ID not in (select TRAV_ID from TBL_TRAVELLER_PASSPORT_DETAILS)
                                            order by T_FNAME
                                        """)
                        data_to_show = cursor.fetchall()
                        print(data_to_show)

                        cursor.execute(f"""
                                select TRAV_ID from TBL_TRAVELLER_PASSPORT_DETAILS
                                where Agent_ID = '{selected_agentID}'
                                    AND PKG_ID = {selected_pkg_id}
                                    AND Convert(date,TourDate,103) = '{formatted_date}'
                        """)
                        guest_flight_data = [row[0] for row in cursor.fetchall()]
                        print(guest_flight_data) 

                        for first, last, trav_id in data_to_show:
                            if trav_id in guest_flight_data:
                                guest_flight_details.append((trav_id, first, last, "Yes"))
                            else:
                                guest_flight_details.append((trav_id, first, last, "No"))

                        print(guest_flight_details)
                    





                if len(data_to_show) == 0:
                
                    no_data_message = f'No data found. Please fill the <a href="/trav-info?get_details=guest_details">Guest Details</a> first.'

          
                url = "https://prodzodiac.cultureholidays.com/api/flight/passport-details"

                # uploaded_file = request.FILES.get("file")

                files = request.FILES.getlist('file')

                for uploaded_file in files:
                    file_result = {
                                "filename": uploaded_file.name,
                                "status": ""
                            }
                    
                    print(uploaded_file)
                    file_mime_type = uploaded_file.content_type

                    files = {"file": (uploaded_file.name, uploaded_file.read(), file_mime_type)}

                    data = {"tourdate": selected_date}

                    try: 

                        response = requests.post(url, files=files, data=data, timeout=60)

                        if response.status_code == 200:
                            passport_details = response.json()
                    

                            if "expiryMessage" in passport_details:
                                check_passport = passport_details["expiryMessage"]
                            

                                if check_passport == "your passport is valid ":
                                    if "extractedDetails" in passport_details:
                                        extracted_details = passport_details["extractedDetails"]

                                        # Store passport details as a list of tuples (similar to flight_list)
                                        passport_list = [
                                            (
                                                extracted_details.get("passport_type", ""),       # 0
                                                extracted_details.get("country_code", ""),       # 1
                                                extracted_details.get("passport_number", ""),    # 2
                                                extracted_details.get("surname", ""),            # 3
                                                extracted_details.get("given_names", ""),        # 4
                                                extracted_details.get("nationality", ""),        # 5
                                                extracted_details.get("date_of_birth", ""),      # 6
                                                extracted_details.get("place_of_birth", ""),     # 7
                                                extracted_details.get("sex", ""),                # 8
                                                extracted_details.get("date_of_issue", ""),      # 9
                                                extracted_details.get("date_of_expiry", ""),     # 10
                                                extracted_details.get("issuing_authority", ""),  # 11
                                            )
                                        ]

                                else:
                                    if passport_details["extractedDetails"].get("date_of_expiry", ""):
                                        expiry_date_str = passport_details["extractedDetails"].get("date_of_expiry", "")
                                        expiry_date = datetime.strptime(expiry_date_str, "%d %b %Y").date()
                                        formatted_date = datetime.strptime(selected_date, "%d/%m/%Y").date()
                                        days_difference = (expiry_date - formatted_date).days


                                        if 0 < days_difference <= 180:
                                            #show_message = "Passport will expires within 180 days!"
                                            file_result['status'] = "Passport will expires within 180 days!"
                                        elif days_difference <= 0:
                                            #show_message = "Your Passport is Expired."
                                            file_result['status'] = "Passport is Expired."

                        else:
                            #show_message = "Something is wrong Please Re-Upload the Passport or Upload the different File Format." 
                            file_result['status'] = "Something is wrong Please Re-Upload the Passport or Upload the different File Format." 

                    except requests.exceptions.Timeout:
                        file_result['status'] = "Upload the different File Format."

                    except requests.exceptions.RequestException as e:
                        file_result['status'] = "Upload the different File Format."  
                        

                    if passport_list:
                        passport_name = passport_list[0][4] + ' ' + passport_list[0][3]
                        print(passport_name)

                        for i in data_to_show:
                            guest_name_1 = i[0]  + ' ' + i[1]
                            guest_name_2 = i[1]  + ' ' + i[0]
                            
                            if passport_name.upper() == guest_name_1 or passport_name.upper() == guest_name_2:
                                try:
                                    with connection.cursor() as cursor:
                                        cursor.execute("""
                                            INSERT INTO TBL_TRAVELLER_PASSPORT_DETAILS
                                            (TRAV_ID, Agent_ID, PKG_ID, TourDate, Passport_Type, Country_Code, Passport_Number, First_Name, Last_Name, Nationality, DOB, Place_of_Birth, Gender, Date_of_Issue, Date_of_Expiry, Issuing_Authority, Passport_URL) 
                                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, '')
                                        """, [
                                            i[2], selected_agentID, selected_pkg_id, formatted_date, 
                                            passport_list[0][0], passport_list[0][1], passport_list[0][2], 
                                            passport_list[0][4], passport_list[0][3], passport_list[0][5],
                                            datetime.strptime(passport_list[0][6], "%d %b %Y").date(), 
                                            passport_list[0][7], passport_list[0][8],
                                            datetime.strptime(passport_list[0][9], "%d %b %Y").date(),
                                            datetime.strptime(passport_list[0][10], "%d %b %Y").date(),
                                            passport_list[0][11]
                                        ])

                                    # Step 2: Only if insert succeeds, upload to S3
                                    bucket_name = 'agentdatas3'
                                    today = datetime.today().strftime("%Y-%m-%d")
                                    original_extension = os.path.splitext(uploaded_file.name)[1]
                                    file_name_for_s3 = f"{i[2]}{original_extension}"
                                    s3_key = f"passports/{today}/{file_name_for_s3}"

                                    file_url = upload_file_to_s3(uploaded_file, bucket_name, s3_key)
                                    print("Uploaded file URL:", file_url)

                                    # Step 3: Update the inserted row with S3 URL
                                    with connection.cursor() as cursor:
                                        cursor.execute("""
                                            UPDATE TBL_TRAVELLER_PASSPORT_DETAILS 
                                            SET Passport_URL = %s 
                                            WHERE TRAV_ID = %s
                                        """, [file_url, i[2]])

                                    file_result['status'] = "Passport is Valid."
                                    break

                                except Exception as e:
                                    try:
                                        print("Start")
                                        with connection.cursor() as cursor:
                                            cursor.execute(f"select 1 from TBL_TRAVELLER_PASSPORT_DETAILS where TRAV_ID = {i[2]}")

                                            check_passport_data = cursor.fetchone()
                                            print(check_passport_data)

                                            if check_passport_data is not None:
                                                file_result['status'] = "Passport is Valid."
                                                break

                                    except Exception as inner_e:
                                        print("Error during insert-check fallback:", inner_e)

                                    print("Insert failed:", e)

                        
                            else:
                                # show_message = "Guest Name is not found in the Guest List or You Already Submit this Passport."
                                file_result['status'] = "Guest Name is not found in the Guest List"

                    results.append(file_result) 

                # return redirect(f'/trav-info?get_details={get_details}')    

            except Exception as e:
                error_message = f"Error: {str(e)}"
                return render(request, 'no_database.html', {})


        elif get_details == "extra_details":
            try:
                if selected_agentID and selected_pkg_id and formatted_date:
                    check_draft_data = ''
                    with connection.cursor() as cursor:
                    
                        cursor.execute(f"SELECT COUNT(*) FROM TBL_TRAVELLER_NAME_DRAFT WHERE PKG_ID = {selected_pkg_id} AND Tourdate = '{formatted_date}' AND AgentID = '{selected_agentID}'")
                        check_draft_data = cursor.fetchone()[0]
                        
                        if check_draft_data > 0:
                            cursor.execute(f"""SELECT Optional_Tour_Taken, Extension_Tour_Taken, Kids_Details, Special_Request
                                            FROM TBL_TRAVELLER_NAME_DRAFT
                                            WHERE PKG_ID = {selected_pkg_id}
                                                AND Tourdate = '{formatted_date}'
                                                AND AgentID = '{selected_agentID}'
                                                ORDER BY RoomNo
                                           """)
                            extra_details = cursor.fetchone()

                            print("Check",extra_details)

                        else: 
                            error_message = f'No data found. Please fill the <a href="/trav-info?get_details=guest_details">Guest Details</a> first.'

                    if optional_tour and extension_tour:
                        with connection.cursor() as cursor:
                            update_query = f"""
                                UPDATE TBL_TRAVELLER_NAME_DRAFT
                                SET Optional_Tour_Taken = %s,
                                    Extension_Tour_Taken = %s,
                                    Kids_Details = %s,
                                    Special_Request = %s
                                WHERE PKG_ID = %s
                                AND Tourdate = %s
                                AND AgentID = %s
                                AND STATUS = 'Active'
                            """
                            cursor.execute(update_query, [
                                optional_tour,
                                extension_tour,
                                kids_info,
                                special_request,
                                selected_pkg_id,
                                formatted_date,
                                selected_agentID
                            ])
                            connection.commit()
                            
                        messages.success(request, "Information Updated Successfully!")
                        return redirect(f'/trav-info?get_details={get_details}')
                        
            except:
                return render(request, 'no_database.html', {})

        executed_query = connection.queries[-1]['sql'] if connection.queries else "No query executed"

        #print(executed_query)

        return render(request, 'trav_info.html', {
            'get_details' : get_details,
            'executed_query' : executed_query,
            'error_message' : error_message,
            'selected_pkg_id' : selected_pkg_id,
           'selected_pkg_title' : selected_pkg_title,
            'year_list' : year_list,
            'selected_year' : selected_year,
            'selected_date' : selected_date,
            #'data_headers' : data_headers,
            'data_to_show' : data_to_show,
            'selected_agentID' : selected_agentID,
            'agent_details' : agent_details,
            'flight_list' : flight_list,
            'selected_trav_id' : selected_trav_id,
            'show_message' : show_message,
            'optional_tour' : optional_tour,
            'extension_tour' : extension_tour,
            'kids_info' : kids_info,
            'special_request' : special_request,
            'extra_details': extra_details,
            'date_to_show' : date_to_show,
            'no_data_message' : no_data_message,
            'upload_results' : results,
            'upload_success' : upload_success,
            'flight_traveller_name' : flight_traveller_name,
            'file_uploaded_type' : file_uploaded_type,
            'no_details_found' : no_details_found,
            'guest_flight_details' : guest_flight_details,
            'guest_trav_id' : guest_trav_id,
            # 'number_of_guest_left' : number_of_guest_left,
            'data_count_show' : data_count_show,

        })
    





@staff_member_required
def agent_data(request):
    data = []
    column_headers = []

    nationality_query = ""
    company_query = ""
    created_date_query = ""
    agent_website_query = ""
    conditional_query = ""
    final_query = ""
 
    selected_nationality = request.GET.getlist('nationality')
    today_date = datetime.today().date()
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    agent_website = request.GET.get('agent_website')
    submit = request.GET.get('submit')
    download_format = request.GET.get('download_format','')
    download = request.GET.get('download','')
    selected_company = request.GET.getlist('company')

    with connection.cursor() as cursor:
        cursor.execute("""
                        SELECT DISTINCT 
                            LTRIM(RTRIM(UPPER(Nationality))) AS Cleaned_Nationality
                        FROM tbl_agent
                        WHERE Nationality IS NOT NULL 
                        AND Nationality <> ''
                        AND LEN(LTRIM(RTRIM(Nationality))) >= 4
                        AND LTRIM(RTRIM(Nationality)) NOT LIKE '%[^A-Za-z0-9 ?/_-]%' 
                        AND LTRIM(RTRIM(Nationality)) LIKE '%[A-Za-z0-9]%'            
                        ORDER BY Cleaned_Nationality;
                        """)
        nationality = [row[0] for row in cursor.fetchall()]

    with connection.cursor() as cursor:
        cursor.execute("""select distinct LTRIM(RTRIM(UPPER(Comp_Name))) AS Cleaned_Company
                        from TBL_Agent
                        WHERE Comp_Name IS NOT NULL
                        AND Comp_Name <> ''
                        AND LEN(LTRIM(RTRIM(Comp_Name))) >= 4
                        AND LTRIM(RTRIM(Comp_Name)) NOT LIKE '%[^A-Za-z0-9 ?/_]%' 
                        AND LTRIM(RTRIM(Comp_Name)) LIKE '%[A-Za-z0-9]%'            
                        ORDER BY Cleaned_Company;
                       """)
        company = [row[0] for row in cursor.fetchall()]

    try:
        if selected_nationality:
            if len(selected_nationality) == 1:
                nationality_query = f"Nationality IN ('{selected_nationality[0]}')"
            else:
                nationality_tuple = "', '".join(selected_nationality)
                nationality_query = f"Nationality IN ('{nationality_tuple}')"
            print(nationality_query)

        if selected_company:
            if len(selected_company) == 1:
                company_query = f"Comp_Name IN ('{selected_company[0]}')"
            else:
                company_tuple = "', '".join(selected_company)
                company_query = f"Comp_Name IN ('{company_tuple}')"
            print(company_query)
        
        if start_date and end_date:
            created_date_query = f"CreatedDate BETWEEN '{start_date}' AND '{end_date}'"
            print(created_date_query)

        if agent_website:
            agent_website_query = f"IS_WEBSITE = '{agent_website}'"
            print(agent_website_query)


        if nationality_query or company_query or created_date_query or agent_website_query :
            filters = [q for q in [nationality_query, company_query, created_date_query, agent_website_query] if q]
            conditional_query= " AND ".join(filters)
            print(conditional_query)

        print(download)
        print(download_format)

        if download == 'true':
            final_query = """SELECT [Id]
                            ,[Name]
                            ,[Address]
                            ,[Contact]
                            ,[Emailid]
                            ,[Nationality]
                            ,[Comp_Name]
                            ,[Type_Bussiness]
                            ,[Date_establishment]
                            ,[Designation]
                            ,[Website]
                            ,[link_Facebook]
                            ,[Membership]
                            ,[Agency_Group]
                            ,[Destination_Sell]
                            ,[Destination_Withus]
                            ,[Suscription]
                            ,[Acceptance]
                            ,[DestWithus_India]
                            ,[DestWithus_Dubai]
                            ,[DestWithus_Srilanka]
                            ,[DestWithus_Maldives]
                            ,[DestWithus_Thailand]
                            ,[DestWithus_Egypt]
                            ,[DestWithus_Turkey]
                            ,[DestWithus_Singapore]
                            ,[DestWithus_Vietnam]
                            ,[DestWithus_Jordan]
                            ,[DestWithus_Indonesia]
                            ,[DestWithus_Nepal]
                            ,[DestWithus_Australia]
                            ,[DestWithus_New_Zealand]
                            ,[DestWithus_Malaysia]
                            ,[DestWithus_China]
                            ,[DestWithus_South_Africa]
                            ,[DestWithus_Kenya]
                            ,[DestWithus_Israel]
                            ,[DestWithus_Greece]
                            ,[DestWithus_Morocco]
                            ,[DestWithus_Europe]
                            ,[AgentID]
                            ,[prifix]
                            ,[other]
                            ,[status]
                            ,[dob]
                            ,[CreditLimit]
                            ,[mobileno]
                            ,[link_Instagram]
                            ,[CreatedDate]
                            ,[UpdatedDate]
                            ,[User_Type]
                            ,[IS_WEBSITE]
                            ,[WebsiteName]
                            ,[DistributerID]
                            ,[AgentType]
                            ,[IsPremium]
                            ,[CashbackCreatedDate]
                            ,[CashbackAmount]
                            ,[ReferralCode]
                            ,[CountryName]
                            ,[CityName]
                            ,[Whatsapp]
                            ,[CityNameNew]
                            ,[StateName]
                            ,[CountryNameNew]
                            ,[LastLogin]
                            ,[PromoteAgency]
                            ,[StepCompleted]
                            ,[Role] 
                            FROM TBL_Agent
							WHERE status = 1
                            """
            if conditional_query:
                final_query += f" WHERE {conditional_query}"
            
            print(final_query)
            with connection.cursor() as cursor:
                cursor.execute(final_query)
                data = cursor.fetchall()
                column_headers = [col[0] for col in cursor.description]

                # print(data)
                # print(column_headers)
            
            if not data:
                error_message = 'No data available for the selected filters... Please wait 3 Seconds'
                return render(request, 'agent_data.html', {
                    'nationality': nationality,
                    'today_date': today_date,
                    'start_date': start_date,
                    'end_date': end_date,
                    'selected_nationality': selected_nationality,
                    'agent_website': agent_website,
                    'data': [],
                    'data_available': False,
                    'company': company,
                    'selected_company': selected_company,
                    'download': download,
                    'error_message': error_message,
                })

            if download_format == 'pdf':
                
                return pdf_download(column_headers, data)
            
            elif download_format == 'csv':
                
                return csv_download(column_headers, data)
            
            elif download_format == 'json':
                
                return json_download(column_headers, data)
            

    except Exception as e:
        pass



    return render(request, 'agent_data.html', {
        'nationality' : nationality,
        'today_date' : today_date,
        'start_date' : start_date,
        'end_date' : end_date,
        'selected_nationality': selected_nationality,
        'agent_website' : agent_website,
        'company' : company,
        'selected_company' : selected_company,


    })


def pdf_download(column_headers,data):
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page(orientation='L')  # Landscape orientation for more space
    #pdf.set_font('Arial', '', 8)

    page_width = pdf.w - 2 * pdf.l_margin  # Calculate usable page width
    max_columns_per_page = 10  # Adjust based on column widths and page size

    # Split columns into chunks
    column_chunks = [
        column_headers[i:i + max_columns_per_page] 
        for i in range(0, len(column_headers), max_columns_per_page)
    ]

    for chunk_index, column_chunk in enumerate(column_chunks):
        # Add a new page for each chunk after the first
        if chunk_index > 0:
            pdf.add_page(orientation='L')

        # Calculate dynamic column width
        column_width = page_width / len(column_chunk)

        pdf.set_font('Arial', 'B', 10)
        # Add headers for this chunk
        for header in column_chunk:
            pdf.cell(column_width, 10, header, border=1, align='C')
        pdf.ln()

        pdf.set_font('Arial', '', 8)
        # Add data rows for this chunk
        for row in data:
            for col_index in range(len(column_chunk)):
                col_idx = col_index + chunk_index * max_columns_per_page
                if col_idx < len(row):  # Check if index is within bounds
                    # Convert value to string and handle None values
                    cell_value = str(row[col_idx]) if row[col_idx] is not None else ""
                    pdf.cell(column_width, 10, cell_value, border=1, align='C')
            pdf.ln()

    # Return the PDF response
    response = HttpResponse(pdf.output(dest='S').encode('latin1'), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="agent_data.pdf"'
    return response


def csv_download(column_headers,data):
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="agent_data.csv"'

    writer = csv.writer(response)
    # Write the header
    writer.writerow(column_headers)

    # Write data rows
    for row in data:
        writer.writerow(row)

    return response


def json_download(column_headers,data):
    data = []
    for row in data:
        row_data = dict(zip(column_headers, row))
        data.append(row_data)

    response = HttpResponse(
        json.dumps(data, indent=4), content_type='application/json'
    )
    response['Content-Disposition'] = 'attachment; filename="agent_data.json"'
    return response



class dashboard:
    @csrf_exempt
    
    def login_page(request):
        print("START")
        print("/login")

        if request.method == 'POST':
            try:
                data = json.loads(request.body)
                username = data.get('username')
                password = data.get('password')
                user_valid = False

                print("Username:", username)
                print("Password:", password)

                if username == 'aman.culture' and password == 'Culture@123':
                    user_valid = True
                else:
                    with connection.cursor() as cursor:
                        cursor.execute("""
                            SELECT StaffID, StaffName FROM TBLStaff 
                            WHERE type = %s AND UserID = %s AND Password = %s AND Status = 1
                        """, ['admin', username, password])
                        result = cursor.fetchone()
                        if result:
                            user_valid = True

                if user_valid:
                    payload = {
                        'username': username,
                        'exp': datetime.utcnow() + timedelta(seconds=int(JWT_EXP_DELTA_SECONDS))
                    }
                    token = jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)
                    
                    print(token)
                    return JsonResponse({
                        'success': True,
                        'message': 'Login successful',
                        'token': token
                    })
                else:
                    return JsonResponse({'success': False, 'message': 'Invalid Username or Password'}, status=401)

            except json.JSONDecodeError:
                return JsonResponse({'success': False, 'message': 'Invalid JSON'}, status=400)

        return JsonResponse({'success': False, 'message': 'Only POST method is allowed'}, status=405)


    @csrf_exempt
    def excel_download_api(request):
        
        if request.method != "POST":
            return HttpResponse("Only POST method is allowed", status=405)

        try:
            body_unicode = request.body.decode('utf-8')
            agents = json.loads(body_unicode)

            if not isinstance(agents, list):
                return HttpResponse("Expected a list of agents", status=400)

            # Preserve order of months as they appear first time in data
            seen_months = OrderedDict()
            for agent in agents:
                for entry in agent.get("data", []):
                    month = entry["Month_Year_Str"]
                    if month not in seen_months:
                        seen_months[month] = None

            all_months = list(seen_months.keys())

            # Define headers
            headers = ["AgentID"] + all_months

            # Build row data
            rows = []
            for agent in agents:
                agent_id = agent.get("AgentID", "")
                month_data = {entry["Month_Year_Str"]: entry["Count"] for entry in agent.get("data", [])}
                row = [agent_id] + [month_data.get(month, 0) for month in all_months]
                rows.append(row)

            # Create Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Agent Data"

            # Write headers
            for col_num, column_title in enumerate(headers, 1):
                ws.cell(row=1, column=col_num, value=column_title)

            # Write data rows
            for row_num, row_data in enumerate(rows, 2):
                for col_num, cell_value in enumerate(row_data, 1):
                    ws.cell(row=row_num, column=col_num, value=cell_value)

            # Return response as Excel file
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="agent_data.xlsx"'

            wb.save(response)
            return response

        except json.JSONDecodeError:
            return HttpResponse("Invalid JSON", status=400)
        except Exception as e:
            return HttpResponse(f"Error processing request: {str(e)}", status=500)
        

    def get_inactive_agents(request):
        print("api/inactive-agents")

        current_year = datetime.now().year
        current_year_1 = current_year-1
        current_year_2 = current_year-2
        current_year_3 = current_year-3

        try:
            with connection.cursor() as cursor:
                cursor.execute(f"""
                        select AgentID, LastLogin as [LastLogin Date], CreatedDate as [Account Created Date] 
                        from TBL_Agent
                        where status = 1
                        AND LastLogin is not Null
                        AND CreatedDate is not Null
                        AND AgentID NOT In (select distinct (Agentid) as Agentid from tbl_agent_culture)
                            """)
                
                executed_query = connection.queries[-1]['sql'] if connection.queries else "No query executed"
                
                print(executed_query)
                
                header = [desc[0] for desc in cursor.description]
                data = cursor.fetchall()
                data_tuples = [tuple(row) for row in data]
                df_agent = pd.DataFrame(data_tuples, columns=header)
            
            df_agent['Account Created Date'] = pd.to_datetime(df_agent['Account Created Date'], errors='coerce')
            df_agent['LastLogin Date'] = pd.to_datetime(df_agent['LastLogin Date'], errors='coerce')    

            df_agent['Year_Of_Account_Creation'] = df_agent['Account Created Date'].dt.year
            df_agent['Year_Of_Last_Login'] = df_agent['LastLogin Date'].dt.year
            df_agent = df_agent[['AgentID', 'Year_Of_Account_Creation', 'Year_Of_Last_Login']].copy()

            total_1 = df_agent[
                (df_agent['Year_Of_Account_Creation'] == current_year_1)
            ].shape[0]

            total_2 = df_agent[
                (df_agent['Year_Of_Account_Creation'] == current_year_2)
            ].shape[0]

            total_3 = df_agent[
                (df_agent['Year_Of_Account_Creation'] == current_year_3)
            ].shape[0]

            overall = df_agent['Year_Of_Account_Creation'].shape[0]

            count_1 = df_agent[
                (df_agent['Year_Of_Account_Creation'] == current_year_1) &
                (df_agent['Year_Of_Last_Login'] != current_year)
            ].shape[0]

            count_2 = df_agent[
                (df_agent['Year_Of_Account_Creation'] == current_year_2) &
                (df_agent['Year_Of_Last_Login'] != current_year)
            ].shape[0]


            count_3 = df_agent[
                (df_agent['Year_Of_Account_Creation'] == current_year_3) &
                (df_agent['Year_Of_Last_Login'] != current_year)
            ].shape[0]

            total_active =df_agent[df_agent['Year_Of_Last_Login'] == current_year].shape[0]

            df_chart = pd.DataFrame({
                'Year': [current_year_3, current_year_2, current_year_1, f'Active in {current_year}'],
                'Total_Account_Created':[total_3, total_2, total_1, overall],
                'Inactive_Agents': [count_3, count_2, count_1, total_active]
                
                })


            return JsonResponse(df_chart.to_dict(orient='records'), safe=False)

        except Exception as e:
            print("Error : ", e)
            return JsonResponse({"error": str(e)}, status=500)



    def agent_login(request):
        print("api/agent-login")
        try:
            with connection.cursor() as cursor:
                # Optimized SQL query - do all processing in database
                cursor.execute("""
                    ;WITH cleaned_data AS (
                        SELECT 
                            UPPER(LTRIM(RTRIM(AGENTID))) as agent_id,
                            LOGINDATE as login_date,
                            YEAR(LOGINDATE) as login_year,
                            DATENAME(MONTH, LOGINDATE) as login_month_name
                        FROM TBL_LOGIN 
                        WHERE AGENTID IS NOT NULL 
                        AND LOGINDATE IS NOT NULL
                        AND Loginby IS NULL
                        AND LTRIM(RTRIM(AGENTID)) != ''
                        AND AgentID NOT In (select distinct (Agentid) as Agentid from tbl_agent_culture)
                    ),
                    unique_logins AS (
                        SELECT DISTINCT
                            agent_id,
                            login_month_name,
                            login_year
                        FROM cleaned_data
                    ),
                    monthly_counts AS (
                        SELECT 
                            login_month_name,
                            login_year,
                            COUNT(agent_id) as agent_count
                        FROM unique_logins
                        GROUP BY login_month_name, login_year
                    ),
                    all_months AS (
                        SELECT 'January' as month_name, 'Jan' as month_short, 1 as month_order
                        UNION ALL SELECT 'February', 'Feb', 2
                        UNION ALL SELECT 'March', 'Mar', 3
                        UNION ALL SELECT 'April', 'Apr', 4
                        UNION ALL SELECT 'May', 'May', 5
                        UNION ALL SELECT 'June', 'Jun', 6
                        UNION ALL SELECT 'July', 'Jul', 7
                        UNION ALL SELECT 'August', 'Aug', 8
                        UNION ALL SELECT 'September', 'Sep', 9
                        UNION ALL SELECT 'October', 'Oct', 10
                        UNION ALL SELECT 'November', 'Nov', 11
                        UNION ALL SELECT 'December', 'Dec', 12
                    )
                    SELECT 
                        am.month_short as Month,
                        am.month_name,
                        COALESCE(mc.login_year, 0) as login_year,
                        COALESCE(mc.agent_count, 0) as agent_count
                    FROM all_months am
                    LEFT JOIN monthly_counts mc ON am.month_name = mc.login_month_name
                    ORDER BY am.month_order, mc.login_year
                """)
                results = cursor.fetchall()
                
            data_dict = {}
            years = set()
            
            for month_short, month_name, year, count in results:
                if month_short not in data_dict:
                    data_dict[month_short] = {'Month': month_short}
                
                if year and year != 0:
                    data_dict[month_short][str(year)] = int(count)
                    years.add(year)
                
            # Ensure all months have all years (fill with 0 if missing)
            years = sorted(years) if years else []
            print(years)
            
            final_data = []
            month_order = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                        'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
            
            for month in month_order:
                row = {'Month': month}
                for year in years:
                    row[str(year)] = data_dict.get(month, {}).get(str(year), 0)
                final_data.append(row)

            return JsonResponse(final_data, safe=False)

        except Exception as e:
            print("Error : ", e)
            return JsonResponse({"error": str(e)}, status=500)


    def get_quarterly_tour(request):
        print('quarterly-tour')

        try:
            with connection.cursor() as cursor:
                cursor.execute(f"""select CONCAT(trim(upper(AgentID)),PackgID,convert(date,tourdate,103)) as [Unique ID], convert(date,tourdate,103) as [Tour Date] 
                                    from TBL_BOOKING
                                    where txn_msg = 'success'
                                        AND Is_cancelled <> 1
                                        AND CreatedDate is NOT NULL
                                        AND tourdate is NOT NULL
                                        AND agentId is NOT NULL
                                        AND PackgID Is NOT NULL
                                        AND CreatedDate != ''
                                        AND tourdate != ''
                                        AND agentId != ''
                                        AND PackgID != '' 
                                        AND Agentid not in (select distinct (Agentid) as Agentid from tbl_agent_culture )	
                                
                            """)

                header = [desc[0] for desc in cursor.description]
                data = cursor.fetchall()
                data_tuples = [tuple(row) for row in data]
                df_booking = pd.DataFrame(data_tuples, columns=header)


            df_booking = df_booking.drop_duplicates(subset=['Unique ID'])

            df_booking['Tour Date'] = pd.to_datetime(df_booking['Tour Date'], errors='coerce')

            df_booking['Tour_Year'] = df_booking['Tour Date'].dt.year

            current_year = datetime.now().year

            valid_years = [current_year - i for i in range(4)]  # [2025, 2024, 2023, 2022]
            df_booking = df_booking[df_booking['Tour_Year'].isin(valid_years)]

            df_booking['Tour_Quarter'] = df_booking['Tour Date'].dt.quarter

            print(df_booking[df_booking['Tour_Year'] == 2022])

            pivot_df = df_booking.groupby(['Tour_Year', 'Tour_Quarter'])['Unique ID'] \
                                .nunique() \
                                .unstack(fill_value=0) \
                                .astype(int)

            # Rename columns to "Quarter 1", "Quarter 2", etc.
            pivot_df.columns = [f"Quarter {q}" for q in pivot_df.columns]

            # Reset index to make "Year" a column
            pivot_df = pivot_df.reset_index().rename(columns={"Tour_Year": "Year"})
            
            return JsonResponse(pivot_df.to_dict(orient='records'), safe=False)
        
        except Exception as e:
            print("Error : ", e)
            return JsonResponse({"error": str(e)}, status=500)



    def most_popular_tourdate(request):
        print('api/popular-date')

        try:
            with connection.cursor() as cursor:
                cursor.execute(f"""SELECT CONCAT(trim(upper(AgentID)),PackgID,convert(date,tourdate,103)) as [Unique ID], convert(date,tourdate,103) as [Tour Date] 
                                    from TBL_BOOKING
                                    where txn_msg = 'success'
                                    AND Is_cancelled <> 1
                                    AND CreatedDate is NOT NULL
                                    AND tourdate is NOT NULL
                                    AND agentId is NOT NULL
                                    AND PackgID Is NOT NULL
                                    AND CreatedDate != ''
                                    AND tourdate != ''
                                    AND agentId != ''
                                    AND PackgID != '' 	
                            """)

                header = [desc[0] for desc in cursor.description]
                data = cursor.fetchall()
                data_tuples = [tuple(row) for row in data]
                df_booking = pd.DataFrame(data_tuples, columns=header)

            df_booking =df_booking.drop_duplicates(subset=['Unique ID'])

            df_booking['Tour Date'] = pd.to_datetime(df_booking['Tour Date'], errors='coerce')
            df_booking = df_booking.dropna(subset=['Tour Date'])
            

            df_booking['Tour_Month'] = df_booking['Tour Date'].dt.month_name()
            df_booking['Tour_Day'] = df_booking['Tour Date'].dt.day  

            df_chart = df_booking[['Tour_Month', 'Tour_Day']].copy()

            month_order = ['January', 'February', 'March', 'April', 'May', 'June',
                        'July', 'August', 'September', 'October', 'November', 'December']
            df_chart['Tour_Month'] = pd.Categorical(df_chart['Tour_Month'], categories=month_order, ordered=True)
            df_chart = df_chart.sort_values('Tour_Month')

            return JsonResponse(df_chart.to_dict(orient='records'), safe=False)

        except Exception as e:
            print("Error : ", e)
            return JsonResponse({"error": str(e)}, status=500)


    ## OLD one
    def txn_tourdate_relation(request):
        print('txn_tourdate_relation')

        try:
            with connection.cursor() as cursor:   
                cursor.execute(f"""select CONCAT(trim(upper(AgentID)), PackgID, convert(date,tourdate,103)) as [Unique ID],  
                            convert(date,tourdate,103) as [Tour Date], 
                            convert(date,CreatedDate,103) as [Created Date], 
                            PackgID
                                    from TBL_BOOKING 
                                    where txn_msg = 'success'
                                    AND Is_cancelled <> 1
                                    AND CreatedDate is not Null
                                    AND PackgID is not Null
                                    AND tourdate is not Null
                                    AND agentId is not Null
                                    AND agentid != ''
                                    AND PackgID  != ''
                                    AND tourdate  != ''
                                    AND CreatedDate != ''
                                    AND year(convert(date,CreatedDate,103)) > 2022
                                    AND Agentid not in (select distinct (Agentid) as Agentid from tbl_agent_culture )
                            """)
                header = [desc[0] for desc in cursor.description]
                data = cursor.fetchall()
                data_tuples = [tuple(row) for row in data]
                df_booking = pd.DataFrame(data_tuples, columns=header)

            df_booking = df_booking.dropna(subset=['Tour Date', 'Created Date', 'PackgID'])

            df_booking['Tour Date'] = pd.to_datetime(df_booking['Tour Date'], errors='coerce')
            df_booking['Created Date'] = pd.to_datetime(df_booking['Created Date'], errors='coerce')
            
            df_booking = df_booking.sort_values('Created Date').drop_duplicates(subset=['Unique ID'], keep='first')

            df_booking['Tour_Month'] = df_booking['Tour Date'].dt.month_name()
            df_booking['Created_Month'] = df_booking['Created Date'].dt.month_name()

            month_order = ['January', 'February', 'March', 'April', 'May', 'June',
                        'July', 'August', 'September', 'October', 'November', 'December']
            
            df_booking['Tour_Month'] = pd.Categorical(df_booking['Tour_Month'], categories=month_order, ordered=True)
            df_booking['Created_Month'] = pd.Categorical(df_booking['Created_Month'], categories=month_order, ordered=True)

            df_booking['PackgID'] = pd.to_numeric(df_booking['PackgID'], errors='coerce')

            packg_ids = request.GET.get('packg_ids')

        
            if packg_ids:
                try:
                    selected_id = int(packg_ids.strip())
                    df_booking = df_booking[df_booking['PackgID'] == selected_id]
                except ValueError:
                    return JsonResponse({'error': 'Invalid packg_ids value'}, safe=False), 400

            # Group by both months and count
            df_chart = df_booking.groupby(['Tour_Month', 'Created_Month']).size().reset_index(name='Count')

            # Sort properly
            df_chart = df_chart.sort_values(['Tour_Month', 'Created_Month'])

            return JsonResponse(df_chart.to_dict(orient='records'), safe=False)
        
        except Exception as e:
            print("Error : ", e)
            return JsonResponse({"error": str(e)}, status=500)


    def notactive_agent(request):
        print("api/active-agent-booking")

        try:
            # Calculate date thresholds once
            today = datetime.now().date()
            one_year_ago = today - timedelta(days=365)
            
            # Calculate cutoff dates for all durations
            cutoff_3_months = today - timedelta(days=90)
            cutoff_6_months = today - timedelta(days=180)
            cutoff_9_months = today - timedelta(days=270)
            cutoff_12_months = today - timedelta(days=365)

            with connection.cursor() as cursor:
                # Single optimized query that gets all data and calculations
                cursor.execute("""
                    ;WITH active_agents AS (
                        SELECT DISTINCT AgentID
                        FROM TBL_Agent
                        WHERE status = 1
                        AND LastLogin IS NOT NULL
                        AND CreatedDate IS NOT NULL
                        AND AgentID NOT In (select distinct (Agentid) as Agentid from tbl_agent_culture)
                        AND CAST(LastLogin AS DATE) >= CAST(DATEADD(DAY, -365, GETDATE()) AS DATE)
                    ),
                    unique_bookings AS (
                        SELECT 
                            agentId,
                            CAST(CreatedDate AS DATE) as booking_date,
                            ROW_NUMBER() OVER (
                                PARTITION BY CONCAT(AgentID, PackgID, tourdate) 
                                ORDER BY CreatedDate
                            ) as rn
                        FROM TBL_BOOKING
                        WHERE txn_msg = 'success'
                        AND Is_cancelled <> 1
                        AND tourdate IS NOT NULL 
                        AND PackgID IS NOT NULL
                        AND agentId IS NOT NULL
                        AND CreatedDate IS NOT NULL
                        AND CreatedDate != ''
                        AND tourdate != ''
                        AND agentId != ''
                        AND PackgID != '' 	      
                    ),
                    recent_bookings AS (
                        SELECT DISTINCT agentId
                        FROM unique_bookings
                        WHERE rn = 1  -- Only first occurrence of each unique booking
                    ),
                    agent_booking_activity AS (
                        SELECT 
                            aa.AgentID,
                            CASE WHEN rb3.agentId IS NOT NULL THEN 1 ELSE 0 END as has_booking_3m,
                            CASE WHEN rb6.agentId IS NOT NULL THEN 1 ELSE 0 END as has_booking_6m,
                            CASE WHEN rb9.agentId IS NOT NULL THEN 1 ELSE 0 END as has_booking_9m,
                            CASE WHEN rb12.agentId IS NOT NULL THEN 1 ELSE 0 END as has_booking_12m
                        FROM active_agents aa
                        LEFT JOIN (
                            SELECT DISTINCT agentId 
                            FROM unique_bookings 
                            WHERE rn = 1 AND booking_date >= %s
                        ) rb3 ON aa.AgentID = rb3.agentId
                        LEFT JOIN (
                            SELECT DISTINCT agentId 
                            FROM unique_bookings 
                            WHERE rn = 1 AND booking_date >= %s
                        ) rb6 ON aa.AgentID = rb6.agentId
                        LEFT JOIN (
                            SELECT DISTINCT agentId 
                            FROM unique_bookings 
                            WHERE rn = 1 AND booking_date >= %s
                        ) rb9 ON aa.AgentID = rb9.agentId
                        LEFT JOIN (
                            SELECT DISTINCT agentId 
                            FROM unique_bookings 
                            WHERE rn = 1 AND booking_date >= %s
                        ) rb12 ON aa.AgentID = rb12.agentId
                    )
                    SELECT 
                        COUNT(*) as total_agents,
                        SUM(has_booking_3m) as with_booking_3m,
                        SUM(has_booking_6m) as with_booking_6m,
                        SUM(has_booking_9m) as with_booking_9m,
                        SUM(has_booking_12m) as with_booking_12m
                    FROM agent_booking_activity
                """, [cutoff_3_months, cutoff_6_months, cutoff_9_months, cutoff_12_months])
                
                result = cursor.fetchone()

            # Process results
            if result:
                total_agents, with_3m, with_6m, with_9m, with_12m = result
                
                results = [
                    {
                        "Duration": "Last 3 Month",
                        "With Booking": int(with_3m or 0),
                        "Without Booking": int(total_agents - (with_3m or 0))
                    },
                    {
                        "Duration": "Last 6 Month", 
                        "With Booking": int(with_6m or 0),
                        "Without Booking": int(total_agents - (with_6m or 0))
                    },
                    {
                        "Duration": "Last 9 Month",
                        "With Booking": int(with_9m or 0), 
                        "Without Booking": int(total_agents - (with_9m or 0))
                    },
                    {
                        "Duration": "Last 12 Month",
                        "With Booking": int(with_12m or 0),
                        "Without Booking": int(total_agents - (with_12m or 0))
                    }
                ]
            else:
                # Fallback if no data
                results = [
                    {"Duration": f"Last {months} Month", "With Booking": 0, "Without Booking": 0}
                    for months in [3, 6, 9, 12]
                ]

            return JsonResponse(results, safe=False)
            
        except Exception as e:
            print("Error : ", e)
            return JsonResponse({"error": str(e)}, status=500)




    def get_quarterly_booking(request):
        print('quarterly-booking')

        try:
            with connection.cursor() as cursor:
                cursor.execute(f"""select CONCAT(trim(upper(AgentID)),PackgID,convert(date,tourdate,103)) as [Unique ID], convert(date,CreatedDate,103) as [Created Date]
                                from TBL_BOOKING
                                where txn_msg = 'success'
                                    AND Is_cancelled <> 1
                                    AND tourdate is not Null 
                                    AND PackgID is not Null
                                    AND AgentId is not Null
                                    AND CreatedDate is not Null
                                    AND tourdate != '' 
                                    AND PackgID != '' 
                                    AND AgentId != '' 
                                    AND CreatedDate != '' 
                                    AND Agentid not in (select distinct (Agentid) as Agentid from tbl_agent_culture )
                                """)

                header = [desc[0] for desc in cursor.description]
                data = cursor.fetchall()
                data_tuples = [tuple(row) for row in data]
                df_booking = pd.DataFrame(data_tuples, columns=header)

            print(df_booking.shape)

            df_booking['Created Date'] = pd.to_datetime(df_booking['Created Date'], errors='coerce')

            df_booking = df_booking.sort_values('Created Date').drop_duplicates(subset=['Unique ID'], keep='first')

            print(df_booking.shape)

            df_booking['Booking_Year'] = df_booking['Created Date'].dt.year

            current_year = datetime.now().year

            valid_years = [current_year - i for i in range(4)]  # [2025, 2024, 2023, 2022]
            df_booking = df_booking[df_booking['Booking_Year'].isin(valid_years)]

            print(df_booking)

            df_booking['Tour_Quarter'] = df_booking['Created Date'].dt.quarter

            print(df_booking)

            pivot_df = df_booking.groupby(['Booking_Year', 'Tour_Quarter'])['Unique ID'] \
                                .nunique() \
                                .unstack(fill_value=0) \
                                .astype(int)

            pivot_df.columns = [f"Quarter {q}" for q in pivot_df.columns]

            pivot_df = pivot_df.reset_index().rename(columns={"Booking_Year": "Year"})

            # Convert to list of dicts
            df_chart = pivot_df  # already a DataFrame

            
            return JsonResponse(df_chart.to_dict(orient='records'),safe=False)

        except Exception as e:
            print("Error : ", e)
            return JsonResponse({"error": str(e)}, status=500)
        

    def frequently_login(request):
        print('api/frequently-login')

        try:
            with connection.cursor() as cursor:
                cursor.execute(f"""
                               ;WITH Agent_Emails AS (
                                select Emailid from tbl_agent
                                where agentid in (select Agentid from tbl_agent_culture)
                                ),
                                Agent_Usernames AS (
                                    select Username from tbl_agent
                                    where agentid in (select Agentid from tbl_agent_culture)
                                )
                                SELECT 
                                    AGENTID, 
                                    LOGINDATE 
                                FROM 
                                    TBL_LOGIN 
                                WHERE 
                                    LOGINDATE > DATEADD(DAY, -100, GETDATE())
                                    AND Loginby IS NULL
                                    AND AGENTID NOT IN (SELECT Emailid FROM Agent_Emails)
                                    AND AGENTID NOT IN (SELECT UserName FROM Agent_Usernames)
                                ORDER BY 
                                    LOGINDATE DESC;
                               """)

                header = [desc[0] for desc in cursor.description]
                data = cursor.fetchall()
                data_tuples = [tuple(row) for row in data]
                df_login = pd.DataFrame(data_tuples, columns=header)

            df_login = df_login.dropna(subset=['LOGINDATE'])
            df_login['LOGINDATE'] = pd.to_datetime(df_login['LOGINDATE'], errors='coerce').dt.date

            df_login['AGENTID'] = df_login['AGENTID'].str.lower().str.strip()
            df_login = df_login[df_login['AGENTID'].notna() & (df_login['AGENTID'] != '')]
            df_login = df_login.drop_duplicates(subset=['AGENTID','LOGINDATE'])

            current_date = datetime.now().date()

            filter_date = int(request.GET.get('filter_date',30))

            cutoff_date = (datetime.now() - timedelta(days=filter_date)).date()

            df_recent = df_login[df_login['LOGINDATE'] >= cutoff_date]

            login_counts = (
                df_recent.groupby('AGENTID')
                .size()
                .reset_index(name='Login_Count')
                .sort_values(by='Login_Count', ascending=False)
            )

            return JsonResponse(login_counts.to_dict(orient='records'),safe=False)
        
        except Exception as e:
                print("Error : ", e)
                return ("Error")
        

    def login_details(request):
        print("api/login-details")
        now = datetime.now()
        current_year, current_month = now.year, now.month

        # Get previous month and year
        if current_month == 1:
            prev_month, prev_year = 12, current_year - 1
        else:
            prev_month, prev_year = current_month - 1, current_year

        try:
            # Optimized SQL query - do filtering and aggregation in database
            with connection.cursor() as cursor:
                # Single optimized query that does most of the work in SQL (SQL Server syntax)
                cursor.execute("""
                    ;WITH cleaned_data AS (
                        SELECT 
                            UPPER(LTRIM(RTRIM(AGENTID))) as agent_id,
                            LOGINDATE as login_date,
                            YEAR(LOGINDATE) as login_year,
                            MONTH(LOGINDATE) as login_month,
                            CAST(LOGINDATE AS DATE) as login_day
                        FROM TBL_LOGIN 
                        WHERE AGENTID IS NOT NULL 
                        AND LOGINDATE IS NOT NULL
                        AND Loginby IS NULL
                        AND LOGINDATE >= DATEADD(MONTH, -25, GETDATE())
                        AND AgentID NOT In (select distinct (Agentid) as Agentid from tbl_agent_culture)
                    ),
                    dedupe_data AS (
                        SELECT 
                            agent_id,
                            login_date,
                            login_year,
                            login_month,
                            login_day,
                            ROW_NUMBER() OVER (
                                PARTITION BY agent_id, 
                                CAST(DATEDIFF(MINUTE, '1900-01-01', login_date) AS BIGINT)
                                ORDER BY login_date
                            ) as rn
                        FROM cleaned_data
                    ),
                    valid_logins AS (
                        SELECT agent_id, login_date, login_year, login_month, login_day 
                        FROM dedupe_data 
                        WHERE rn = 1
                    ),
                    daily_stats AS (
                        SELECT 
                            login_year, 
                            login_month, 
                            login_day,
                            COUNT(*) as daily_logins,
                            COUNT(DISTINCT agent_id) as daily_agents
                        FROM valid_logins
                        GROUP BY login_year, login_month, login_day
                    )
                    SELECT 
                        -- Current month metrics
                        COUNT(CASE WHEN login_year = %s AND login_month = %s THEN 1 END) as current_total_logins,
                        COUNT(DISTINCT CASE WHEN login_year = %s AND login_month = %s THEN agent_id END) as current_unique_agents,
                               
                        -- Previous month metrics  
                        COUNT(CASE WHEN login_year = %s AND login_month = %s THEN 1 END) as prev_total_logins,
                        COUNT(DISTINCT CASE WHEN login_year = %s AND login_month = %s THEN agent_id END) as prev_unique_agents
                    FROM valid_logins
                """, [
                    current_year, current_month,  # current total
                    current_year, current_month,  # current unique
                    prev_year, prev_month,        # prev total
                    prev_year, prev_month         # prev unique
                ])
                
                metrics = cursor.fetchone()
                
                # Get daily averages separately (simpler approach)
                cursor.execute("""
                    ;WITH cleaned_data AS (
                        SELECT 
                            UPPER(LTRIM(RTRIM(AGENTID))) as agent_id,
                            LOGINDATE as login_date,
                            YEAR(LOGINDATE) as login_year,
                            MONTH(LOGINDATE) as login_month,
                            CAST(LOGINDATE AS DATE) as login_day
                        FROM TBL_LOGIN 
                        WHERE AGENTID IS NOT NULL 
                        AND Loginby IS NULL
                        AND LOGINDATE IS NOT NULL
                        AND (
                            (YEAR(LOGINDATE) = %s AND MONTH(LOGINDATE) = %s) OR
                            (YEAR(LOGINDATE) = %s AND MONTH(LOGINDATE) = %s)
                        )
                        AND AgentID NOT In (select distinct (Agentid) as Agentid from tbl_agent_culture)
                    ),
                    dedupe_data AS (
                        SELECT 
                            agent_id,
                            login_date,
                            login_year,
                            login_month,
                            login_day,
                            ROW_NUMBER() OVER (
                                PARTITION BY agent_id, 
                                CAST(DATEDIFF(MINUTE, '1900-01-01', login_date) AS BIGINT)
                                ORDER BY login_date
                            ) as rn
                        FROM cleaned_data
                    ),
                    valid_logins AS (
                        SELECT agent_id, login_date, login_year, login_month, login_day 
                        FROM dedupe_data 
                        WHERE rn = 1
                    ),
                    daily_stats AS (
                        SELECT 
                            login_year, 
                            login_month, 
                            login_day,
                            COUNT(*) as daily_logins,
                            COUNT(DISTINCT agent_id) as daily_agents
                        FROM valid_logins
                        GROUP BY login_year, login_month, login_day
                    )
                    SELECT 
                        login_year,
                        login_month,
                        AVG(CAST(daily_logins AS FLOAT) / daily_agents) as avg_logins_per_agent
                    FROM daily_stats
                    WHERE daily_agents > 0
                    GROUP BY login_year, login_month
                """, [
                    current_year, current_month,  # current month
                    prev_year, prev_month      # prev month
                ])
                
                executed_query = connection.queries[-1]['sql'] if connection.queries else "No query executed"
                print(executed_query)
                
                avg_results = cursor.fetchall()

                cursor.execute("""
                    SELECT 
                        login_year as Year,
                        COUNT(DISTINCT agent_id) as Login_Count
                    FROM (
                        SELECT 
                            UPPER(LTRIM(RTRIM(AGENTID))) as agent_id,
                            YEAR(LOGINDATE) as login_year
                        FROM TBL_LOGIN 
                        WHERE AGENTID IS NOT NULL 
                        AND Loginby IS NULL
                        AND LOGINDATE IS NOT NULL
                        AND Loginby IS NULL
                        AND AgentID NOT In (select distinct (Agentid) as Agentid from tbl_agent_culture)	
                    ) yearly_data
                    GROUP BY login_year
                    ORDER BY login_year
                """)
                
                yearly_data = cursor.fetchall()

            # Process results
            (current_logins, current_unique_agents, previous_logins, previous_unique_agents) = metrics
            
            # Process daily averages
            current_avg = 0
            previous_avg = 0
            
            for year, month, avg_val in avg_results:
                if year == current_year and month == current_month:
                    current_avg = avg_val or 0
                elif year == prev_year and month == prev_month:
                    previous_avg = avg_val or 0

            # Calculate percentage changes
            def safe_percentage_change(current, previous):
                if previous == 0:
                    return 100.0 if current > 0 else 0.0
                return ((current - previous) / previous) * 100

            logins_percentage_change = safe_percentage_change(current_logins, previous_logins)
            unique_percentage_change = safe_percentage_change(current_unique_agents, previous_unique_agents)
            avg_percentage_change = safe_percentage_change(current_avg, previous_avg)

            # Format yearly data
            login_counts_list = [{'Year': int(year), 'Login_Count': int(count)} 
                            for year, count in yearly_data]

            # Build response
            response_data = {
                'login_counts_per_year': login_counts_list,
                'agent_per_day_login_avg': round(float(current_avg or 0), 2),
                'avg_percentage_change': round(float(avg_percentage_change), 2),
                'total_logins': int(current_logins or 0),
                'logins_percentage_change': round(float(logins_percentage_change), 2),
                'total_unique_agents': int(current_unique_agents or 0),
                'unique_percentage_change': round(float(unique_percentage_change), 2),
            }

            return JsonResponse(response_data, safe=False)

        except Exception as e:
            return JsonResponse({'error': str(e), 'status': 'failed'}, safe=False, status=500)


    def most_searched_tour(request): #Need to add more things
        print('HolidaysSearch')

        try:
            with connection.cursor() as cursor:
                cursor.execute(f"""
                                select pkgID, CreatedDate, AgentId, Fare
                                from Tbl_HolidaysSearch
                                where fare is not Null 
                                    AND Fare > 0
                                    AND AgentId is Not NULL 
                                    AND AgentId != ''
                                    AND pkgID > 0
                                    AND CreatedDate is not Null
                                    AND CreatedDate != ''
                                """)

                header = [desc[0] for desc in cursor.description]
                data = cursor.fetchall()
                data_tuples = [tuple(row) for row in data]
                df_holidayssearch = pd.DataFrame(data_tuples, columns=header)

            df_holidayssearch['CreatedDate'] = df_holidayssearch['CreatedDate'].dt.date

            df_holidayssearch = df_holidayssearch.drop_duplicates(subset=['pkgID', 'CreatedDate', 'AgentId'])

            most_frequent_pkgids = (
                df_holidayssearch.groupby('CreatedDate')['pkgID']
                .agg(lambda x: x.value_counts().idxmax())
                .reset_index()
            )

            most_frequent_pkgids.columns = ['CreatedDate', 'pkgID']

            with connection.cursor() as cursor:
                cursor.execute("SELECT PKG_ID as pkgID, PKG_TITLE FROM TBL_PKG_DETAILS")
                pkg_header = [desc[0] for desc in cursor.description]
                pkg_data = cursor.fetchall()
                pkg_data = [tuple(row) for row in pkg_data]
                df_pkg = pd.DataFrame(pkg_data, columns=pkg_header)

            # Merge to get pkgName
            merged = pd.merge(most_frequent_pkgids, df_pkg, on='pkgID', how='left')

            merged = merged.sort_values(by='CreatedDate', ascending=False)

            return JsonResponse(merged.to_dict(orient='records'), safe=False)

        except Exception as e:
            print("Error : ", e)
            return JsonResponse({"error": str(e)}, status=500)

    def yearly_searched_tour(request):
        print('api/yearly-searched-tour')
        try:
        
            with connection.cursor() as cursor:
                cursor.execute('''select PKG_ID, PKG_TITLE 
                               from TBL_PKG_DETAILS
                                where Status = 1
                                AND AGENTID = ''
                                AND AGENTID not in (select distinct (Agentid) as Agentid from tbl_agent_culture)
                                ''')
                
                pkg_header = [desc[0] for desc in cursor.description]
                pkg_data = cursor.fetchall()
                df_pkg = pd.DataFrame(pkg_data, columns=pkg_header)
                df_pkg['PKG_ID'] = df_pkg['PKG_ID'].astype(str)  # Ensure string type for safe merging

            # --- Most Searched Tours ---
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT pkgid, CreatedDate, Agentid, Fare 
                    FROM Tbl_HolidaysSearch
                    WHERE AgentId IS NOT NULL AND AgentId != ''
                        AND pkgID > 0
                        AND CreatedDate IS NOT NULL AND CreatedDate != ''
                        AND Fare > 0
                        AND YEAR(CONVERT(date, CreatedDate, 103)) > (YEAR(GETDATE()) - 3)
                """)
                header = [desc[0] for desc in cursor.description]
                data = cursor.fetchall()
                df_holidays = pd.DataFrame(data, columns=header)

            df_holidays['CreatedDate'] = pd.to_datetime(df_holidays['CreatedDate'], errors='coerce')
            df_holidays['tour_search_year'] = df_holidays['CreatedDate'].dt.year
            df_holidays['pkgid'] = df_holidays['pkgid'].astype(str)

            most_searched = (
                df_holidays.groupby('tour_search_year')['pkgid']
                .agg(lambda x: x.value_counts().idxmax())
                .reset_index()
            )
            most_searched = most_searched.merge(df_pkg, left_on='pkgid', right_on='PKG_ID', how='left')
            most_searched = most_searched[['tour_search_year', 'PKG_TITLE']]
            most_searched.columns = ['year', 'most_searched_tour']

            # --- Most Booked Tours ---
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT CONCAT(trim(upper(AgentID)), PackgID, convert(date,tourdate,103)) AS UniqueID,
                       convert(date,tourdate,103) as tourdate, convert(date,CreatedDate,103) as CreatedDate, PackgID, trim(upper(AgentID)) as AgentID
                    FROM TBL_BOOKING
                    WHERE txn_msg = 'success'
                        AND (PayMode = 'Deposit' OR PayMode = 'Full')
                        AND TRY_CONVERT(DECIMAL(10,2), USDamt) >= 100.00
                        AND tourdate IS NOT NULL
                        AND PackgID IS NOT NULL AND PackgID != ''
                        AND PackgID > 0
                        AND Is_cancelled <> 1
                        AND AgentId IS NOT NULL
                        AND AgentId != ''
                        AND CreatedDate IS NOT NULL
                        AND YEAR(CONVERT(date, CreatedDate, 103)) > (YEAR(GETDATE()) - 3)
                        AND AgentId NOT IN (SELECT DISTINCT (Agentid) AS Agentid FROM tbl_agent_culture)
                """)
                header = [desc[0] for desc in cursor.description]
                data = cursor.fetchall()
                df_booking = pd.DataFrame(data, columns=header)

            df_booking = df_booking.drop_duplicates(subset=['UniqueID'])
            df_booking['CreatedDate'] = pd.to_datetime(df_booking['CreatedDate'], errors='coerce')
            df_booking = df_booking.sort_values('CreatedDate').drop_duplicates(subset=['UniqueID'], keep='first')

            df_booking['Created_year'] = df_booking['CreatedDate'].dt.year
            df_booking['PackgID'] = df_booking['PackgID'].astype(str)

            most_booked = (
                df_booking.groupby('Created_year')['PackgID']
                .agg(lambda x: x.value_counts().idxmax())
                .reset_index()
            )
            most_booked = most_booked.merge(df_pkg, left_on='PackgID', right_on='PKG_ID', how='left')
            most_booked = most_booked[['Created_year', 'PKG_TITLE']]
            most_booked.columns = ['year', 'most_booked_tour']


            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT country, YEAR(TRY_CONVERT(date, CreatedDate, 103)) AS CreatedYear
                    FROM Tbl_HolidaysSearch
                    WHERE country IS NOT NULL 
                        AND country != ''
                        AND YEAR(CONVERT(date, CreatedDate, 103)) > (YEAR(GETDATE()) - 3)
                    ORDER BY SRID
                """)
                data = cursor.fetchall()
                columns = [col[0] for col in cursor.description]

            # Step 1: Create DataFrame
            df = pd.DataFrame(data, columns=columns)

            # Step 2: Ensure country is a string, then split and explode
            df['country'] = df['country'].astype(str).str.split(',')
            df = df.explode('country')

            # Step 3: Clean country values
            df['country'] = df['country'].str.strip()
            df = df[df['country'].str.match(r'^[A-Za-z\s]+$')]  # Optional: filter alphabetic only

            # Step 4: Rename columns
            df.rename(columns={'CreatedYear': 'year', 'country': 'most_search_Country'}, inplace=True)

            # Step 5: Group and count
            search_counts = df.groupby(['year', 'most_search_Country']).size().reset_index(name='No_of_Searches')

            # Step 6: Get most searched country per year
            most_searched_country = (
                search_counts.sort_values(['year', 'No_of_Searches'], ascending=[True, False])
                .groupby('year')
                .first()
                .reset_index()
            )
            most_searched_country = most_searched_country.drop(columns=['No_of_Searches'])

            # Output
            print(most_searched_country)



            # --- Final Merge ---
            final_df = pd.merge(most_searched, most_booked, on='year', how='outer')
            final_df = pd.merge(final_df, most_searched_country, on='year', how='outer')

            final_df = final_df.sort_values('year').fillna('')

            return JsonResponse(final_df.to_dict(orient='records'), safe=False)

        except Exception as e:
            print("Error:", e)
            return JsonResponse({"error": str(e)}, status=500)


    def monthly_searched_tour(request):
        print('api/monthly-searched-tour')
        year = request.GET.get('year')

        if not year:
            return JsonResponse({'error': 'Parameter is required'}, status=400)

        try:
            with connection.cursor() as cursor:
                cursor.execute('''select PKG_ID, PKG_TITLE 
                               from TBL_PKG_DETAILS
                                where Status = 1
                                AND AGENTID = ''
                                AND AGENTID not in (select distinct (Agentid) as Agentid from tbl_agent_culture)
                                ''')
                
                pkg_header = [desc[0] for desc in cursor.description]
                pkg_data = cursor.fetchall()
                df_pkg = pd.DataFrame(pkg_data, columns=pkg_header)
                df_pkg['PKG_ID'] = df_pkg['PKG_ID'].astype(str) 

            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT pkgid, CreatedDate, Agentid, Fare 
                    FROM Tbl_HolidaysSearch
                    WHERE AgentId IS NOT NULL AND AgentId != ''
                        AND pkgID > 0
                        AND CreatedDate IS NOT NULL AND CreatedDate != ''
                        AND Fare > 0
                        AND YEAR(CONVERT(date, CreatedDate, 103)) = %s
                """,[year])
                header = [desc[0] for desc in cursor.description]
                data = cursor.fetchall()
                df_holidays = pd.DataFrame(data, columns=header)

            df_holidays['CreatedDate'] = pd.to_datetime(df_holidays['CreatedDate'], errors='coerce')
            df_holidays['tour_search_month'] = df_holidays['CreatedDate'].dt.month_name()
            # Add numeric month for sorting
            df_holidays['month_num'] = df_holidays['CreatedDate'].dt.month
            df_holidays['pkgid'] = df_holidays['pkgid'].astype(str)

            most_searched = (
                df_holidays.groupby('tour_search_month')['pkgid']
                .agg(lambda x: x.value_counts().idxmax())
                .reset_index()
            )
            
            # Add month number to most_searched dataframe
            month_to_num = {month: i for i, month in enumerate(pd.date_range(start=f'2023-01-01', periods=12, freq='ME').strftime('%B'), 1)}
            most_searched['month_num'] = most_searched['tour_search_month'].map(month_to_num)

            most_searched = most_searched.merge(df_pkg, left_on='pkgid', right_on='PKG_ID', how='left')
            most_searched = most_searched[['tour_search_month', 'PKG_TITLE', 'month_num']]
            most_searched.columns = ['month', 'most_searched_tour', 'month_num']


            #--- Most Booked Tours ---
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT CONCAT(trim(upper(AgentID)), PackgID, convert(date,tourdate,103)) AS UniqueID,
                       convert(date,tourdate,103) as tourdate, convert(date,CreatedDate,103) as CreatedDate, PackgID, trim(upper(AgentID)) as AgentID
                    FROM TBL_BOOKING
                    WHERE txn_msg = 'success'
                        AND (PayMode = 'Deposit' OR PayMode = 'Full')
                        AND TRY_CONVERT(DECIMAL(10,2), USDamt) >= 100.00
                        AND tourdate IS NOT NULL
                        AND PackgID IS NOT NULL AND PackgID != ''
                        AND PackgID > 0
                        AND Is_cancelled <> 1
                        AND AgentId IS NOT NULL
                        AND AgentId is != ''
                        AND CreatedDate IS NOT NULL
                        AND YEAR(CONVERT(date, CreatedDate, 103)) = %s
                        AND AgentId NOT IN (SELECT DISTINCT (Agentid) AS Agentid FROM tbl_agent_culture)
                """,[year])

                header = [desc[0] for desc in cursor.description]
                data = cursor.fetchall()
                df_booking = pd.DataFrame(data, columns=header)

            df_booking = df_booking.drop_duplicates(subset=['UniqueID'])
            df_booking['CreatedDate'] = pd.to_datetime(df_booking['CreatedDate'], errors='coerce')
            df_booking = df_booking.sort_values('CreatedDate').drop_duplicates(subset=['UniqueID'], keep='first')

            df_booking['Created_month'] = df_booking['CreatedDate'].dt.month_name()
            # Add numeric month for sorting
            df_booking['month_num'] = df_booking['CreatedDate'].dt.month
            df_booking['PackgID'] = df_booking['PackgID'].astype(str)

            most_booked = (
                df_booking.groupby('Created_month')['PackgID']
                .agg(lambda x: x.value_counts().idxmax())
                .reset_index()
            )
            
            # Add month number to most_booked dataframe
            most_booked['month_num'] = most_booked['Created_month'].map(month_to_num)

            most_booked = most_booked.merge(df_pkg, left_on='PackgID', right_on='PKG_ID', how='left')
            most_booked = most_booked[['Created_month', 'PKG_TITLE', 'month_num']]
            most_booked.columns = ['month', 'most_booked_tour', 'month_num']

            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT country, month(TRY_CONVERT(date, CreatedDate, 103)) AS created_month
                    FROM Tbl_HolidaysSearch
                    WHERE country IS NOT NULL 
                        AND country != ''
                        AND YEAR(CONVERT(date, CreatedDate, 103)) = %s
                    ORDER BY SRID
                """,[year])

                data = cursor.fetchall()
                columns = [col[0] for col in cursor.description]

            # Step 1: Create DataFrame
            df = pd.DataFrame(data, columns=columns)


            # Step 2: Ensure country is a string, then split and explode
            df['country'] = df['country'].astype(str).str.split(',')
            df = df.explode('country')

            # Step 3: Clean country values
            df['country'] = df['country'].str.strip()
            df = df[df['country'].str.match(r'^[A-Za-z\s]+$')]  # Optional: filter alphabetic only

            # Step 4: Rename columns
            df.rename(columns={'created_month': 'created_month', 'country': 'most_search_Country'}, inplace=True)

            # Step 5: Group and count
            search_counts = df.groupby(['created_month', 'most_search_Country']).size().reset_index(name='No_of_Searches')

            # Step 6: Get most searched country per year
            most_searched_country = (
                search_counts.sort_values(['created_month', 'No_of_Searches'], ascending=[True, False])
                .groupby('created_month')
                .first()
                .reset_index()
            )

            most_searched_country = most_searched_country.drop(columns=['No_of_Searches'])
            most_searched_country['month_num'] = most_searched_country['created_month']
            most_searched_country['month'] = most_searched_country['month_num'].apply(lambda x: datetime(1900, x, 1).strftime('%B'))
            most_searched_country = most_searched_country.drop(columns=['created_month'])

            # --- Final Merge ---
            final_df = pd.merge(most_searched, most_booked, on=['month', 'month_num'], how='outer')
            print(final_df) 

            final_df = pd.merge(final_df, most_searched_country, on=['month', 'month_num'], how='outer')
            print(final_df) 

            # Sort by month number instead of month name
            final_df = final_df.sort_values('month_num').fillna('')
            
            # Remove the month_num column before returning the result
            final_df = final_df[['month', 'most_searched_tour', 'most_booked_tour', 'most_search_Country']]

            return JsonResponse(final_df.to_dict(orient='records'), safe=False)

        except Exception as e:
            print("Error:", e)
            return JsonResponse({"error": str(e)}, status=500)


    def daily_searched_tour(request):
        print('api/daily-searched-tour')
        year = request.GET.get('year')
        month = request.GET.get('month')

        if not year or not month:
            return JsonResponse({'error': 'Parameter is required'}, status=400)
        
        if month:
            month_number = datetime.strptime(month, "%B").month
            month = month_number

        try:
            # Fetch Package Titles only once
            with connection.cursor() as cursor:
                cursor.execute('''select PKG_ID, PKG_TITLE 
                               from TBL_PKG_DETAILS
                                where Status = 1
                                AND AGENTID = ''
                                AND AGENTID not in (select distinct (Agentid) as Agentid from tbl_agent_culture)
                                ''')
                
                pkg_header = [desc[0] for desc in cursor.description]
                pkg_data = cursor.fetchall()
                df_pkg = pd.DataFrame(pkg_data, columns=pkg_header)
                df_pkg['PKG_ID'] = df_pkg['PKG_ID'].astype(str)  # Ensure string type for safe merging

            # --- Most Searched Tours ---
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT pkgid, CreatedDate, Agentid, Fare 
                    FROM Tbl_HolidaysSearch
                    WHERE AgentId IS NOT NULL AND AgentId != ''
                        AND pkgID > 0
                        AND CreatedDate IS NOT NULL AND CreatedDate != ''
                        AND Fare > 0
                        AND YEAR(CONVERT(date, CreatedDate, 103)) = %s
                        AND MONTH(CONVERT(date, CreatedDate, 103)) = %s
                """,[year,month])

                header = [desc[0] for desc in cursor.description]
                data = cursor.fetchall()
                df_holidays = pd.DataFrame(data, columns=header)

            df_holidays['CreatedDate'] = pd.to_datetime(df_holidays['CreatedDate'], errors='coerce')
            df_holidays['tour_search_date'] = df_holidays['CreatedDate'].dt.date
            df_holidays['pkgid'] = df_holidays['pkgid'].astype(str)

            most_searched = (
                df_holidays.groupby('tour_search_date')['pkgid']
                .agg(lambda x: x.value_counts().idxmax())
                .reset_index()
            )
            most_searched = most_searched.merge(df_pkg, left_on='pkgid', right_on='PKG_ID', how='left')
            most_searched = most_searched[['tour_search_date', 'PKG_TITLE']]
            most_searched.columns = ['date', 'most_searched_tour']


            # --- Most Booked Tours ---
            with connection.cursor() as cursor:
                cursor.execute("""
                        SELECT CONCAT(trim(upper(AgentID)), PackgID, convert(date,tourdate,103)) AS UniqueID,
                       convert(date,tourdate,103) as tourdate, convert(date,CreatedDate,103) as CreatedDate, PackgID, trim(upper(AgentID)) as AgentID
                    FROM TBL_BOOKING
                    WHERE txn_msg = 'success'
                        AND (PayMode = 'Deposit' OR PayMode = 'Full')
                        AND TRY_CONVERT(DECIMAL(10,2), USDamt) >= 100.00
                        AND tourdate IS NOT NULL
                        AND PackgID IS NOT NULL AND PackgID != ''
                        AND PackgID > 0
                        AND Is_cancelled <> 1
                        AND AgentId IS NOT NULL
                        AND AgentId != ''
                        AND CreatedDate IS NOT NULL
                        AND agentid not in (select distinct (Agentid) as Agentid from tbl_agent_culture)
                        AND YEAR(CONVERT(date, CreatedDate, 103)) = %s
                        AND MONTH(CONVERT(date, CreatedDate, 103)) = %s
                """,[year, month])

                header = [desc[0] for desc in cursor.description]
                data = cursor.fetchall()
                df_booking = pd.DataFrame(data, columns=header)

            df_booking = df_booking.drop_duplicates(subset=['UniqueID'])
            df_booking['CreatedDate'] = pd.to_datetime(df_booking['CreatedDate'], errors='coerce')
            df_booking = df_booking.sort_values('CreatedDate').drop_duplicates(subset=['UniqueID'], keep='first')

            df_booking['Created_date'] = df_booking['CreatedDate'].dt.date
            df_booking['PackgID'] = df_booking['PackgID'].astype(str)

            most_booked = (
                df_booking.groupby('Created_date')['PackgID']
                .agg(lambda x: x.value_counts().idxmax())
                .reset_index()
            )
            most_booked = most_booked.merge(df_pkg, left_on='PackgID', right_on='PKG_ID', how='left')
            most_booked = most_booked[['Created_date', 'PKG_TITLE']]
            most_booked.columns = ['date', 'most_booked_tour']

            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT country, TRY_CONVERT(date, CreatedDate, 103) AS CreatedDate
                    FROM Tbl_HolidaysSearch
                    WHERE country IS NOT NULL 
                        AND country != ''
                        AND YEAR(CONVERT(date, CreatedDate, 103)) = %s
                        AND MONTH(CONVERT(date, CreatedDate, 103)) = %s
                    ORDER BY SRID
                """,[year, month])
                
                data = cursor.fetchall()
                columns = [col[0] for col in cursor.description]

            # Step 1: Create DataFrame
            df = pd.DataFrame(data, columns=columns)

            # Step 2: Ensure country is a string, then split and explode
            df['country'] = df['country'].astype(str).str.split(',')
            df = df.explode('country')

            # Step 3: Clean country values
            df['country'] = df['country'].str.strip()
            df = df[df['country'].str.match(r'^[A-Za-z\s]+$')]  # Optional: filter alphabetic only

            # Step 4: Rename columns
            df.rename(columns={'CreatedDate': 'date', 'country': 'most_search_Country'}, inplace=True)

            # Step 5: Group and count
            search_counts = df.groupby(['date', 'most_search_Country']).size().reset_index(name='No_of_Searches')

            # Step 6: Get most searched country per year
            most_searched_country = (
                search_counts.sort_values(['date', 'No_of_Searches'], ascending=[True, False])
                .groupby('date')
                .first()
                .reset_index()
            )
            most_searched_country = most_searched_country.drop(columns=['No_of_Searches'])

            # Output
            print(most_searched_country)

            # --- Final Merge ---
            final_df = pd.merge(most_searched, most_booked, on='date', how='outer')

            final_df = pd.merge(final_df, most_searched_country, on='date', how='outer')

            final_df = final_df.sort_values('date').fillna('')

            return JsonResponse(final_df.to_dict(orient='records'), safe=False)


        except Exception as e:
            print("Error:", e)
            return JsonResponse({"error": str(e)}, status=500)
        

    def overall_country_search(request):
        selected_country = request.GET.get('country')

        print(selected_country)

        try:
            if selected_country:
                selected_country = selected_country.upper()
                if selected_country == 'UNITED ARAB EMIRATES':
                    selected_country = 'DUBAI'

                print(selected_country)
                with connection.cursor() as cursor:
                    cursor.execute(f"""SELECT 
                                        b.AgentId, 
                                        b.CreatedDate, 
                                        UPPER(TRIM(a.PKG_TITLE)) AS pkgName, 
                                        UPPER(TRIM(
											CASE 
												WHEN a.Country = 'Bali' THEN 'Indonesia' 
												ELSE a.Country 
											END
										)) AS CountryName
                                    FROM
                                        Tbl_HolidaysSearch b
                                    INNER JOIN 
                                        TBL_PKG_DETAILS a ON a.PKG_ID = b.pkgID
                                    WHERE 
                                        a.Status = 1
                                        AND b.AgentId IS NOT NULL
                                        AND b.AgentId != ''
                                        AND ',' + UPPER(TRIM(
											CASE 
												WHEN a.Country = 'Bali' THEN 'Indonesia' 
												ELSE a.Country 
											END
										)) + ',' LIKE %s
                                        """,[f'%,{selected_country},%'])
                    
                    executed_query = connection.queries[-1]['sql'] if connection.queries else "No query executed"
                
                    print(executed_query)
                    
                    data = cursor.fetchall()
                    columns = [row[0] for row in cursor.description]

                df_country = pd.DataFrame(data, columns=columns)    

                #print(df_country.head())

                df_country['CreatedDate'] = pd.to_datetime(df_country['CreatedDate'], format='%d/%m/%Y', errors='coerce')
                df_country['CountryName'] = df_country['CountryName'].str.split(r'\s*,\s*')
                df_country = df_country.explode('CountryName')
                df_country['CountryName'] = df_country['CountryName'].str.strip().str.upper()

                df_country['CountryName'] = df_country['CountryName'].replace({
                    'BALI': 'INDONESIA',
                    'DUBAI': 'UNITED ARAB EMIRATES'
                })

                df_country.sort_values(by=['AgentId', 'pkgName', 'CountryName', 'CreatedDate'], ascending=[True, True, True, True], inplace=True)

                df_country['time_diff'] = df_country.groupby(['AgentId', 'pkgName', 'CountryName'])['CreatedDate'].diff(-1).abs()

                # Keep rows where time_diff is either NaT or >= 1 minute
                filtered_country_df = df_country[(df_country['time_diff'].isna()) | (df_country['time_diff'] >= pd.Timedelta(minutes=1))].copy()
                filtered_country_df.drop(columns='time_diff', inplace=True)

                # Remove bad country entries
                filtered_country_df = filtered_country_df[filtered_country_df['CountryName'].notnull()]
                filtered_country_df = filtered_country_df[filtered_country_df['CountryName'].str.strip().ne("")]
                filtered_country_df = filtered_country_df[~filtered_country_df['CountryName'].isin(["0", "1", "NA", "NULL"])]
                

                country_package_groupby = (
                    filtered_country_df.groupby('pkgName')['AgentId']
                    .count()
                    .reset_index()
                    .rename(columns={'pkgName': 'Package_Name', 'AgentId': 'No_of_Searches'})
                    .sort_values(by='No_of_Searches', ascending=False)
                    .head(3)
                )

                country_package_groupby = country_package_groupby.to_dict(orient='records')

            else:
                country_package_groupby = ''

            with connection.cursor() as cursor:
                cursor.execute(f"""
                            SELECT b.AgentId, b.CreatedDate, b.pkgID, b.SRID, UPPER(TRIM(a.Country)) as CountryName
                            FROM
                                Tbl_HolidaysSearch b
                            INNER JOIN 
                                TBL_PKG_DETAILS a ON a.PKG_ID = b.pkgID
                            WHERE 
                                a.Status = 1
                                AND b.AgentId IS NOT NULL
                                AND b.AgentId != ''
                            """)
                data = cursor.fetchall()
                columns = [row[0] for row in cursor.description]

            df = pd.DataFrame(data, columns=columns)
            df['CreatedDate'] = pd.to_datetime(df['CreatedDate'], format='%d/%m/%Y', errors='coerce')
            df['CountryName'] = df['CountryName'].str.split(r'\s*,\s*')
            df = df.explode('CountryName')
            df['CountryName'] = df['CountryName'].str.strip().str.upper()

            df['CountryName'] = df['CountryName'].replace({
                'BALI': 'INDONESIA',
                'DUBAI': 'UNITED ARAB EMIRATES'
            })


            df.sort_values(by=['AgentId', 'pkgID', 'CountryName', 'CreatedDate'], ascending=[True, True, True, True], inplace=True)

            # Identify rows to keep based on 1-minute difference
            df['time_diff'] = df.groupby(['AgentId', 'pkgID', 'CountryName'])['CreatedDate'].diff(-1).abs()

            
            # Keep rows where time_diff is either NaT or >= 1 minute
            filtered_df = df[(df['time_diff'].isna()) | (df['time_diff'] >= pd.Timedelta(minutes=1))].copy()

            # Drop the helper column
            filtered_df.drop(columns='time_diff', inplace=True)

            # Remove bad country entries
            filtered_df = filtered_df[filtered_df['CountryName'].notnull()]
            filtered_df = filtered_df[filtered_df['CountryName'].str.strip().ne("")]
            filtered_df = filtered_df[~filtered_df['CountryName'].isin(["0", "1", "NA", "NULL"])]

            country_groupby = (
                filtered_df.groupby('CountryName')['AgentId']
                .count()
                .reset_index()
                .rename(columns={'CountryName': 'Country', 'AgentId': 'No_of_Searches'})
            )

            country_groupby = country_groupby.to_dict(orient='records')

            response_data = {
                'country_groupby': country_groupby,
                'country_package_groupby': country_package_groupby
            }

            return JsonResponse(response_data, safe=False)
        
        except Exception as e:
            print("Error:", e)
            return JsonResponse({"error": str(e)}, status=500)


    def flyer_overview(request):
        print("flyer-download")

        try:
            with connection.cursor() as cursor:
                # Single optimized query that handles undefined AGENT_ID lookup
                cursor.execute("""
                    SELECT 
                        CASE 
                            WHEN af.AGENT_ID = 'undefined' THEN COALESCE(a.AgentID, 'undefined')
                            ELSE af.AGENT_ID 
                        END as AGENT_ID,
                        af.NAME,
                        af.EMAIL,
                        af.CREATED_DATE,
                        UPPER(LTRIM(RTRIM(af.FLYER_TITLE))) as FLYER_TITLE,
                        af.AMOUNT,
                        YEAR(CONVERT(DATE, af.CREATED_DATE, 103)) as flyer_created_year
                    FROM TBL_AGENT_FLYER af
                    LEFT JOIN tbl_agent a ON (af.AGENT_ID = 'undefined' AND af.EMAIL = a.UserName)
                    WHERE af.FLYER_TITLE IS NOT NULL
                        AND af.FLYER_TITLE != ''
                        AND af.AGENT_ID IS NOT NULL
                        AND af.AGENT_ID != ''
                        AND TRY_CONVERT(DECIMAL(10,2), af.AMOUNT) > 0.00
                        AND af.FLYER_STATUS = 1
                        AND YEAR(CONVERT(DATE, af.CREATED_DATE, 103)) >= (YEAR(GETDATE()) - 3)
                        
                    ORDER BY af.CREATED_DATE DESC
                """)
                
                # Fetch all data at once
                rows = cursor.fetchall()
                columns = [desc[0] for desc in cursor.description]
                
            # Process data without pandas for better performance
            processed_data = []
            
            for row in rows:
                row_dict = dict(zip(columns, row))
                
                if row_dict['AGENT_ID'] != 'undefined':
                    processed_data.append(row_dict)
            
            seen_combinations = set()
            unique_data = []
            
            for row in processed_data:
                key = (row['FLYER_TITLE'], row['AMOUNT'], row['AGENT_ID'])
                if key not in seen_combinations:
                    seen_combinations.add(key)
                    unique_data.append(row)
            
            # Use efficient data structures for counting
            from collections import defaultdict, Counter
            
            # Count flyers per year
            year_counts = Counter()
            
            # Count flyers per agent (storing name as well)
            agent_counts = defaultdict(lambda: {'NAME': '', 'count': 0})
            
            # Count flyer titles
            title_counts = Counter()
            
            # Single pass through data for all calculations
            for row in unique_data:
                year = row['flyer_created_year']
                agent_id = row['AGENT_ID']
                agent_name = row['NAME']
                flyer_title = row['FLYER_TITLE']
                
                # Count by year
                year_counts[year] += 1
                
                # Count by agent
                agent_counts[agent_id]['NAME'] = agent_name
                agent_counts[agent_id]['count'] += 1
                
                # Count by title
                title_counts[flyer_title] += 1
            
            flyer_counts_per_year = [
                {'Year': year, 'FlyerCount': count} 
                for year, count in sorted(year_counts.items())
            ]
            
            flyer_count_by_agent = [
                {'AGENT_ID': agent_id, 'NAME': data['NAME'], 'FlyerCount': data['count']}
                for agent_id, data in sorted(
                    agent_counts.items(), 
                    key=lambda x: x[1]['count'], 
                    reverse=True
                )[:50]  # Top 50 agents
            ]
            
            flyer_title_count = [
                {'FLYER_TITLE': title, 'FlyerCount': count}
                for title, count in title_counts.most_common(10)  # Top 10 titles
            ]
            
            final_result = {
                'flyer_counts_per_year': flyer_counts_per_year,
                'Agent_with_most_flyer_create': flyer_count_by_agent,
                'Most_Used_Flyer_Title': flyer_title_count,
            }
            
            return JsonResponse(final_result, safe=False)

        except Exception as e:
            print("Error:", e)
            return JsonResponse({"error": str(e)}, status=500)


    def booking_overviews(request):
        print("api/booking-details")
        try:
            current_year = datetime.now().year

            with connection.cursor() as cursor:
                cursor.execute(f"""
                    SELECT CONCAT(Upper(trim(AgentID)), PackgID, convert(date,tourdate,103)) AS UniqueID,
                        convert(date,tourdate,103) as tourdate, convert(date,CreatedDate,103) as CreatedDate, PackgID, trim(upper(AgentID)) as AgentID, PayMode
                    FROM TBL_BOOKING
                    WHERE txn_msg = 'success'
                        AND Is_cancelled <> 1       
                        AND CreatedDate IS NOT NULL
                        AND tourdate IS NOT NULL
                        AND AgentID IS NOT NULL
                        AND PackgID IS NOT NULL
                        AND agentId != ''
                        AND PackgID != ''
                        AND YEAR(CONVERT(date, CreatedDate, 103)) > 2022
                        AND YEAR(CONVERT(date, tourdate, 103)) > 2022
                        AND AgentID NOT IN (SELECT DISTINCT (Agentid) AS Agentid FROM tbl_agent_culture)
                    ORDER BY CreatedDate DESC
                """)
                header = [desc[0] for desc in cursor.description]
                data = cursor.fetchall()
                df_booking = pd.DataFrame(data, columns=header)

            # Convert to datetime
            df_booking['CreatedDate'] = pd.to_datetime(df_booking['CreatedDate'], format='%d/%m/%Y', errors='coerce')
            df_booking['tourdate'] = pd.to_datetime(df_booking['tourdate'], format='%d/%m/%Y', errors='coerce')

            # Find the minimum CreatedDate for each UniqueID
            df_booking = df_booking.sort_values('CreatedDate').drop_duplicates(subset=['UniqueID'], keep='first')

            # ---- Yearly CreatedDate Analysis ----
            df_booking['CreatedYear'] = df_booking['CreatedDate'].dt.year

            unique_counts_by_year = (
                df_booking.groupby('CreatedYear')['UniqueID']
                .nunique()  # Changed from count to nunique for consistency
                .reset_index()
                .rename(columns={'UniqueID': 'BookingCount'})
            )

            # ---- Monthly CreatedDate Analysis ----
            df_booking['CreatedMonthNum'] = df_booking['CreatedDate'].dt.month
            df_booking['CreatedMonthName'] = df_booking['CreatedDate'].dt.strftime('%B')  # Full month name

            # Group by year and month and count bookings
            monthly_created_counts = (
                df_booking.groupby(['CreatedYear', 'CreatedMonthNum', 'CreatedMonthName'])['UniqueID']
                .nunique() 
                .reset_index()
                .rename(columns={'UniqueID': 'BookingCount'})
                .sort_values(['CreatedYear', 'CreatedMonthNum'])  
            )

            # Format as requested
            monthly_created_by_year = []
            for year, group in monthly_created_counts.groupby('CreatedYear'):
                year_data = []
                for _, row in group.iterrows():
                    month_obj = {row['CreatedMonthName']: row['BookingCount']}
                    year_data.append(month_obj)
                monthly_created_by_year.append({str(year): year_data})

            return JsonResponse({
                'BookingCountByCreatedDateYear': unique_counts_by_year.to_dict(orient='records'),
                'BookingCountByCreatedDateMonth': monthly_created_by_year,
        
            }, safe=False)

        except Exception as e:
            print("Error:", e)
            return JsonResponse({"error": str(e)}, status=500)


    def booking_kpi(request):
        try:
            # Database query to fetch booking data
            with connection.cursor() as cursor:
                cursor.execute(f"""
                    SELECT CONCAT(Upper(trim(AgentID)), PackgID, convert(date,tourdate,103)) AS UniqueID,
                        convert(date,tourdate,103) as tourdate, convert(date,CreatedDate,103) as CreatedDate, PackgID, trim(upper(AgentID)) as agentId, PayMode             
                    FROM TBL_BOOKING
                    WHERE txn_msg = 'success'
                        AND Is_cancelled <> 1       
                        AND CreatedDate IS NOT NULL
                        AND tourdate IS NOT NULL
                        AND agentId IS NOT NULL
                        AND PackgID IS NOT NULL
                        AND agentId != ''
                        AND PackgID != ''
                        AND YEAR(CONVERT(date, CreatedDate, 103)) > 2022
                        AND YEAR(CONVERT(date, tourdate, 103)) > 2022
                        AND agentId NOT IN (SELECT DISTINCT (Agentid) AS Agentid FROM tbl_agent_culture)
                    ORDER BY CreatedDate DESC
                """)
                
                header = [desc[0] for desc in cursor.description]
                data = cursor.fetchall()
                df_booking = pd.DataFrame(data, columns=header)
            
            # Convert date strings to datetime objects
            df_booking['agentId'] = df_booking['agentId'].astype(str).str.strip()
            df_booking['CreatedDate'] = pd.to_datetime(df_booking['CreatedDate'], format='%d/%m/%Y', errors='coerce')
            df_booking['tourdate'] = pd.to_datetime(df_booking['tourdate'], format='%d/%m/%Y', errors='coerce')
            
            # Get earliest CreatedDate for each UniqueID
            min_created_df = df_booking.groupby('UniqueID')['CreatedDate'].min().reset_index()
            
            unique_bookings = pd.merge(
                min_created_df,
                df_booking[['UniqueID', 'tourdate']].drop_duplicates(),
                on='UniqueID'
            )
            
            current_year = datetime.now().year
            print(current_year)
            
            bookings_current_year = unique_bookings[unique_bookings['CreatedDate'].dt.year == current_year]
            bookings_current_year['BookingMonth'] = bookings_current_year['CreatedDate'].dt.month
            monthly_booking_counts = bookings_current_year.groupby('BookingMonth')['UniqueID'].count()
            monthly_booking_average = monthly_booking_counts.mean()
            
            print("Monthly Average Bookings:", monthly_booking_average)
    
            current_year_bookings = df_booking[df_booking['CreatedDate'].dt.year == current_year]

            unique_agents_count = current_year_bookings['agentId'].nunique()

            print("Unique Agents in Current Year:", unique_agents_count)

            return JsonResponse({
                'monthly_booking_average': monthly_booking_average,
                'unique_agents_count': unique_agents_count,
                #'monthly_tour_average': monthly_tour_average,
                #'monthly_tour_counts': monthly_tour_counts.to_dict()
            })

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)


    def agent_booking_report(request):
        print("api/agent-booking-report")

        current_year = datetime.now().year

        selected_year = int(request.GET.get('year'))
        selected_id = request.GET.get('agentId')
        end_year = current_year + 1

        target_years = list(range(selected_year + 1, end_year + 1))

        if selected_id :
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT trim(upper(agentId)) as agentId, PackgID, 
                            min(PackageName) as PackageName,
                            convert(date,tourdate,103) as tourdate 
                            FROM tbl_booking
                            WHERE txn_msg = 'success'
                                AND Is_cancelled <> 1       
                                AND CreatedDate IS NOT NULL
                                AND tourdate IS NOT NULL
                                AND agentId  IS NOT NULL
                                AND PackgID IS NOT NULL
                                AND PackgID != ''
                                AND CreatedDate != ''
                                AND agentId = %s
                                AND YEAR(CONVERT(date, tourdate, 103)) >= %s
                                AND agentId NOT In (select distinct (Agentid) as Agentid from tbl_agent_culture)
                            group by agentId, PackgID, tourdate
                            ORDER BY CONVERT(date, tourdate, 103) DESC
                    """,[selected_id, selected_year])
                
                header = [desc[0] for desc in cursor.description]
                data = cursor.fetchall()
                df_booking = pd.DataFrame(data, columns=header)
               
            return JsonResponse(df_booking.to_dict(orient='records'), safe=False)
        
        else:
            with connection.cursor() as cursor:
                cursor.execute("""
                                SELECT CONCAT(trim(upper(agentId)), PackgID, convert(date,tourdate,103)) AS UniqueID, trim(upper(agentId)) as agentId, PackageName, PackgID, convert(date,tourdate,103) as tourdate
                                    FROM tbl_booking
                                    WHERE txn_msg = 'success'
                                        AND Is_cancelled <> 1       
                                        AND CreatedDate IS NOT NULL
                                        AND tourdate IS NOT NULL
                                        AND agentId  IS NOT NULL
                                        AND PackgID IS NOT NULL
                                        AND CreatedDate != ''
                                        AND tourdate != ''
                                        AND agentId != ''
                                        AND PackgID != ''
                                        AND YEAR(CONVERT(date, tourdate, 103)) >= %s
                                        AND agentId NOT In (select distinct (Agentid) as Agentid from tbl_agent_culture)
                                    ORDER BY CONVERT(date, tourdate, 103) DESC
                                """,[selected_year])

                executed_query = connection.queries[-1]['sql'] if connection.queries else "No query executed"
                
                print(executed_query)

                header = [desc[0] for desc in cursor.description]
                data = cursor.fetchall()
                df_booking = pd.DataFrame(data, columns=header)

            df_booking = df_booking.drop_duplicates(subset=['UniqueID'])
            df_booking['agentId'] = df_booking['agentId'].astype(str).str.strip().str.upper()
            df_booking['tourdate'] = pd.to_datetime(df_booking['tourdate'], format='%d/%m/%Y', errors='coerce')
            df_booking['touryear'] = df_booking['tourdate'].dt.year

            filtered_df = df_booking[df_booking['touryear'] == selected_year]
            agents_in_selected_year = filtered_df['agentId'].unique()

            agent_year_matrix = []
            for agent in agents_in_selected_year:
                agent_data = {'agentId': agent}

                total = df_booking[
                    (df_booking['agentId'] == agent) & (df_booking['touryear'] >= selected_year)
                ]['UniqueID'].nunique()
                agent_data['TotalBookings'] = total

                for y in target_years:
                    has_booking = not df_booking[
                        (df_booking['agentId'] == agent) & (df_booking['touryear'] == y)
                    ].empty
                    agent_data['Booked for ' + str(y)] = 'Yes' if has_booking else 'No'
                agent_year_matrix.append(agent_data)

            agent_year_matrix = sorted(agent_year_matrix, key=lambda x: x['TotalBookings'], reverse=True)
                    
            df_result = pd.DataFrame(agent_year_matrix)

            agent_ids = df_result['agentId'].tolist()

            with connection.cursor() as cursor:
                format_strings = ','.join(['%s'] * len(agent_ids))
                cursor.execute(f"""
                    SELECT AgentID, Name, convert(date,LastLogin,103) as LastLogin 
                    FROM TBL_Agent 
                    WHERE AgentID IN ({format_strings})
                """, agent_ids)
                agent_details = cursor.fetchall()

            # Convert to DataFrame
            df_agent_info = pd.DataFrame(agent_details, columns=['AgentID', 'Name', 'LastLogin'])
            df_agent_info['AgentID'] = df_agent_info['AgentID'].astype(str).str.upper().str.strip()

            # Merge with result
            df_result['agentId'] = df_result['agentId'].astype(str).str.upper().str.strip()
            df_result = df_result.merge(df_agent_info, how='left', left_on='agentId', right_on='AgentID')
            df_result.drop(columns=['AgentID'], inplace=True)

            latest_tour_info = (
                df_booking.sort_values(['agentId', 'tourdate'], ascending=[True, False])
                .drop_duplicates(subset='agentId', keep='first')  # Latest tour per agent
                [['agentId', 'tourdate', 'PackageName']]
                .rename(columns={'tourdate': 'LatestTourDate', 'PackageName': 'LatestPackage'})
            )

            df_result = df_result.merge(latest_tour_info, how='left', on='agentId')
            df_result['LatestTourDate'] = df_result['LatestTourDate'].dt.strftime('%Y-%m-%d')

            df_result.fillna({'Name': '', 'LastLogin': '', 'Last TourDate': '', 'Last PackageName': ''}, inplace=True)

            print(df_result.head())

            with connection.cursor() as cursor:
                cursor.execute(f"""select TRAV_ID, CREATED_BY, convert(date,TourDate,103) as TourDate, Pkgid from TBL_TRAVELLER_NAME
                                    where CREATED_BY is not NULL
                                        AND CREATED_BY != ''
                                        AND Status1 = 'ACTIVE'
                                        AND TourDate is not Null
                                        AND TourDate != ''
                                        AND Pkgid is not Null
                                        AND Pkgid != ''
                                        """)

                header = [desc[0] for desc in cursor.description]   
                data = cursor.fetchall()
                df_traveller = pd.DataFrame(data, columns=header)
            
            df_traveller['CREATED_BY'] = df_traveller['CREATED_BY'].astype(str).str.strip().str.upper()

            traveller_counts = df_traveller.groupby('CREATED_BY')['TRAV_ID'].count().reset_index()
            traveller_counts.rename(columns={'CREATED_BY': 'agentId', 'TRAV_ID': 'TravellerCount'}, inplace=True)

            df_result = df_result.merge(traveller_counts, on='agentId', how='left')

            # Fill missing values with 0 for agents with no travellers
            df_result['TravellerCount'] = df_result['TravellerCount'].fillna(0).astype(int)

            print(df_result.head(12))

            # Optionally return as JSON
            return JsonResponse(df_result.to_dict(orient='records'), safe=False)


    def most_search_package(request):
        print('api/most-search-package')

        try:
            country_param = request.GET.get('country') 
            selected_country = country_param.strip().upper() if country_param else None

            year_param = request.GET.get('year')  
            selected_year = int(year_param) if year_param else None

            with connection.cursor() as cursor:
                cursor.execute(f"""SELECT 
                                    UPPER(TRIM(a.PKG_TITLE)) AS pkgName, 
                                    b.CreatedDate, 
                                    b.AgentId, 
                                    a.Country
                                FROM 
                                    Tbl_HolidaysSearch b
                                INNER JOIN 
                                    TBL_PKG_DETAILS a ON a.PKG_ID = b.pkgID
                                WHERE 
                                    a.Status = 1
                                    AND b.AgentId IS NOT NULL
                                    AND b.AgentId != ''
                                ORDER BY 
                                    b.AgentId;
                                    """)
                header = [desc[0] for desc in cursor.description]
                data = cursor.fetchall()
                df_package = pd.DataFrame(data, columns=header)
            
            df_package['CreatedDate'] = pd.to_datetime(df_package['CreatedDate'], format='%d/%m/%Y', errors='coerce')

            df_full_data = df_package.copy()

            if selected_year:
                df_package['CreatedYear'] = df_package['CreatedDate'].dt.year
                df_package = df_package[df_package['CreatedYear'] == selected_year].copy()

            df_package.sort_values(by=['AgentId', 'pkgName', 'CreatedDate'], ascending=[True, True, True], inplace=True)

            # Identify rows to keep based on 1-minute difference
            df_package['time_diff'] = df_package.groupby(['AgentId', 'pkgName'])['CreatedDate'].diff(-1).abs()

            # Keep rows where time_diff is either NaT or >= 1 minute
            filtered_df = df_package[(df_package['time_diff'].isna()) | (df_package['time_diff'] >= pd.Timedelta(minutes=1))].copy()

            # Drop the helper column
            filtered_df.drop(columns='time_diff', inplace=True)

            if selected_country:
                filtered_df['Country'] = filtered_df['Country'].str.split(r'\s*,\s*')
                filtered_df = filtered_df.explode('Country')
                filtered_df['Country'] = filtered_df['Country'].str.strip().str.upper()
                filtered_df['Country'] = filtered_df['Country'].replace({
                    'BALI': 'INDONESIA',
                })
                filtered_df = filtered_df[filtered_df['Country'] == selected_country]

            if filtered_df.empty:
                print("No data for the selected year.")
                return JsonResponse([], safe=False)
            
            top_10 = filtered_df['pkgName'].value_counts().head(10)
            top_pkg_names = top_10.index.tolist()

            current_year = datetime.now().year
            last_three_years = [current_year - 2, current_year - 1, current_year]

            full_df = df_full_data[df_full_data['pkgName'].isin(top_pkg_names)].copy()
            full_df['Year'] = full_df['CreatedDate'].dt.year
            full_df['Month'] = full_df['CreatedDate'].dt.strftime('%B')
            full_df = full_df[full_df['Year'].isin(last_three_years)]

            month_order = ['January', 'February', 'March', 'April', 'May', 'June',
                       'July', 'August', 'September', 'October', 'November', 'December']

            grouped = (
                full_df.groupby(['pkgName', 'Year', 'Month'])
                .size()
                .reset_index(name='count')
            )

            grouped['Month'] = pd.Categorical(grouped['Month'], categories=month_order, ordered=True)
            grouped = grouped.sort_values(by=['pkgName', 'Year', 'Month'])

            # --- STEP 3: Structure result ---
            result = []
            for pkg in top_pkg_names:
                pkg_data = grouped[grouped['pkgName'] == pkg]
                yearly_data = {}
                for year in last_three_years:
                    monthly_counts = {month: 0 for month in month_order}
                    year_data = pkg_data[pkg_data['Year'] == year]
                    for _, row in year_data.iterrows():
                        monthly_counts[row['Month']] = row['count']
                    yearly_data[str(year)] = monthly_counts

                result.append({
                    'pkgName': pkg,
                    'count': int(top_10[pkg]),
                    'monthly_data': yearly_data
                })

            return JsonResponse(result, safe=False)

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)   
    
        except Exception as e:
            print("Error:", e)
            return JsonResponse({"error": str(e)}, status=500)
    


    def agent_query(request):
        print("api/agent-query")

        format = request.GET.get("format")

        try:
            with connection.cursor() as cursor:
                cursor.execute(f"""
                        select
                            (select top 1 agentid from tbl_agent where Emailid=tm.EMAIL_ID or username=tm.EMAIL_ID ) as AGENT_ID, 
                            (select top 1 Emailid from tbl_agent where Emailid=tm.EMAIL_ID or username=tm.EMAIL_ID ) as EMAIL_ID,
                            (select top 1 Name from tbl_agent where Emailid=tm.EMAIL_ID or username=tm.EMAIL_ID ) as Name,
                            CREATED_DATE
                        from TBL_MESSAGE TM
                        where Year(convert(date,TM.CREATED_DATE,103)) >= (year(getdate())-3)
                            and (tm.EMAIL_ID in(select emailid from tbl_agent where Emailid=tm.EMAIL_ID ) or tm.EMAIL_ID in(select username from tbl_agent where username=tm.EMAIL_ID))
                        order by TM.MSG_ID desc
                """)
                header = [desc[0] for desc in cursor.description]
                data = cursor.fetchall()
                df_query = pd.DataFrame(data, columns=header)

            df_query = df_query[df_query['AGENT_ID'].notna() & (df_query['AGENT_ID'].str.strip() != '')]
            df_query['AGENT_ID'] = df_query['AGENT_ID'].astype(str).str.strip().str.upper()
            df_query['EMAIL_ID'] = df_query['EMAIL_ID'].astype(str).str.strip().str.lower()
            df_query['Name'] = df_query['Name'].astype(str).str.strip().str.title()

            df_query['CREATED_DATE'] = pd.to_datetime(df_query['CREATED_DATE'], errors='coerce')
            df_query['QUERY_YEAR'] = df_query['CREATED_DATE'].dt.year

            print(df_query)

            query_summary = (
                df_query.groupby(['AGENT_ID','EMAIL_ID', 'Name', 'QUERY_YEAR'])
                .size()
                .reset_index(name='TotalQuery')
                .sort_values(by='TotalQuery', ascending=False)
            )

            print(query_summary.head(5))

            with connection.cursor() as cursor:
                cursor.execute(f"""
                        SELECT CONCAT(trim(upper(agentId)), PackgID, convert(date,tourdate,103)) AS UniqueID, trim(upper(agentId)) as agentId
                            FROM tbl_booking
                            WHERE txn_msg = 'success'
                                    AND Is_cancelled <> 1       
                                    AND CreatedDate IS NOT NULL
                                    AND tourdate IS NOT NULL
                                    AND agentId  IS NOT NULL
                                    AND PackgID IS NOT NULL
                                    AND CreatedDate != ''
                                    AND tourdate != ''
                                    AND agentId != ''
                                    AND PackgID != ''
                                    AND agentid NOT IN (SELECT DISTINCT (Agentid) AS Agentid FROM tbl_agent_culture)
                            ORDER BY CONVERT(date, tourdate, 103) DESC
                        """)
                header = [desc[0] for desc in cursor.description]
                data = cursor.fetchall()
                df_booking = pd.DataFrame(data, columns=header)

            df_booking = df_booking.drop_duplicates(subset=['UniqueID'])
            df_booking['agentId'] = df_booking['agentId'].astype(str).str.strip().str.upper()

            df_booking = df_booking.drop(columns=['UniqueID'])

            booking_summary = (
                df_booking.groupby('agentId')
                .size()
                .reset_index(name='TotalBooking')
                .sort_values(by='TotalBooking', ascending=False)
            )

            print(booking_summary.head(5))

            booked_agents = set(booking_summary['agentId'])

            agents_with_queries_only = query_summary[~query_summary['AGENT_ID'].isin(booked_agents)].sort_values(by='TotalQuery', ascending=False)

            print(agents_with_queries_only)

            df_pivot = agents_with_queries_only.pivot_table(
                index=['AGENT_ID', 'EMAIL_ID', 'Name'],
                columns='QUERY_YEAR',
                values='TotalQuery',
                fill_value=0
            ).reset_index()

            df_pivot.columns.name = None

            # Optionally rename year columns to "Query_2024", etc.
            df_pivot = df_pivot.rename(columns=lambda x: f"Query_in_{x}" if isinstance(x, int) else x)

            # Now add TotalQuery column (sum across all year columns)
            year_columns = [col for col in df_pivot.columns if col.startswith("Query_in_")]
            df_pivot["TotalQuery"] = df_pivot[year_columns].sum(axis=1)

            print(df_pivot)
            
            if format == 'excel':
                response = HttpResponse(content_type='text/csv')
                response['Content-Disposition'] = 'attachment; filename="Agents_without_booking.csv"'
                
                agents_with_queries_only.to_csv(path_or_buf=response, index=False)
                return response

            return JsonResponse(df_pivot.to_dict(orient='records'), safe=False)

        except Exception as e:
            print("Error:", e)
            return JsonResponse({"error": str(e)}, status=500)


    def agent_login_trend(request):
        print("api/agent-login-trend")
        try:
            selected_year = ''
            agent_id = request.GET.get('agentId', '').strip().upper()

            year_param = request.GET.get('year', None)
            if year_param:
                selected_year = [y.strip() for y in year_param.split(',')]
            else:
                selected_year = []

            print(agent_id)
            print(selected_year)


            # Get pagination params from GET, default page=1, page_size=10
            page = int(request.GET.get('page', 1))
            page_size = int(request.GET.get('pageSize', 50))
            
            if page < 1:
                page = 1
            if page_size < 1:
                page_size = 10

            params = []
            if agent_id:
                query_part_1 = "AND b.AgentID = %s"
                params.append(agent_id)
            else:
                query_part_1 = ""

            # Step 1: Get all distinct AgentIDs for pagination (only those active in last 3 years)
            with connection.cursor() as cursor:
                cursor.execute(f"""
                    SELECT DISTINCT b.AgentID 
                    FROM tbl_agent b
                    JOIN tbl_login a 
                        ON (a.AGENTID = b.UserName OR a.AGENTID = b.Emailid)
                    WHERE b.AgentID IS NOT NULL AND b.AgentID != ''
                        AND len(b.AgentID) > 5	
                        {query_part_1}
                        AND YEAR(CONVERT(DATE,a.LOGINDATE,103)) > (YEAR(GETDATE()) - 3)
                        AND a.Loginby IS NULL
                        AND b.Status = 1
                        AND b.AgentID NOT IN (
                            select distinct (Agentid) as Agentid from tbl_agent_culture
                        )
                    ORDER BY b.AgentID
                """, params)

                executed_query = connection.queries[-1]['sql'] if connection.queries else "No query executed"
                print(executed_query)

                all_agents = [row[0].strip().upper() for row in cursor.fetchall()]

            total_agents = len(all_agents)
            total_pages = (total_agents + page_size - 1) // page_size

            # Pagination slice on agent list
            start_idx = (page - 1) * page_size
            end_idx = start_idx + page_size
            paged_agents = all_agents[start_idx:end_idx]    

            if not paged_agents:
                return JsonResponse({
                    'page': page,
                    'page_size': page_size,
                    'total_agents': total_agents,
                    'total_pages': total_pages,
                    'results': []
                })

            # Step 2: Fetch only data for these paged agents
            placeholders = ', '.join(['%s'] * len(paged_agents))
            
            if selected_year:
                query_part_2 = f"AND YEAR(CONVERT(DATE,a.LOGINDATE,103)) IN ({', '.join(selected_year)})"
            else:
                query_part_2 = ''
            
            print(query_part_2)

            query = f"""
                    SELECT a.LOGINDATE, b.AgentID, b.Name
                    FROM tbl_agent b
                    LEFT JOIN tbl_login a 
                        ON (a.AGENTID = b.UserName OR a.AGENTID = b.Emailid)
                        AND YEAR(CONVERT(DATE,a.LOGINDATE,103)) > (YEAR(GETDATE()) - 3)
                        {query_part_2}
                    WHERE b.AgentID IN ({placeholders})
                    ORDER BY a.LOGINDATE
                """
            print(query)

            #print(paged_agents)
            with connection.cursor() as cursor:
                cursor.execute(query, paged_agents)

                executed_query = connection.queries[-1]['sql'] if connection.queries else "No query executed"
                print(executed_query)

                header = [desc[0] for desc in cursor.description]
                data = cursor.fetchall()

            df_login = pd.DataFrame(data, columns=header)
            df_login['LOGINDATE'] = pd.to_datetime(df_login['LOGINDATE'], errors='coerce')
            df_login['AgentID'] = df_login['AgentID'].astype(str).str.strip().str.upper()
            df_login['Month_Year'] = df_login['LOGINDATE'].dt.to_period('M').dt.to_timestamp()
            df_login['Month_Year_Str'] = df_login['Month_Year'].dt.strftime('%b %Y')

            agent_names = df_login[['AgentID', 'Name']].drop_duplicates('AgentID').set_index('AgentID')['Name'].to_dict()

            grouped = df_login.groupby(['AgentID', 'Month_Year']).size().reset_index(name='Count')

            current_year = datetime.now().year
            current_month_start = pd.Timestamp(datetime.now().strftime('%Y-%m-01'))

            if selected_year:
                years = sorted(int(y) for y in selected_year)
                start = pd.Timestamp(f'{years[0]}-01-01')

                if current_year in years:
                    # Cap at current month if current year is included
                    end = current_month_start
                else:
                    end = pd.Timestamp(f'{years[-1]}-12-01')
            else:
                start_year = current_year - 2
                start = pd.Timestamp(f'{start_year}-01-01')
                end = current_month_start

            all_months = pd.date_range(start=start, end=end, freq='MS')

            # Step 3: Create full cartesian product (paged AgentIDs × all months)
            all_combinations = pd.MultiIndex.from_product([paged_agents, all_months], names=['AgentID', 'Month_Year']).to_frame(index=False)

            final = pd.merge(all_combinations, grouped, on=['AgentID', 'Month_Year'], how='left')
            final['Count'] = final['Count'].fillna(0).astype(int)
            final['Month_Year_Str'] = final['Month_Year'].dt.strftime('%b %Y')
            final['Name'] = final['AgentID'].map(agent_names)

            final.sort_values(by=['AgentID', 'Month_Year'], inplace=True)

            output = []
            for agent_id, group in final.groupby('AgentID'):
                output.append({
                    'AgentID': agent_id,
                    'Name': group['Name'].iloc[0] if not group['Name'].isnull().all() else None,
                    'data': group[['Month_Year_Str', 'Count']].to_dict(orient='records')
                })

            return JsonResponse({
                'page': page,
                'page_size': page_size,
                'total_agents': total_agents,
                'total_pages': total_pages,
                'results': output
            }, safe=False)

        except Exception as e:
            print("Error:", e)
            return JsonResponse({"error": str(e)}, status=500)
       

    def agent_query_with_booking(request):
        print("api/agent-query-with-booking")

        format = request.GET.get("format")

        agentid = request.GET.get('agentId', '').strip().upper()
        year = request.GET.get('year', '').strip()

        print(agentid)
        print(year)

        try:
            with connection.cursor() as cursor:
                cursor.execute(f""" 
                        select
                            (select top 1 agentid from tbl_agent where Emailid=tm.EMAIL_ID or username=tm.EMAIL_ID ) as AGENT_ID, 
                            (select top 1 Emailid from tbl_agent where Emailid=tm.EMAIL_ID or username=tm.EMAIL_ID ) as EMAIL_ID,
                            (select top 1 Name from tbl_agent where Emailid=tm.EMAIL_ID or username=tm.EMAIL_ID ) as Name,
                            CREATED_DATE
                        from TBL_MESSAGE TM
                        where Year(convert(date,TM.CREATED_DATE,103)) >= (year(getdate())-3)
                            and (tm.EMAIL_ID in(select emailid from tbl_agent where Emailid=tm.EMAIL_ID and AgentId not in (select distinct (Agentid) as Agentid from tbl_agent_culture)) or tm.EMAIL_ID in(select username from tbl_agent where username=tm.EMAIL_ID and AgentId not in (select distinct (Agentid) as Agentid from tbl_agent_culture)))
                        order by TM.MSG_ID desc
                """)

                executed_query = connection.queries[-1]['sql'] if connection.queries else "No query executed"
                print(executed_query)

                header = [desc[0] for desc in cursor.description]
                data = cursor.fetchall()
                df_query = pd.DataFrame(data, columns=header)

            df_query = df_query[df_query['AGENT_ID'].notna() & (df_query['AGENT_ID'].str.strip() != '')]
            df_query['AGENT_ID'] = df_query['AGENT_ID'].astype(str).str.strip().str.upper()
            df_query['EMAIL_ID'] = df_query['EMAIL_ID'].astype(str).str.strip().str.lower()
            df_query['Name'] = df_query['Name'].astype(str).str.strip().str.title()

            df_query['CREATED_DATE'] = pd.to_datetime(df_query['CREATED_DATE'], errors='coerce')
            df_query['QUERY_YEAR'] = df_query['CREATED_DATE'].dt.year


            query_summary = (
                df_query.groupby(['AGENT_ID', 'EMAIL_ID', 'Name', 'QUERY_YEAR'])
                .size()     
                .reset_index(name='TotalQuery')
                .sort_values(by='TotalQuery', ascending=False)
            )

            print(query_summary.head(5))


            with connection.cursor() as cursor:
                cursor.execute(f"""
                            SELECT 
                                CONCAT(TRIM(UPPER(agentId)), PackgID, convert(date,tourdate,103)) AS UniqueID,
                                TRIM(UPPER(agentId)) as agentId,
                                MIN(CreatedDate) AS FirstCreatedDate
                            FROM tbl_booking
                            WHERE 
                                txn_msg = 'success' 
                                AND Is_cancelled <> 1 
                                AND CreatedDate IS NOT NULL
                                AND tourdate IS NOT NULL 
                                AND agentId IS NOT NULL
                                AND PackgID IS NOT NULL
                                AND CreatedDate != ''
                                AND tourdate != ''
                                AND agentId != ''
                                AND PackgID != ''
                                AND agentId NOT In (select distinct (Agentid) as Agentid from tbl_agent_culture)
                            GROUP BY 
                                agentId, PackgID, tourdate
                            ORDER BY 
                                MIN(CONVERT(date, tourdate, 103)) DESC
                        """)
                executed_query = connection.queries[-1]['sql'] if connection.queries else "No query executed"
                print(executed_query)
                header = [desc[0] for desc in cursor.description]
                data = cursor.fetchall()
                df_booking = pd.DataFrame(data, columns=header)

           
            df_booking = df_booking.drop_duplicates(subset=['UniqueID'])
            df_booking['agentId'] = df_booking['agentId'].astype(str).str.strip().str.upper()

            df_booking = df_booking.drop(columns=['UniqueID'])

            booking_summary = (
                df_booking.groupby('agentId')
                .size()
                .reset_index(name='TotalBooking')
                .sort_values(by='TotalBooking', ascending=False)
            
            )

            print(booking_summary.head(5))


            booked_agents = set(booking_summary['agentId'])

           
            agents_with_queries_only = query_summary[query_summary['AGENT_ID'].isin(booked_agents)]

            agents_with_queries_only = agents_with_queries_only.merge(
                booking_summary[['agentId', 'TotalBooking']],
                left_on='AGENT_ID',
                right_on='agentId',
                how='left'
            ).drop(columns=['agentId'])

            print(agents_with_queries_only.head())

            
            df_pivot = agents_with_queries_only.pivot_table(
                index=['AGENT_ID', 'EMAIL_ID', 'Name', 'TotalBooking'],
                columns='QUERY_YEAR',
                values='TotalQuery',
                fill_value=0
            ).reset_index()


            df_pivot.columns.name = None

            # Optionally rename year columns to "Query_2024", etc.
            df_pivot = df_pivot.rename(columns=lambda x: f"Query_in_{x}" if isinstance(x, int) else x)

            # Now add TotalQuery column (sum across all year columns)
            year_columns = [col for col in df_pivot.columns if col.startswith("Query_in_")]
            df_pivot["TotalQuery"] = df_pivot[year_columns].sum(axis=1)

            print(df_pivot)

            if agentid and year:

                print(agentid)
                print(year)

                with connection.cursor() as cursor:
                    cursor.execute(f"""select CREATED_DATE, MSG_TYPE, Country from TBL_MESSAGE
                                        where AGENT_ID ='{agentid}'
                                        and year(convert(Date,CREATED_DATE,103)) = {year}
                                                                """)
                    
                    df_agent_message = pd.DataFrame(cursor.fetchall(), columns=[desc[0] for desc in cursor.description])

                    return JsonResponse(df_agent_message.to_dict(orient='records'), safe=False)
                

            if format == 'excel':
                response = HttpResponse(content_type='text/csv')
                response['Content-Disposition'] = 'attachment; filename="Agents_with_booking.csv"'
                
                agents_with_queries_only.to_csv(path_or_buf=response, index=False)
                return response

            return JsonResponse(df_pivot.to_dict(orient='records'), safe=False)


        except Exception as e:
            print("Error:", e)
            return JsonResponse({"error": str(e)}, status=500)
        


    def average_booking_report(request):
        print("api/average-booking")

        selected_pkgid = request.GET.get('pkg_id')
        start_date_str = request.GET.get('startDate')
        end_date_str = request.GET.get('endDate')

        try:
            if start_date_str:
                start_date = datetime.strptime(start_date_str, '%d-%m-%Y').date()
            else:
                start_date = None

            if end_date_str:
                end_date = datetime.strptime(end_date_str, '%d-%m-%Y').date()
            else:
                end_date = None

            with connection.cursor() as cursor:
                cursor.execute(f"""SELECT CONCAT(Upper(trim(agentId)), PackgID, convert(date,tourdate,103)) AS UniqueID, Upper(trim(agentId)) as agentId, convert(date,CreatedDate,103) as CreatedDate, convert(date,tourdate,103)) as tourdate, PackgID
                            FROM tbl_booking
                            WHERE txn_msg = 'success'
                                AND Is_cancelled <> 1       
                                AND CreatedDate IS NOT NULL
                                AND tourdate IS NOT NULL
                                AND agentId  IS NOT NULL
                                AND PackgID IS NOT NULL
                                AND CreatedDate != ''
                                AND tourdate != ''
                                AND agentId != ''
                                AND PackgID != ''
                                AND year(convert(date,CreatedDate,103)) > 2022
                                AND AgentId not In (select distinct (Agentid) as Agentid from tbl_agent_culture)
                            ORDER BY CONVERT(date, tourdate, 103) DESC
                               """)
                header = [desc[0] for desc in cursor.description]
                data = cursor.fetchall()
                df_booking = pd.DataFrame(data, columns=header)

            df_booking['tourdate'] = pd.to_datetime(df_booking['tourdate'], format='%d/%m/%Y', errors='coerce')
            df_booking['CreatedDate'] = pd.to_datetime(df_booking['CreatedDate'], format='%d/%m/%Y', errors='coerce')
            df_booking = df_booking.sort_values('CreatedDate').drop_duplicates(subset=['UniqueID'], keep='first')


            if selected_pkgid:
                try:
                    df_booking = df_booking[df_booking['PackgID'] == selected_pkgid]
                    print(df_booking)
                    
                except ValueError:  
                    return JsonResponse({'error': 'Invalid packg_ids value'}, safe=False), 400
                
            if start_date:
                df_booking = df_booking[df_booking['CreatedDate'].dt.date >= start_date]
            if end_date:
                df_booking = df_booking[df_booking['CreatedDate'].dt.date <= end_date]

            print(df_booking)
                
            bookings_per_day = df_booking.groupby(df_booking['CreatedDate'].dt.date).size()

            no_of_agent_with_booking = df_booking['agentId'].nunique()
            print(no_of_agent_with_booking)

            no_of_booking = df_booking['UniqueID'].nunique()
            print(no_of_booking)

          
            if bookings_per_day.empty:
                average_all_days = 0.0
            else:
                # Create full date range based on filtered data
                full_range_start = start_date
                
                full_range_end = end_date

                date_range = pd.date_range(start=full_range_start, end=full_range_end)

                # Fill missing days with 0
                bookings_full_range = bookings_per_day.reindex(date_range.date, fill_value=0)

                # Calculate mean
                average_all_days = bookings_full_range.mean()

                #average_all_days = round(average_all_days, 2)

            print("📊 Average bookings per day (filtered):", round(average_all_days, 2))

            response_data = {
                        'Average_bookings_per_day': round(average_all_days, 2),
                        'No_Of_Agents' : no_of_agent_with_booking,
                        'Total_Bookings' : no_of_booking,
                        }

            return JsonResponse(response_data, safe=False)

        except Exception as e:
            print("Error:", e)
            return JsonResponse({"error": str(e)}, status=500)


    def api_get_country_list(request):
        print('api/get-country-list')
        try:
            with connection.cursor() as cursor:
                cursor.execute(f"""exec GetCountryList""")

                header = [desc[0] for desc in cursor.description]
                data = cursor.fetchall()
            country_data = pd.DataFrame(data, columns=header)

            final_data = country_data[['CountryCode', 'CountryName']].sort_values('CountryName')

            return JsonResponse(final_data.to_dict(orient='records'), safe=False)

        except Exception as e:
            print("Error:", e)
            return JsonResponse({"error": str(e)}, status=500)



    def api_get_pkg_title(request):
        print('api/get-pkg-title')
        country_code = request.GET.get('countryCode')
        
        try:
            with connection.cursor() as cursor:
                cursor.execute("""
                                SELECT distinct trim(upper(a.PKG_TITLE)) as PKG_TITLE, a.PKG_ID
                                FROM TBL_PKG_DETAILS a 
                                JOIN TBL_PKG_ROOM_RATE b ON a.PKG_ID = b.PKG_ID
                                WHERE (a.AGENTID is NULL or a.AGENTID='')
                                and a.AGENTID NOT In (select distinct (Agentid) as Agentid from tbl_agent_culture)
                                AND a.Status=1
                                AND a.CountryCode LIKE %s
                                AND b.Status = 1
                                and b.RATE_AVIAL_DATE Is not Null 
                                AND b.RATE_AVIAL_DATE != ''
                                ORDER by PKG_TITLE
                               """, (f"%{country_code}%",))
                
                header = [desc[0] for desc in cursor.description]
                data = cursor.fetchall()

            pkg_title = pd.DataFrame(data, columns=header)

            return JsonResponse(pkg_title.to_dict(orient='records'), safe=False)


        except Exception as e:
            print("Error:", e)
            return JsonResponse({"error": str(e)}, status=500)
        

    def guest_details_by_package(request):
        print('api/active-cancel-guest')

        pkg_id = request.GET.get('pkgid')
        active_guest_counts = []
        remove_guest_counts = []

        try:
            with connection.cursor() as cursor:
                cursor.execute("""
                                ;WITH ValidBookings AS (
                                    SELECT DISTINCT QueryID
                                    FROM TBL_BOOKING
                                    WHERE txn_msg = 'success' and is_cancelled <> 1 
                                        AND CreatedDate is NOT NULL
                                        AND tourdate is NOT NULL
                                        AND agentId is NOT NULL
                                        AND PackgID Is NOT NULL
                                        AND CreatedDate != ''
                                        AND tourdate != ''
                                        AND agentId != ''
                                        AND PackgID != ''
                                        AND agentid NOT In (select distinct (Agentid) as Agentid from tbl_agent_culture)
                                )
                                SELECT
                                    TN.TourDate,
                                    year(convert(date,TN.tourdate, 103)) as Year,
                                    
                                    -- Total Active Bookings
                                    SUM(CASE WHEN TN.Status1 = 'Active' THEN 1 ELSE 0 END) AS Active_Guest,
                                    
                                    -- Total Removed Bookings
                                    SUM(CASE WHEN TN.Status1 = 'Remove' THEN 1 ELSE 0 END) AS Remove_Guest

                                FROM TBL_TRAVELLER_NAME TN
                                WHERE TN.Pkgid = %s
                                AND TN.PaxDepositAmount > 0
                                AND CONVERT(DATE, TN.TourDate, 103) >= CONVERT(DATE, GETDATE())
                                AND CONVERT(DATE, TN.TourDate, 103) < DATEADD(YEAR, 3, GETDATE())
                                AND TN.PKG_QUERY_ID IN (SELECT QueryID FROM ValidBookings)

                                GROUP BY TN.TourDate 
                                ORDER BY CONVERT(DATE, TN.TourDate, 103);
                            """, [pkg_id])
       
                header = [row[0] for row in cursor.description]
                data = cursor.fetchall()
                 
            pkg_date_details = pd.DataFrame(data, columns=header)

            return JsonResponse (pkg_date_details.to_dict(orient='records'), safe=False)

        except Exception as e:
            print("Error:", e)
            return JsonResponse({"error": str(e)}, status=500)
        

    def customize_report(request):
        print('api/customize-report')
        pkg_id = request.GET.get('pkgId')

        if pkg_id:
            query = f'AND Packgid = {pkg_id}'
        else:
            query = ''
        
        current_year = datetime.now().year

        try:
            with connection.cursor() as cursor:
                cursor.execute(f"""
                        SELECT
                            DATENAME(MONTH, CreatedDate) AS MonthName,
                            SUM(CASE WHEN YEAR(CreatedDate) = YEAR(GETDATE()) - 3 THEN cast(USDamt as money) ELSE 0 END) AS Previous_3_Year,
                            SUM(CASE WHEN YEAR(CreatedDate) = YEAR(GETDATE()) - 2 THEN cast(USDamt as money) ELSE 0 END) AS Previous_2_Year,
                            SUM(CASE WHEN YEAR(CreatedDate) = YEAR(GETDATE()) - 1 THEN cast(USDamt as money) ELSE 0 END) AS Previous_Year,
                            SUM(CASE WHEN YEAR(CreatedDate) = YEAR(GETDATE()) THEN cast(USDamt as money) ELSE 0 END) AS Current_Year
                        FROM TBL_BOOKING
                        WHERE txn_msg = 'success' 
                            AND Is_cancelled <> 1       
                            AND CreatedDate is NOT NULL
                            AND tourdate is NOT NULL
                            AND agentId is NOT NULL
                            AND PackgID Is NOT NULL
                            AND CreatedDate != ''
                            AND tourdate != ''
                            AND agentId != ''
                            AND PackgID != ''
                            AND agentId NOT In (select distinct (Agentid) as Agentid from tbl_agent_culture)
                            {query}
                        GROUP BY DATENAME(MONTH, CreatedDate), MONTH(CreatedDate)
                        ORDER BY MONTH(CreatedDate)
                        """)
                data = cursor.fetchall()
                header = [col[0] for col in cursor.description]

            createddate_report =  pd.DataFrame(data, columns=header)

            createddate_report = createddate_report.rename(
                columns={
                    "Previous_3_Year": current_year - 3,
                    "Previous_2_Year": current_year - 2,
                    "Previous_Year" : current_year - 1,
                    "Current_Year" : current_year
                }
            )
            createddate_report_json = createddate_report.set_index('MonthName').to_dict(orient='index')



            with connection.cursor() as cursor:
                cursor.execute(f"""
                        SELECT
                            DATENAME(MONTH, convert(date,tourdate,103)) AS MonthName,
                            SUM(CASE WHEN YEAR(convert(date,tourdate,103)) = YEAR(GETDATE()) - 3 THEN cast(USDamt as money) ELSE 0 END) AS Previous_3_Year,
                            SUM(CASE WHEN YEAR(convert(date,tourdate,103)) = YEAR(GETDATE()) - 2 THEN cast(USDamt as money) ELSE 0 END) AS Previous_2_Year,
                            SUM(CASE WHEN YEAR(convert(date,tourdate,103)) = YEAR(GETDATE()) - 1 THEN cast(USDamt as money) ELSE 0 END) AS Previous_Year,
                            SUM(CASE WHEN YEAR(convert(date,tourdate,103)) = YEAR(GETDATE()) THEN cast(USDamt as money) ELSE 0 END) AS Current_Year,
                            SUM(CASE WHEN YEAR(convert(date,tourdate,103)) = YEAR(GETDATE()) + 1 THEN cast(USDamt as money) ELSE 0 END) AS Next_Year
                        FROM TBL_BOOKING
                        WHERE txn_msg = 'success' 
                            AND Is_cancelled <> 1       
                            AND CreatedDate is NOT NULL
                            AND tourdate is NOT NULL
                            AND agentId is NOT NULL
                            AND PackgID Is NOT NULL
                            AND CreatedDate != ''
                            AND tourdate != ''
                            AND agentId != ''
                            AND PackgID != ''
                            AND agentId NOT In (select distinct (Agentid) as Agentid from tbl_agent_culture)
                               {query}
                        GROUP BY DATENAME(MONTH, convert(date,tourdate,103)), MONTH(convert(date,tourdate,103))
                        ORDER BY MONTH(convert(date,tourdate,103))
                        """)
                data = cursor.fetchall()
                header = [col[0] for col in cursor.description]

            tourdate_report =  pd.DataFrame(data, columns=header)

            tourdate_report = tourdate_report.rename(
                columns={
                    "Previous_3_Year": current_year - 3,
                    "Previous_2_Year": current_year - 2,
                    "Previous_Year" : current_year - 1,
                    "Current_Year" : current_year,
                    "Next_Year" : current_year + 1
                }
            )
            tourdate_report_json = tourdate_report.set_index('MonthName').to_dict(orient='index')




            with connection.cursor() as cursor:
                cursor.execute(f"""
                            SELECT CONCAT(UPPER(LTRIM(RTRIM(agentId))), PackgID, convert(date,tourdate,103)) AS UniqueID, trim(upper(agentId)) as agentId, convert(date,CreatedDate,103) as CreatedDate, convert(date,tourdate,103) as tourdate, PackgID, USDamt
                            from TBL_BOOKING
                            where txn_msg = 'success'
                            AND Is_cancelled <> 1       
                            AND CreatedDate is NOT NULL
                                AND tourdate is NOT NULL
                                AND agentId is NOT NULL
                                AND PackgID Is NOT NULL
                                AND CreatedDate != ''
                                AND tourdate != ''
                                AND agentId != ''
                                AND PackgID != ''
                                AND agentId NOT In (select distinct (Agentid) as Agentid from tbl_agent_culture)
                                {query}
                                
                               """)
                data = cursor.fetchall()
                header = [col[0] for col in cursor.description]
            df_booking = pd.DataFrame(data, columns=header)

            
            with connection.cursor() as cursor:
                cursor.execute(f"""
                                select trav_id, trim(upper(created_by)) as created_by, convert(date,tourdate,103) as tourdate, Pkgid from TBL_TRAVELLER_NAME
                                where PKG_QUERY_ID IN (
                                    SELECT trim(upper(QueryID)) as QueryID
                                    from TBL_BOOKING
                                    where txn_msg = 'success'
                                    AND Is_cancelled <> 1       
                                    AND CreatedDate is NOT NULL
                                        AND tourdate is NOT NULL
                                        AND agentId is NOT NULL
                                        AND PackgID Is NOT NULL
                                        AND CreatedDate != ''
                                        AND tourdate != ''
                                        AND agentId != ''
                                        AND PackgID != ''
                                        AND QueryID != '' AND QueryID != '-'
                                        AND agentId NOT In (select distinct (Agentid) as Agentid from tbl_agent_culture)
                                        {query}
                                        AND year(convert(date, CreatedDate, 103)) >= 2022
                                    )
                                AND status1 = 'Active'
                                AND PaxDepositAmount > 99
                               """)
                data2 = cursor.fetchall()
                header2 = [col[0] for col in cursor.description]
            df_traveller = pd.DataFrame(data2, columns=header2)

            df_traveller['tourdate'] = pd.to_datetime(df_traveller['tourdate'], format='%d/%m/%Y', errors='coerce')
            df_traveller['TourYear'] = df_traveller['tourdate'].dt.year

            traveller_count = df_traveller.groupby('TourYear').size().reset_index(name='Total_Travellers')
            traveller_count.rename(columns={'TourYear': 'Year'}, inplace=True)


            # Convert date columns
            df_booking['tourdate'] = pd.to_datetime(df_booking['tourdate'], format='%d/%m/%Y', errors='coerce')
            df_booking['CreatedDate'] = pd.to_datetime(df_booking['CreatedDate'], format='%d/%m/%Y', errors='coerce')
            df_booking['USDamt'] = pd.to_numeric(df_booking['USDamt'], errors='coerce')


            # 1. USD Total (no duplicate removal)
            df_booking['CreatedYear'] = df_booking['CreatedDate'].dt.year
            usd_total = df_booking.groupby('CreatedYear')['USDamt'].sum().reset_index()
            usd_total.columns = ['Year', 'Total_Revenue']

            # 2. Unique Bookings and Agents (after dropping duplicates)
            df_unique = df_booking.sort_values('CreatedDate').drop_duplicates(subset=['UniqueID'], keep='first')
            df_unique['CreatedYear'] = df_unique['CreatedDate'].dt.year
            booking_agent_stats = df_unique.groupby('CreatedYear').agg({
                'UniqueID': 'nunique',
                'agentId': 'nunique'
            }).reset_index()
            booking_agent_stats.columns = ['Year', 'Total_Booking', 'Total_Agents']

            # Merge both 
            yearly_stats = pd.merge(booking_agent_stats, usd_total, on='Year')

            

            yearly_stats = pd.merge(yearly_stats, traveller_count, on='Year', how='left')

            yearly_stats['Total_Travellers'] = yearly_stats['Total_Travellers'].fillna(0).astype(int)

            yearly_stats = yearly_stats[yearly_stats['Year'] > 2021]

            print(yearly_stats)

            yearly_stats_json = yearly_stats.set_index('Year').to_dict(orient='index')

            response_data = {
                        'Booking_report': createddate_report_json,
                        'Tourdate_report': tourdate_report_json,
                        'Data_for_table' : yearly_stats_json,
                        }


            return JsonResponse(response_data, safe=False)

        except Exception as e:
            print("Error:", e)
            return JsonResponse({"error": str(e)}, status=500)



    def series_booking_overview(request):
        print('api/series-booking')

        try:
            pkg_id = request.GET.get('pkgId')
            print(pkg_id)

            current_year = datetime.now().year
            
            query = ""
            params = []

            if pkg_id:
                try:
                    int(pkg_id)  # Ensures pkg_id is numeric
                    query = "AND PackgID = %s"
                    params.append(pkg_id)
                except ValueError:
                    raise ValueError("Invalid Package ID")
            
            
            with connection.cursor() as cursor:
                cursor.execute(f"""
                    SELECT
                        DATENAME(MONTH, CreatedDate) AS MonthName,
                        SUM(CASE WHEN YEAR(CreatedDate) = YEAR(GETDATE()) - 3 THEN cast(USDamt as money) ELSE 0 END) AS Previous_3_Year,
                        SUM(CASE WHEN YEAR(CreatedDate) = YEAR(GETDATE()) - 2 THEN cast(USDamt as money) ELSE 0 END) AS Previous_2_Year,
                        SUM(CASE WHEN YEAR(CreatedDate) = YEAR(GETDATE()) - 1 THEN cast(USDamt as money) ELSE 0 END) AS Previous_Year,
                        SUM(CASE WHEN YEAR(CreatedDate) = YEAR(GETDATE()) THEN cast(USDamt as money) ELSE 0 END) AS Current_Year
                    FROM TBL_BOOKING
                    WHERE txn_msg = 'success' 
                        AND Is_cancelled <> 1       
                        AND CreatedDate IS NOT NULL
                        AND tourdate IS NOT NULL
                        AND agentId IS NOT NULL
                        AND PackgID IS NOT NULL
                        AND CreatedDate != ''
                        AND tourdate != ''
                        AND agentId != ''
                        AND PackgID != ''
                        AND agentId NOT In (select distinct (Agentid) as Agentid from tbl_agent_culture)
                        {query}
                    GROUP BY DATENAME(MONTH, CreatedDate), MONTH(CreatedDate)
                    ORDER BY MONTH(CreatedDate)
                """, params)

                data = cursor.fetchall()

                if not data:
                    raise ValueError("No data found for the given package ID")

                headers = [col[0] for col in cursor.description]

            df = pd.DataFrame(data, columns=headers)

            df = df.rename(columns={
                "Previous_3_Year": str(current_year - 3),
                "Previous_2_Year": str(current_year - 2),
                "Previous_Year": str(current_year - 1),
                "Current_Year": str(current_year)
            })

            print(df)

            df_json = df.set_index('MonthName').to_dict(orient='index')

            return JsonResponse(df_json, safe=False)

        
        except Exception as e:
            print("Error:", e)
            return JsonResponse({"error": str(e)}, status=500)


    def customize_booking_overview(request):
        print('api/customize-booking')
        country_code = request.GET.get('countryCode')

        current_year = datetime.now().year

        query = ''
        query1 = ''
        params = []
        params1 = []
        if country_code:
            query = "AND CountryCode LIKE %s"
            params.append(f'%{country_code}%')

            query1 = "AND p.CountryCode LIKE %s"
            params1.append(f'%{country_code}%')
        
        try:

            with connection.cursor() as cursor:
                cursor.execute(f"""
                                SELECT
                                    DATENAME(MONTH, CreatedDate) AS MonthName,
                                    SUM(CASE WHEN YEAR(CreatedDate) = YEAR(GETDATE()) - 3 THEN cast(USDamt as money) ELSE 0 END) AS Previous_3_Year,
                                    SUM(CASE WHEN YEAR(CreatedDate) = YEAR(GETDATE()) - 2 THEN cast(USDamt as money) ELSE 0 END) AS Previous_2_Year,
                                    SUM(CASE WHEN YEAR(CreatedDate) = YEAR(GETDATE()) - 1 THEN cast(USDamt as money) ELSE 0 END) AS Previous_Year,
                                    SUM(CASE WHEN YEAR(CreatedDate) = YEAR(GETDATE()) THEN cast(USDamt as money) ELSE 0 END) AS Current_Year
                                FROM TBL_BOOKING
                                WHERE txn_msg = 'success' 
                                    AND Is_cancelled <> 1       
                                    AND CreatedDate is NOT NULL
                                    AND tourdate is NOT NULL
                                    AND agentId is NOT NULL
                                    AND PackgID Is NOT NULL
                                    AND CreatedDate != ''
                                    AND tourdate != ''
                                    AND agentId != ''
                                    AND PackgID != ''
                                    AND agentId NOT In (select distinct (Agentid) as Agentid from tbl_agent_culture)
                                    and PackgID In (
                                        select PKG_ID
                                        from TBL_PKG_DETAILS
                                        WHERE AGENTID is not NULL 
                                        and  AGENTID != ''
                                        AND Status=1
                                        {query}
                                    )
                                GROUP BY DATENAME(MONTH, CreatedDate), MONTH(CreatedDate)
                                ORDER BY MONTH(CreatedDate)
                            """,params) 
                
                data = cursor.fetchall()
                headers = [col[0] for col in cursor.description]
  
            customize_booking_overview = pd.DataFrame(data, columns=headers)


            customize_booking_overview = customize_booking_overview.rename(
                    columns={
                        "Previous_3_Year": current_year - 3,
                        "Previous_2_Year": current_year - 2,
                        "Previous_Year" : current_year - 1,
                        "Current_Year" : current_year
                    }
                )
            
            print(customize_booking_overview)
            
            customize_booking_overview_json = customize_booking_overview.set_index('MonthName').to_dict(orient='index')


            with connection.cursor() as cursor:
                cursor.execute(f"""
                            SELECT 
                                b.PackgID, 
                                CONVERT(date, b.tourdate, 103) AS tourdate,
                                MIN(TRIM(p.PKG_TITLE)) AS PKG_TITLE,
                                MIN(p.AGENTID) AS AGENTID
                            FROM tbl_booking b
                            INNER JOIN TBL_PKG_DETAILS p ON TRY_CAST(b.PackgID AS INT) = p.PKG_ID
                            WHERE 
                                b.txn_msg = 'success'
                                AND b.agentid not In (select distinct (Agentid) as Agentid from tbl_agent_culture)
                                AND b.Is_cancelled <> 1       
                                AND YEAR(CONVERT(date, b.tourdate, 103)) >= YEAR(GETDATE()) - 3
                                AND p.AGENTID IS NOT NULL
                                AND p.AGENTID != ''
                                AND p.Status = 1
                                {query1}
                            GROUP BY 
                                b.PackgID, 
                                CONVERT(date, b.tourdate, 103)
                            ORDER BY 
                                b.PackgID
                        """, params1)
                
                


                package_data = cursor.fetchall()
                package_headers = (row[0] for row in cursor.description)

            df_package = pd.DataFrame(package_data, columns=package_headers)


    

            all_agent_ids = set()
            df_package['AGENTID'].dropna().apply(lambda x: all_agent_ids.update([aid.strip() for aid in x.split(',') if aid.strip()]))

            agent_lookup = {}
            if all_agent_ids:
                placeholders = ','.join(['%s'] * len(all_agent_ids))
                with connection.cursor() as cursor:
                    cursor.execute(f"""
                        SELECT Name, AgentID 
                        FROM TBL_Agent
                        WHERE LEN(AgentID) > 5 AND AgentID IN ({placeholders}) AND STATUS = 1
                    """, list(all_agent_ids))
                    agent_info = cursor.fetchall()
                    agent_lookup = {aid: name for name, aid in agent_info}

            def get_agent_names(agentid_raw):
                if not agentid_raw:
                    return ''
                agent_ids = [aid.strip() for aid in agentid_raw.split(',') if aid.strip()]
                return ', '.join([agent_lookup.get(aid, '') for aid in agent_ids])

            df_package['AgentName'] = df_package['AGENTID'].apply(get_agent_names)



            # 2. Get Total Amounts in Bulk
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT 
                        PackgID, 
                        CONVERT(date, tourdate, 103) AS tourdate, 
                        SUM(CAST(USDamt AS MONEY)) AS Total_Amount
                    FROM tbl_booking
                    WHERE txn_msg = 'success' 
                        AND Is_cancelled <> 1      
                        AND CreatedDate is NOT NULL
                        AND tourdate is NOT NULL
                        AND agentId is NOT NULL
                        AND PackgID Is NOT NULL
                        AND CreatedDate != ''
                        AND tourdate != ''
                        AND agentId != ''
                        AND PackgID != '' 
                        AND agentId NOT In (select distinct (Agentid) as Agentid from tbl_agent_culture)      
                    GROUP BY PackgID, CONVERT(date, tourdate, 103)
                """)
                amount_data = cursor.fetchall()

            df_amount = pd.DataFrame(amount_data, columns=['PackgID', 'tourdate', 'Total_Amount'])

            # 3. Get Total Guest Counts in Bulk
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT 
                        Pkgid, 
                        CONVERT(date, TourDate, 103) AS tourdate, 
                        COUNT(*) AS Total_Guest
                    FROM tbl_traveller_name
                    WHERE Status1 = 'Active'      
                        AND CREATED_BY NOT In (select distinct (Agentid) as Agentid from tbl_agent_culture)
                    GROUP BY Pkgid, CONVERT(date, TourDate, 103)
                """)
                guest_data = cursor.fetchall()

            df_guest = pd.DataFrame(guest_data, columns=['PackgID', 'tourdate', 'Total_Guest'])

            # 4. Merge all
            df_package = df_package.merge(df_amount, on=['PackgID', 'tourdate'], how='left')
            df_package = df_package.merge(df_guest, on=['PackgID', 'tourdate'], how='left')

            df_package['Total_Amount'] = pd.to_numeric(df_package['Total_Amount'], errors='coerce').fillna(0).round(2)
            df_package['Total_Guest'] = df_package['Total_Guest'].fillna(0).astype(int)

            df_package['TourYear'] = pd.to_datetime(df_package['tourdate'], errors='coerce').dt.year


            df_package['tourdate'] = pd.to_datetime(df_package['tourdate'], errors='coerce')
            df_package['tourdate'] = df_package['tourdate'].dt.strftime('%Y-%m-%d')  # Ensure date is JSON serializable 
            df_package = df_package.sort_values('Total_Amount', ascending=False)
            df_package['Total_Amount'] = df_package['Total_Amount'].map(lambda x: int(round(float(x), 2)))

            df_package = df_package[['AgentName', 'AGENTID', 'PackgID', 'tourdate', 'PKG_TITLE', 'Total_Amount', 'Total_Guest', 'TourYear']]
        
            # print(df_package)

            df_package = df_package.to_dict(orient='records') 

            
        

            #df_package = df_package.to_dict()


            final_result = {
                'customize_package_booking' : customize_booking_overview_json,
                'customize_package_details' : df_package,
            }



            return JsonResponse (final_result, safe=False)

        except Exception as e:
            print("Error:", e)
            return JsonResponse({"error": str(e)}, status=500)


    def api_get_emp_id(request):
        try:
            with connection.cursor() as cursor:
                cursor.execute(f"""
                                SELECT upper(trim(StaffName)) as StaffName, EmpId from tblStaff
                                where Status = 1
                                and EmpId LIKE 'CH0%'
                                order by StaffName
                               """)
                data = cursor.fetchall()
                headers = [row[0] for row in cursor.description]

            df = pd.DataFrame(data, columns=headers)

            return JsonResponse(df.to_dict(orient='records'), safe=False)

        except Exception as e:
            print("Error:", e)
            return JsonResponse({"error": str(e)}, status=500)


    def tour_sale_by_staff(request):
        # print("api/emp-sales")
        try:
            current_year = datetime.now().year
            empid = request.GET.get('empId')
            
            query = ''
           
            params = []
            if empid:
                query = "AND SalesID = %s"
                params.append(f'{empid}')

            
            with connection.cursor() as cursor:
                cursor.execute(f"""
                                SELECT
                                    DATENAME(MONTH, CreatedDate) AS MonthName,
                                    SUM(CASE WHEN YEAR(CreatedDate) = YEAR(GETDATE()) - 3 THEN cast(USDamt as money) ELSE 0 END) AS Previous_3_Year,
                                    SUM(CASE WHEN YEAR(CreatedDate) = YEAR(GETDATE()) - 2 THEN cast(USDamt as money) ELSE 0 END) AS Previous_2_Year,
                                    SUM(CASE WHEN YEAR(CreatedDate) = YEAR(GETDATE()) - 1 THEN cast(USDamt as money) ELSE 0 END) AS Previous_Year,
                                    SUM(CASE WHEN YEAR(CreatedDate) = YEAR(GETDATE()) THEN cast(USDamt as money) ELSE 0 END) AS Current_Year
                                FROM TBL_BOOKING
                                WHERE txn_msg = 'success' 
                               AND Is_cancelled <> 1       
                                    AND CreatedDate is NOT NULL
                                    AND tourdate is NOT NULL
                                    AND agentId is NOT NULL
                                    AND PackgID Is NOT NULL
                                    AND CreatedDate != ''
                                    AND tourdate != ''
                                    AND agentId != ''
                                    AND PackgID != ''
                                    AND agentId NOT In (select distinct (Agentid) as Agentid from tbl_agent_culture)
                                    and PackgID In (
                                        select PKG_ID
                                        from TBL_PKG_DETAILS
                                        WHERE AGENTID is not NULL 
                                        and  AGENTID != ''
                                        AND Status=1
                                        {query}
                                    )
                                GROUP BY DATENAME(MONTH, CreatedDate), MONTH(CreatedDate)
                                ORDER BY MONTH(CreatedDate)
                               """, params)

                data = cursor.fetchall()
                headers = [row[0] for row in cursor.description]

            df_emp_sale = pd.DataFrame(data, columns=headers)

            df_emp_sale = df_emp_sale.rename(
                    columns={
                        "Previous_3_Year": current_year - 3,
                        "Previous_2_Year": current_year - 2,
                        "Previous_Year" : current_year - 1,
                        "Current_Year" : current_year
                    }
                )
            
            # print(df_emp_sale)

            df_emp_sale_json = df_emp_sale.set_index('MonthName').to_dict(orient='index') # For Bar Chart

            # result = {
            #     'emp-sales-report' : df_emp_sale_json
            #     }

            return JsonResponse(df_emp_sale_json, safe=False)
        
        except Exception as e:
            print("Error:", e)
            return JsonResponse({"error": str(e)}, status=500)
        

    def tour_sale_by_staff_by_year(request):
        print("api/emp-sales-year")
        try:
            empid = request.GET.get('empId')
            print(empid)

            selected_year = request.GET.get('year')

            print(selected_year)

            query = ''
            query2 = ''
            params = []
            if empid:
                query = "AND SalesID = %s"
                params.append(f'{empid}')

            if selected_year:
                query2 = f'HAVING YEAR(MIN(CreatedDate)) = {selected_year}'

    
            with connection.cursor() as cursor:
                cursor.execute(f"""
                                select trim(upper(CountryCode)) as CountryCode
                                    from TBL_PKG_DETAILS
                                    WHERE AGENTID is not NULL
                                    and  AGENTID != ''
                                    AND Status=1
                                    AND PKG_ID IN (Select packgid 
                                                    from TBL_BOOKING
                                                    where txn_msg = 'success'
                                                    and Is_cancelled <> 1
                                                    AND CreatedDate is NOT NULL
                                                    AND tourdate is NOT NULL
                                                    AND agentId is NOT NULL
                                                    AND PackgID Is NOT NULL
                                                    AND CreatedDate != ''
                                                    AND tourdate != ''
                                                    AND agentId != ''
                                                    AND PackgID != ''
                                                    AND agentId NOT In (select distinct (Agentid) as Agentid from tbl_agent_culture)
                                                    group by agentId, PackgID, convert(date,tourdate,103)
                                                    {query2}
                                                    )
                                    {query}
                               """, params)
                executed_query = connection.queries[-1]['sql'] if connection.queries else "No query executed"
                print(executed_query)

                data1 = cursor.fetchall()

                if not data1:
                    raise ValueError("No data found for the Given Year")
                
                headers1 = [row[0] for row in cursor.description]

            df_emp_country_sale = pd.DataFrame(data1, columns=headers1)

       

            
            df_emp_country_sale['CountryCode'] = df_emp_country_sale['CountryCode'].apply(
                lambda x: [c.strip() for c in x.split(',')] if isinstance(x, str) else [x]
            )

            # Explode the list into multiple rows
            df_emp_country_sale = df_emp_country_sale.explode('CountryCode').reset_index(drop=True)

            country_counts = df_emp_country_sale['CountryCode'].value_counts().reset_index()
            country_counts.columns = ['CountryCode', 'Count']


            
            unique_country_codes = country_counts['CountryCode'].dropna().unique().tolist()

            print(unique_country_codes)

           
            placeholders = ','.join(['%s'] * len(unique_country_codes))

            # Query all CountryNames in one go
            with connection.cursor() as cursor:
                cursor.execute(f"""
                    SELECT CountryCode, trim(upper(CountryName)) as CountryName
                    FROM tbl_country
                    WHERE CountryCode IS NOT NULL
					AND CountryName IS NOT NULL
					AND CountryCode != ''
					AND CountryName != ''
                    AND Status = 1 AND CountryCode IN ({placeholders})
                """, unique_country_codes)

                results = cursor.fetchall()

            # Create map and assign country names
            country_map = {code: name for code, name in results}

            print(country_map)
        
            country_counts['CountryName'] = country_counts['CountryCode'].map(country_map)

            country_counts = country_counts.drop(columns='CountryCode')

            country_counts['CountryName'] = country_counts['CountryName'].replace(
                    {
                        'BALI': 'BALI/INDONESIA', 
                        'INDONESIA': 'BALI/INDONESIA'
                    })

            country_counts = country_counts.sort_values('Count', ascending=False).reset_index(drop=True)

            # Keep top 4 rows
            top_countries = country_counts.iloc[:4]

            # Sum the rest
            other_total = country_counts.iloc[4:]['Count'].sum()

            

            # Create "Other" row only if there are extra countries
            if other_total > 0:
                other_row = pd.DataFrame([{'CountryName': 'Other', 'Count': other_total}])
                final_country_counts = pd.concat([top_countries, other_row], ignore_index=True)
            else:
                final_country_counts = top_countries

            print(final_country_counts)

            country_counts_json = final_country_counts.to_dict(orient='records')
            
            

            return JsonResponse (country_counts_json, safe=False)
        
        except Exception as e:
            print("Error:", e)
            return JsonResponse({"error": str(e)}, status=500)


    def country_of_customize_booking_month(request):
        print('api/country-customize-booking')

        try:
            with connection.cursor() as cursor:
                cursor.execute(f"""
                                select PackgID, convert(date,tourdate,103) as tourdate,
                                MIN(CreatedDate) AS CreatedDate
                                from tbl_booking
                                where txn_msg='success'
                                    AND Is_cancelled <> 1 
                                    AND CreatedDate is NOT NULL
                                    AND agentId is NOT NULL
                                    AND CreatedDate != ''
                                    AND agentId != ''
                                    AND agentId NOT In (select distinct (Agentid) as Agentid from tbl_agent_culture)
                                    and year(convert(date, tourdate,103)) >= 2022
                                    and PackgID in (select PKG_ID from TBL_PKG_DETAILS
                                                    where Status = 1
                                                    and AGENTID LIKE '%CHAGT%'
                                                    )
                                Group by agentId, PackgID, convert(date,tourdate,103)
                               """)
                
                booking_data = cursor.fetchall()
                booking_header = [row[0] for row in cursor.description]
                df_booking = pd.DataFrame(booking_data, columns=booking_header)

                print(df_booking)

                cursor.execute(f"""select PKG_ID, CountryCode, trim(upper(Country)) as Country from TBL_PKG_DETAILS
                                    where Status = 1
                                    AND AGENTID NOT IN (select distinct (Agentid) as Agentid from tbl_agent_culture)
                                    and AGENTID LIKE '%CHAGT%'
                               """)
                pkg_data = cursor.fetchall()
                pkg_header = [row[0] for row in cursor.description]
                df_pkg = pd.DataFrame(pkg_data, columns=pkg_header)

                print(df_pkg)

            df_pkg['CountryCode'] = df_pkg['CountryCode'].astype(str).str.split(',')
            df_pkg['Country'] = df_pkg['Country'].astype(str).str.split(',')

            df_pkg['country_pairs'] = df_pkg.apply(lambda row: list(zip(row['CountryCode'], row['Country'])), axis=1)

            # Step 3: Explode the zipped list
            df_pkg = df_pkg.explode('country_pairs').reset_index(drop=True)

            # Step 4: Split back into separate columns
            df_pkg[['CountryCode', 'Country']] = pd.DataFrame(df_pkg['country_pairs'].tolist(), index=df_pkg.index)

            # Step 5: Drop the temporary column
            df_pkg.drop(columns='country_pairs', inplace=True)


            print(df_pkg)

            #df_package = df_package.merge(df_amount, on=['PackgID', 'tourdate'], how='left')
            #df_result = df_result.merge(df_agent_info, how='left', left_on='agentId', right_on='AgentID')

            df_booking['PackgID'] = df_booking['PackgID'].astype(str)
            df_pkg['PKG_ID'] = df_pkg['PKG_ID'].astype(str)


            df_booking['CreatedDate'] = pd.to_datetime(df_booking['CreatedDate'], format='%d/%m/%Y', errors='coerce')
            df_booking['CreatedYear'] = df_booking['CreatedDate'].dt.year
            df_booking['CreatedMonth'] = df_booking['CreatedDate'].dt.strftime('%B')


            df_final = df_booking.merge(df_pkg, how='left', left_on='PackgID', right_on='PKG_ID')

         

            df_final['Country'] = df_final['Country'].replace(
                    {
                        'BALI': 'INDONESIA', 
                    })

       

            df_final = df_final.drop_duplicates(subset=[ 'PackgID', 'CreatedDate', 'CreatedYear', 'PKG_ID', 'CountryCode','Country'])


            print(df_final)

            df_final['CreatedMonth'] = df_final['CreatedMonth'].str.capitalize()

            # Step 1: Get the most frequent country per (year, month)
            mode_country = (
                df_final.groupby(['CreatedYear', 'CreatedMonth'])['Country']
                .agg(lambda x: x.mode().iloc[0] if not x.mode().empty else None)
                .reset_index()
            )

            # mode_country = (
            #     df_final.groupby(['CreatedYear', 'CreatedMonth'])['Country']
            #     .agg(lambda x: '/'.join(sorted(x.mode())) if not x.mode().empty else '')
            #     .reset_index()
            # )



            print(mode_country)

            # Step 2: Pivot to get months as columns, years as rows
            pivot_table = mode_country.pivot(index='CreatedYear', columns='CreatedMonth', values='Country')

            # Step 3: Ensure all 12 months are present
            month_order = list(month_name)[1:]  # ['January', 'February', ..., 'December']
            pivot_table = pivot_table.reindex(columns=month_order)

            # Optional: Fill missing cells with blank (default is NaN)
            pivot_table = pivot_table.fillna('')

            # Reset index to convert back to a DataFrame
            pivot_table = pivot_table.reset_index()

            print(pivot_table)

            result_json = pivot_table.to_dict(orient='records')

            return JsonResponse(result_json, safe=False)



        except Exception as e:
            print("Error:", e)
            return JsonResponse({"error": str(e)}, status=500)
        

    def query(request):
        print("api/query-data")
        try:


            with connection.cursor() as cursor:
                cursor.execute(f"""SELECT
                    DATENAME(MONTH, CREATED_DATE) AS MonthName,
                    SUM(CASE WHEN YEAR(CREATED_DATE) = YEAR(GETDATE()) - 3 THEN 1 ELSE 0 END) AS Previous_3_Year,
                    SUM(CASE WHEN YEAR(CREATED_DATE) = YEAR(GETDATE()) - 2 THEN 1 ELSE 0 END) AS Previous_2_Year,
                    SUM(CASE WHEN YEAR(CREATED_DATE) = YEAR(GETDATE()) - 1 THEN 1 ELSE 0 END) AS Previous_Year,
                    SUM(CASE WHEN YEAR(CREATED_DATE) = YEAR(GETDATE()) THEN 1 ELSE 0 END) AS Current_Year
                FROM TBL_MESSAGE
                WHERE CREATED_DATE IS NOT NULL and MSG_TYPE = 'CustomizedTrip' 
                    AND AGENT_ID NOT IN (select distinct (Agentid) as Agentid from tbl_agent_culture)
                GROUP BY DATENAME(MONTH, CREATED_DATE), MONTH(CREATED_DATE)
                ORDER BY MONTH(CREATED_DATE)
                            """)
                
                total_query_data = cursor.fetchall()
                total_query_header = [row[0] for row in cursor.description]
                df_total_query = pd.DataFrame(total_query_data, columns=total_query_header)

                current_year = datetime.now().year
                df_total_query = df_total_query.rename(columns=
                                                       {"Previous_3_Year" : current_year - 3,
                                                        "Previous_2_Year" : current_year - 2,
                                                        "Previous_Year" : current_year - 1,
                                                        "Current_Year" : current_year,
                                                        })


                print(df_total_query)



                cursor.execute(f"""SELECT 
                                    a.EMAIL_ID, 
                                    a.CREATED_DATE as QUERY_CREATED_DATE,
                                    b.AgentID
                                FROM 
                                    TBL_MESSAGE a 
                                LEFT JOIN TBL_Agent b ON a.EMAIL_ID = b.Emailid
                                WHERE 
                                    a.CREATED_DATE IS NOT NULL 
                                    AND b.AgentID NOT In (select distinct (Agentid) as Agentid from tbl_agent_culture)   
                                    AND a.MSG_TYPE = 'CustomizedTrip'
                                    AND YEAR(CONVERT(date, a.CREATED_DATE, 103)) >= 2022
                                    AND (b.status = 1 OR b.status IS NULL);
                               """)
                query_data = cursor.fetchall()
                query_data_header = [row[0] for row in cursor.description]
                df_query_data = pd.DataFrame(query_data, columns=query_data_header)
                print(df_query_data.head())


                cursor.execute(f"""select concat(trim(upper(agentid)), convert(date,tourdate,103), PackgID) as UniqueID, 
                                trim(upper(agentid)) as AgentId,
                                MIN(CreatedDate) AS FirstCreatedDate
                            from TBL_BOOKING
                            where txn_msg = 'success'
                                and YEAR(CONVERT(date, CreatedDate, 103)) >= 2022
                                and Is_cancelled <> 1
                                AND tourdate is NOT NULL
                                AND agentId is NOT NULL
                                AND PackgID Is NOT NULL
                                AND CreatedDate != ''
                                AND tourdate != ''
                                AND agentId != ''
                                AND PackgID != ''
                                AND agentId NOT In (select distinct (Agentid) as Agentid from tbl_agent_culture) 
                            GROUP BY 
                                agentId, PackgID, tourdate
                               """)
                booking_data = cursor.fetchall()
                booking_header = [row[0] for row in cursor.description]
                df_booking = pd.DataFrame(booking_data, columns=booking_header)
                print(df_booking.head())



            #if df_query_data['AgentID'] in df_booking['AgentId'] and df_query_data['QUERY_CREATED_DATE'] > df_booking['FirstCreatedDate']:
            
            df_query_data['QUERY_CREATED_DATE'] = pd.to_datetime(df_query_data['QUERY_CREATED_DATE'])
            df_booking['FirstCreatedDate'] = pd.to_datetime(df_booking['FirstCreatedDate'])

            # Normalize AgentID case for reliable merge
            df_query_data['AgentID'] = df_query_data['AgentID'].str.upper().str.strip()
            df_booking['AgentId'] = df_booking['AgentId'].str.upper().str.strip()

            # Merge and filter
            merged_df = pd.merge(df_query_data, df_booking, left_on='AgentID', right_on='AgentId', how='inner')
            filtered_df = merged_df[merged_df['QUERY_CREATED_DATE'] > merged_df['FirstCreatedDate']]

            # Extract year and month
            filtered_df['Year'] = filtered_df['QUERY_CREATED_DATE'].dt.year
            filtered_df['MonthName'] = filtered_df['QUERY_CREATED_DATE'].dt.strftime('%B')


            print(filtered_df)
            years = [current_year - 3, current_year - 2, current_year - 1, current_year]

            filtered_df = filtered_df[filtered_df['Year'].isin(years)]

            # Pivot the table: rows = MonthName, columns = Year, values = number of distinct agents
            pivot_df = filtered_df.pivot_table(
                index='MonthName',
                columns='Year',
                values='AgentID',
                aggfunc='nunique',  # You can use 'count' if you want raw counts instead
                fill_value=0
            )

            # Ensure MonthName is in calendar order
            month_order = ['January', 'February', 'March', 'April', 'May', 'June', 
                        'July', 'August', 'September', 'October', 'November', 'December']
            pivot_df = pivot_df.reindex(month_order).fillna(0)

            # Ensure all year columns exist (some may be missing if no data)
            for y in years:
                if y not in pivot_df.columns:
                    pivot_df[y] = 0

            # Reorder columns as per year order
            pivot_df = pivot_df[[y for y in years]]

            # Reset index to bring MonthName as a column
            pivot_df.reset_index(inplace=True)

            # Optional: convert to int
            pivot_df[years] = pivot_df[years].astype(int)

            print(pivot_df)
            df_total_solved = pivot_df.copy()

            df_query_melted = df_total_query.melt(id_vars='MonthName', var_name='Year', value_name='Total_Query')
            df_solved_melted = df_total_solved.melt(id_vars='MonthName', var_name='Year', value_name='Total_Solved')

            # Merge the two melted DataFrames on MonthName and Year
            merged_df = pd.merge(df_query_melted, df_solved_melted, on=['MonthName', 'Year'], how='outer').fillna(0)

            # Convert values to integers
            merged_df['Total_Query'] = merged_df['Total_Query'].astype(int)
            merged_df['Total_Solved'] = merged_df['Total_Solved'].astype(int)

            # Sort for consistent output
            merged_df['Year'] = merged_df['Year'].astype(str)
            merged_df = merged_df.sort_values(by=['Year', 'MonthName'])

            # Group by year and build nested dictionary
            final_output = []
            month_order = ['January', 'February', 'March', 'April', 'May', 'June', 
                'July', 'August', 'September', 'October', 'November', 'December']

            for year, group in merged_df.groupby('Year'):
                # Reindex months in correct order
                group = group.set_index('MonthName').reindex(month_order).reset_index()

                monthly_data = {
                    row['MonthName']: {
                        "Total_Query": row['Total_Query'],
                        "Total_Solved": row['Total_Solved']
                    }
                    for _, row in group.iterrows()
                }
                final_output.append({year: monthly_data})

            print("Print")
            # Output JSON
            print(json.dumps(final_output, indent=2))

            print("Response")

            return JsonResponse(final_output, safe=False)





        except Exception as e:
            print("Error:", e)
            return JsonResponse({"error": str(e)}, status=500)
        

    def top_10_searches(request):
        print("api/top-10-searches")
        try:
            final_data = []
            days = [1,7,30]
            headers = ["Package Name", "No of Searches"]  

            for day in days:
                with connection.cursor() as cursor:
                    cursor.execute(
                    f"""WITH BaseData AS (
                            SELECT 
                                UPPER(LTRIM(RTRIM(a.PKG_TITLE))) AS pkgName, 
                                TRY_CAST(b.CreatedDate AS DATETIME) AS CreatedDate, 
                                b.AgentId, 
                                a.Country
                            FROM 
                                Tbl_HolidaysSearch b
                            INNER JOIN 
                                TBL_PKG_DETAILS a ON a.PKG_ID = b.pkgID
                            WHERE 
                                a.Status = 1
                                AND b.AgentId IS NOT NULL
                                AND b.AgentId != ''
                                AND b.CREATEDDATE >= DATEADD(DAY, -{day}, GETDATE())
                        ),
                        WithTimeDiff AS (
                            SELECT *,
                                LEAD(CreatedDate) OVER (PARTITION BY AgentId, pkgName ORDER BY CreatedDate ASC) AS NextCreatedDate
                            FROM BaseData
                        ),
                        Filtered AS (
                            SELECT *,
                                DATEDIFF(SECOND, CreatedDate, NextCreatedDate) AS DiffInSeconds
                            FROM WithTimeDiff
                            WHERE NextCreatedDate IS NULL OR ABS(DATEDIFF(SECOND, CreatedDate, NextCreatedDate)) >= 60
                        ),
                        PkgCounts AS (
                            SELECT 
                                pkgName,
                                COUNT(*) AS [count]
                            FROM Filtered
                            GROUP BY pkgName
                        )
                        SELECT TOP 10 *
                        FROM PkgCounts
                        ORDER BY [count] DESC;
                    """)
                    
                    mail_data = cursor.fetchall()
                    headers = [col[0] for col in cursor.description]
                    
                    formatted_rows = [dict(zip(headers, row)) for row in mail_data]

                    final_data.append({
                        "day": day,
                        "data": formatted_rows
                    })
            
            return JsonResponse(final_data, safe=False)


        except Exception as e:
            print("Error:", e)
            return JsonResponse({"error": str(e)}, status=500)
        

    def payment_gateway_report(request):
        print('api/payment-gateway-report')
        try:
            days = [1,7,30]
            summary_dict = defaultdict(dict)

            for day in days:
                if day == 1:
                    query = 'CreatedDate >= DATEADD(HOUR, -24, GETDATE())'
                else:
                    query = f'CreatedDate >= DATEADD(DAY, -{day}, GETDATE())'

                with connection.cursor() as cursor:
                    cursor.execute(f"""
                        SELECT 
                            UPPER(TRIM(txn_msg)) AS txn_msg, 
                            UPPER(TRIM(bank)) AS bank
                        FROM tbl_booking
                        WHERE 
                            {query}
                            AND txn_msg IN ('FAILURE', 'HOLD', 'SUCCESS')
                            AND bank IS NOT NULL
                            AND bank NOT IN ('', 'WALLET', 'BANKDEPOSIT', 'CASHBACK', 'CASH PAYMENT', 'Giftcard')
                            AND agentId NOT In (select distinct (Agentid) as Agentid from tbl_agent_culture)  

                    """)
                    data = cursor.fetchall()
                    header = [col[0] for col in cursor.description]
                    df = pd.DataFrame(data, columns=header)
                
                

                # Aggregate counts
                result = (
                    df.groupby('bank')['txn_msg']
                    .value_counts()
                    .unstack(fill_value=0)
                    .reset_index()
                )



                # Ensure all columns exist
                for col in ['FAILURE', 'HOLD', 'SUCCESS']:
                    if col not in result.columns:
                        result[col] = 0

                # Total & Rate
                result['Total Request'] = result[['FAILURE', 'HOLD', 'SUCCESS']].sum(axis=1)
                result['Success Rate'] = ((result['SUCCESS'] / result['Total Request']) * 100).fillna(0).round(1)

                # Store data column-wise per day
                for _, row in result.iterrows():
                    bank = row['bank']
                    summary_dict[bank][str(day)] = {
                        'FAILURE': int(row.get('FAILURE', 0)),
                        'HOLD': int(row.get('HOLD', 0)),
                        'SUCCESS': int(row.get('SUCCESS', 0)),
                        'Total_Request': int(row['Total Request']),
                        'Success_Rate': f"{row['Success Rate']}%"
                    }


            for bank in summary_dict:
                for day in map(str, days):
                    if day not in summary_dict[bank]:
                        summary_dict[bank][day] = {
                            'FAILURE': 0,
                            'HOLD': 0,
                            'SUCCESS': 0,
                            'Total_Request': 0,
                            'Success_Rate': '0.0%'
                        }
            for bank in summary_dict:
                summary_dict[bank] = dict(sorted(summary_dict[bank].items(), key=lambda x: int(x[0])))

            # Final output
            print(summary_dict)
            ordered_summary = OrderedDict(sorted(summary_dict.items()))

            return JsonResponse(ordered_summary, safe=False)
        

        except Exception as e:
            print("Error:", e)
            return JsonResponse({"error": str(e)}, status=500)


    def new_added_guest(request):
        try:
            with connection.cursor() as cursor:
                cursor.execute(f"""WITH Last24Hours AS (
                                        SELECT DISTINCT PAXID, CREATED_DATE
                                        FROM TBL_TRAVELLER_PAYMENT
                                        WHERE TXN_ID IN (
                                            SELECT txn_id
                                            FROM TBL_BOOKING
                                            WHERE txn_msg = 'success'
                                            AND Is_cancelled <> 1 
                                            AND CREATEDDATE >= DATEADD(DAY, -30, GETDATE())
                                        )
                                    ),
                                    ValidBookings AS (
                                        SELECT DISTINCT QueryID
                                        FROM TBL_BOOKING
                                        WHERE txn_msg = 'success'
                                        and agentid not in (select distinct (Agentid) as Agentid from tbl_agent_culture)
                                    )

                                    SELECT *
                                    FROM (
                                        SELECT
                                            TN.Pkgid,
                                            PD.PKG_TITLE,  -- Include title from package details

                                            COUNT(DISTINCT CASE
                                                WHEN L24.PAXID IS NOT NULL AND TN.PaxDepositAmount > 0 AND TN.PaxDepositAmount < 201 AND L24.CREATED_DATE >= DATEADD(DAY, -1, GETDATE()) THEN TN.TRAV_ID
                                            END) AS Last_24Hrs_Booking,

                                            COUNT(DISTINCT CASE
                                                WHEN L24.PAXID IS NOT NULL AND TN.PaxDepositAmount > 0 AND TN.PaxDepositAmount < 201 AND L24.CREATED_DATE >= DATEADD(DAY, -7, GETDATE()) THEN TN.TRAV_ID
                                            END) AS Last_7Days_Booking,

                                            COUNT(DISTINCT CASE
                                                WHEN L24.PAXID IS NOT NULL AND TN.PaxDepositAmount > 0 AND TN.PaxDepositAmount < 201 AND L24.CREATED_DATE >= DATEADD(DAY, -30, GETDATE()) THEN TN.TRAV_ID
                                            END) AS Last_30Days_Booking

                                        FROM TBL_TRAVELLER_NAME TN
                                        LEFT JOIN Last24Hours L24 ON TN.TRAV_ID = L24.PAXID
                                        INNER JOIN TBL_PKG_DETAILS PD ON TN.Pkgid = PD.PKG_ID
                                        WHERE TN.Status1 = 'Active'
                                        AND TN.PaxDepositAmount > 0
                                        AND CONVERT(DATE, TN.TourDate, 103) >= CONVERT(DATE, GETDATE())
                                        AND TN.PKG_QUERY_ID IN (SELECT QueryID FROM ValidBookings)
                                        GROUP BY TN.Pkgid, PD.PKG_TITLE
                                    ) AS FilteredResults
                                    WHERE Last_24Hrs_Booking > 0
                                    OR Last_7Days_Booking > 0
                                    OR Last_30Days_Booking > 0
                                    ORDER BY CAST(Pkgid AS INT)
                            """)

                mail_data = cursor.fetchall()

            headers = ["Package_ID", "Package_Name", "24Hrs", "7Days", "30Days"] 

            df = pd.DataFrame(mail_data, columns=headers)

            return JsonResponse(df.to_dict(orient='records'), safe=False)



        except Exception as e:
            print("Error:", e)
            return JsonResponse({"error": str(e)}, status=500)


    def inactive_agent(request):
        #http://127.0.0.1:8000/inactive-agent?column=tourdate&day=365
        #http://127.0.0.1:8000/inactive-agent?column=tourdate&day=31-12-2020
        #http://127.0.0.1:8000/inactive-agent?column=tourdate&day=31-12-2020&format=excel
        print("inactive-agent")
        try:
            format = request.GET.get('format')
            column = request.GET.get('column')
            day_param = request.GET.get('day')
            cutoff_date = None
            
            if column.upper() == 'TOURDATE':
                column = 'tourdate'
            elif column.upper() == 'CREATEDDATE':
                column = 'CreatedDate'

            print(day_param)

            if day_param:
                day_param = day_param.strip()

                try:
                    # Try parsing as an integer (number of days ago)
                    days_ago = int(day_param)
                    cutoff_date = datetime.now() - timedelta(days=days_ago)
                except ValueError:
                    try:
                        # Try parsing as a specific date (supports multiple formats if needed)
                        cutoff_date = datetime.strptime(day_param,  "%d-%m-%Y")
                    except ValueError:
                        # Optionally handle bad input gracefully
                        return HttpResponseBadRequest("Invalid date format. Use either a number of days or DD-MM-YYYY format.")


            print(cutoff_date)

            with connection.cursor() as cursor:
                cursor.execute(f"""select upper(trim(agentid)) as agentid, convert(date,tourdate,103) as tourdate, convert(date,CreatedDate,103) as CreatedDate, PackgID
                                    from TBL_BOOKING
                                WHERE txn_msg = 'success' 
                                    AND Is_cancelled <> 1       
                                    AND CreatedDate is NOT NULL
                                    AND tourdate is NOT NULL
                                    AND agentId is NOT NULL
                                    AND PackgID Is NOT NULL
                                    AND CreatedDate != ''
                                    AND tourdate != ''
                                    AND agentId != ''
                                    AND PackgID != ''
                                    AND agentid NOT In (select distinct (Agentid) as Agentid from tbl_agent_culture)
                               """)
                booking_data = cursor.fetchall()
                booking_header = [row[0] for row in cursor.description] 
                booking_df = pd.DataFrame(booking_data, columns=booking_header)

                cursor.execute(f"""select Upper(trim(AgentID)) as AgentID from TBL_Agent
                                    where status = 1 
                                    and AgentID is Not NULL
                                    and AgentID != ''
                                    AND AgentID NOT In (select distinct (Agentid) as Agentid from tbl_agent_culture)
                                    
                               """)
                agent_data = cursor.fetchall()
                agent_header = [row[0] for row in cursor.description]
                agent_df = pd.DataFrame(agent_data, columns=agent_header)

            print(booking_df.shape)
            print(agent_df.shape)


            booking_df['CreatedDate'] = pd.to_datetime(booking_df['CreatedDate'], errors='coerce')
            booking_df['tourdate'] = pd.to_datetime(booking_df['tourdate'], errors='coerce')


            # booking_df['agentid'] = booking_df['agentid'].astype(str).str.upper().str.strip()
            # agent_df['AgentID'] = agent_df['AgentID'].astype(str).str.upper().str.strip()


            booking_df = booking_df.dropna(subset=['agentid', 'PackgID', 'tourdate', 'CreatedDate'])

            booking_df['UniqueID'] = booking_df['agentid'] + '_' + booking_df['PackgID'].astype(str) + '_' + booking_df['tourdate'].dt.strftime('%Y-%m-%d')

            booking_df = booking_df.drop_duplicates(subset=['UniqueID'])

            booking_df = booking_df.drop(columns=['PackgID', 'UniqueID'])


            created_dates = booking_df.groupby('agentid', as_index=False)[f'{column}'].max()

            booking_agents = booking_df['agentid']
            master_agents = agent_df['AgentID']
            all_agent_ids = pd.Series(pd.concat([booking_agents, master_agents]).unique(), name='AgentID')

            

            all_agents_with_dates = pd.merge(
                all_agent_ids.to_frame(), 
                created_dates, 
                left_on='AgentID', right_on='agentid', 
                how='left'
            ).drop(columns=['agentid'])


            # cutoff_date = datetime.now() - timedelta(days=365)

            # cutoff_date = datetime.strptime("01-01-2024", "%d-%m-%Y")

            old_or_null_agents = all_agents_with_dates[
                (all_agents_with_dates[f'{column}'].isna()) | 
                (all_agents_with_dates[f'{column}'] < cutoff_date)
            ].copy()


            old_or_null_agents[f'{column}'] = old_or_null_agents[f'{column}'].dt.strftime('%Y-%m-%d')


            old_or_null_agents = old_or_null_agents[['AgentID', f'{column}']]



            if format == 'excel':
                response = HttpResponse(content_type='text/csv')
                response['Content-Disposition'] = 'attachment; filename="old_or_null_agents.csv"'
                
                old_or_null_agents.to_csv(path_or_buf=response, index=False)
                return response
            

            return JsonResponse(old_or_null_agents.to_dict(orient='records'), safe=False)



        except Exception as e:
            print("Error:", e)
            return JsonResponse({"error": str(e)}, status=500)


    def mail_report(request):
        print("api/mail-report")
        try:
            with connection.cursor() as cursor:
                cursor.execute(f"""
                                SELECT 
                                    CONCAT(trim(upper(RID)),trim(upper(type)),convert(date,mailsenddate,103)) as ID,
                                    min(convert(date,Created_date,103)) as MailReadDate, 
                                    convert(date,mailsenddate,103) as MailSendDate,
                                    DATEDIFF(DAY, convert(date,mailsenddate,103), min(convert(date,Created_date,103))) AS DaysBetweenReadAndSend
                                FROM 
                                    TBL_EMAIL_READNJS
                                where Type != 'messagealertmail'
                                    AND  IsActive = 1
                                    AND Created_date IS NOT NULL
                                    AND MailSendDate IS NOT NULL
                                GROUP BY trim(upper(RID)), trim(upper(type)), convert(date,mailsenddate,103)       
                               """)

                mail_report = cursor.fetchall()
                headers = [col[0] for col in cursor.description]
            df_mail_report = pd.DataFrame(mail_report, columns=headers)
            
            # Convert to JSON
            df_mail_report = df_mail_report.drop(columns=['MailReadDate', 'MailSendDate'])

            df_mail_report = df_mail_report.groupby('DaysBetweenReadAndSend').count().reset_index()

            print(df_mail_report)

            same_day = int(df_mail_report.loc[df_mail_report['DaysBetweenReadAndSend'] == 0, 'ID'].sum())
            next_day = int(df_mail_report.loc[df_mail_report['DaysBetweenReadAndSend'] == 1, 'ID'].sum())
            before_7_days = int(df_mail_report.loc[df_mail_report['DaysBetweenReadAndSend'].between(2, 7), 'ID'].sum())
            after_7_days = int(df_mail_report.loc[df_mail_report['DaysBetweenReadAndSend'] > 7, 'ID'].sum())
            Total = same_day + next_day + before_7_days + after_7_days

            same_day_percentage = round(float(same_day / Total * 100),2) if Total > 0 else 0
            next_day_percentage = round(float(next_day / Total * 100),2) if Total > 0 else 0
            before_7_days_percentage = round(float(before_7_days / Total * 100),2) if Total > 0 else 0
            after_7_days_percentage = round(float(after_7_days / Total * 100),2) if Total > 0 else 0


            summary = [{
                "SameDay": {
                    "Count": same_day,
                    "Percentage": f"{same_day_percentage}%"
                },
                "NextDay": {
                    "Count": next_day,
                    "Percentage": f"{next_day_percentage}%"
                },
                "Between 3rd & 7th Day": {
                    "Count": before_7_days,
                    "Percentage": f"{before_7_days_percentage}%"
                },
                "After7Days": {
                    "Count": after_7_days,
                    "Percentage": f"{after_7_days_percentage}%"
                }
            }]

            return JsonResponse(summary, safe=False)



        except Exception as e:
            print("Error:", e)
            return JsonResponse({"error": str(e)}, status=500)


    def mail_not_read(request):
        print("api/mail-not-read")

        date_filter = ""
        template_filter = ""
        params = []

        start_date = request.GET.get("startdate")
        end_date = request.GET.get("enddate")
        selected_template = request.GET.get("template")

        if end_date:
            end_date = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1)
            end_date = end_date.strftime('%Y-%m-%d')

        if start_date and end_date:
            date_filter = "AND CAST(Created_date AS DATE) BETWEEN %s AND %s"
            params.extend([start_date, end_date])

        if selected_template:
            template_filter = "AND mail_type = %s"
            params.append(selected_template)

        try:
            with connection.cursor() as cursor:

                # Dropdown values
                cursor.execute("""
                    SELECT template_name, template_name AS value 
                    FROM tbl_automatedmail_details 
                    ORDER BY template_name
                """)
                template_dropdown_filter = pd.DataFrame(cursor.fetchall(), columns=[col[0] for col in cursor.description])

                # Mails not read (using EXISTS instead of NOT IN + no CONCAT)
                cursor.execute(f"""
                    SELECT 
                        ID, Emailid, CAST(Created_date AS DATE) AS Created_date
                    FROM 
                        AutoMailSendDeatilsNJS a
                    WHERE 
                        IsActive = 1
                        AND Created_date IS NOT NULL
                        {date_filter}
                        {template_filter}
                        AND NOT EXISTS (
                            SELECT 1
                            FROM TBL_EMAIL_READNJS r
                            WHERE 
                                r.IsActive = 1
                                AND r.Created_date IS NOT NULL
                                AND r.MailSendDate IS NOT NULL
                                AND LOWER(RTRIM(LTRIM(r.RID))) = LOWER(RTRIM(LTRIM(a.Emailid)))
                                AND CAST(r.MailSendDate AS DATE) = CAST(a.Created_date AS DATE)
                                AND LOWER(RTRIM(LTRIM(r.type))) = LOWER(RTRIM(LTRIM(a.mail_type)))
                        )
                    ORDER BY CAST(Created_date AS DATE) DESC
                """, params)
                df_mail_not_read = pd.DataFrame(cursor.fetchall(), columns=[col[0] for col in cursor.description])

                # All mails sent
                cursor.execute(f"""
                    SELECT 
                        ID, Emailid, CAST(Created_date AS DATE) AS Created_date
                    FROM 
                        AutoMailSendDeatilsNJS
                    WHERE 
                        IsActive = 1
                        AND Created_date IS NOT NULL
                        {date_filter}
                        {template_filter}
                    ORDER BY CAST(Created_date AS DATE) DESC
                """, params)
                df_total_send_mail = pd.DataFrame(cursor.fetchall(), columns=[col[0] for col in cursor.description])

            # Group and compute stats
            df_mail_not_read = df_mail_not_read.groupby('Created_date')['ID'].count().reset_index()
            df_mail_not_read.rename(columns={"ID": "No_Of_Mails_Not_Read"}, inplace=True)

            df_total_send_mail = df_total_send_mail.groupby('Created_date')['ID'].count().reset_index()
            df_total_send_mail.rename(columns={
                "Created_date": "MailSendDate",
                "ID": "Total_Send_Mail"
            }, inplace=True)

            df_final = pd.merge(
                df_total_send_mail,
                df_mail_not_read,
                how='left',
                left_on='MailSendDate',
                right_on='Created_date'
            ).drop(columns='Created_date')

            df_final['No_Of_Mails_Not_Read'] = df_final['No_Of_Mails_Not_Read'].fillna(0).astype(int)
            df_final['No_of_Mail_Read'] = df_final['Total_Send_Mail'] - df_final['No_Of_Mails_Not_Read']
            df_final['Mail_Read_Percentage'] = (
                (df_final['No_of_Mail_Read'] / df_final['Total_Send_Mail']) * 100
            ).round(2).astype(str) + '%'

            df_final = df_final.sort_values('MailSendDate', ascending=False)

            result = {
                "template_filter": template_dropdown_filter.to_dict(orient='records'),
                "data": df_final.to_dict(orient='records')
            }

            return JsonResponse(result, safe=False)

        except Exception as e:
            print("Error:", e)
            return JsonResponse({"error": str(e)}, status=500)


    def mail_For_booking(request):
        print("api/mail-for-booking")
        try:
        
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT 
                        TRIM(UPPER(AgentID)) AS AgentID, 
                        TRIM(LOWER(Emailid)) AS Emailid, 
                        CONVERT(DATE, Created_date, 103) AS Created_date 
                    FROM AutoMailSendDeatilsNJS
                    WHERE 
                        mail_type = 'HoldyourSpaces'
                        AND IsActive = 1
                        AND Created_date >= DATEADD(DAY, -45, GETDATE())
                           
                """)
                mail_send_data = cursor.fetchall()

                if mail_send_data == []:
                    print("No Data")
                    return JsonResponse ([], safe=False)
                    
                mail_send_header = [col[0] for col in cursor.description]

            df_mail_send = pd.DataFrame(mail_send_data, columns=mail_send_header)
            

            # Step 2: Group mail send counts per date
            df_group = df_mail_send.groupby('Created_date')['AgentID'].count().reset_index()
            df_group = df_group.rename(columns={'AgentID': 'Total_Send_Mail'})

            

            # Step 3: Fetch all bookings in one query
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT 
                        TRIM(UPPER(AgentID)) AS AgentID, 
                        CONVERT(DATE, CreatedDate, 103) AS BookingDate 
                    FROM tbl_booking
                    WHERE 
                        txn_msg = 'success'
                        AND Is_cancelled <> 1
                        AND PackgID IS NOT NULL AND PackgID != ''
                        AND tourdate IS NOT NULL AND tourdate != ''
                        AND CreatedDate >= DATEADD(DAY, -60, GETDATE())  -- buffer period
                """)
                booking_data = cursor.fetchall()
                booking_header = [col[0] for col in cursor.description]

            df_bookings = pd.DataFrame(booking_data, columns=booking_header)

            print(df_bookings)

            # Step 4: Merge and calculate date differences in memory
            result_rows = []

            for index, row in df_group.iterrows():
                created_date = row['Created_date']
                total_agents = row['Total_Send_Mail']

                # Agents mailed on this date
                agent_list = df_mail_send[df_mail_send['Created_date'] == created_date]['AgentID'].tolist()

                # Filter bookings for these agents
                agent_bookings = df_bookings[df_bookings['AgentID'].isin(agent_list)]

                # Compute the min booking date per agent
                agent_earliest_booking = (
                    agent_bookings[agent_bookings['BookingDate'] >= created_date]
                    .groupby('AgentID')['BookingDate']
                    .min()
                    .reset_index()
                )

                # Compute day difference
                agent_earliest_booking['DayDiff'] = (agent_earliest_booking['BookingDate'] - created_date).apply(lambda x: x.days)

                # Count how many fall in each category
                day7 = (agent_earliest_booking['DayDiff'] < 8).sum()
                day15 = ((agent_earliest_booking['DayDiff'] >= 8) & (agent_earliest_booking['DayDiff'] < 16)).sum()
                day30 = ((agent_earliest_booking['DayDiff'] >= 16) & (agent_earliest_booking['DayDiff'] < 31)).sum()

                result_rows.append({
                    'Created_date': created_date,
                    'Day_7': int(day7),
                    'Day_15': int(day15),
                    'Day_30': int(day30),
                    'Total_Send_Mail': total_agents
                })

            summary_df = pd.DataFrame(result_rows)
            

            summary_df = summary_df.sort_values('Created_date', ascending=False)

            print(summary_df)

            return JsonResponse(summary_df.to_dict(orient='records'), safe=False)

        except Exception as e:
            print("Error:", e)
            return JsonResponse({"error": str(e)}, status=500)


    def new_guest_added_details(request):
        print("api/new-added-guest-details")
        try:
            pkg_id = request.GET.get('pkg_id')
            days = request.GET.get('days')

            print(pkg_id)
            print(days)

            query = ''
            
            if days == 1:
                query = "CreatedDate >= DATEADD(HOUR, -24, GETDATE())"
                

            else:
                query = f"CreatedDate >= DATEADD(DAY, -{days}, GETDATE())"
        
                

            with connection.cursor() as cursor:
                cursor.execute(f"""WITH ValidBookings AS (
                                        SELECT DISTINCT QueryID
                                        FROM TBL_BOOKING
                                        WHERE txn_msg = 'success'
                                    ),
                                    RecentPayments AS (
                                        SELECT DISTINCT PAXID
                                        FROM TBL_TRAVELLER_PAYMENT
                                        WHERE TXN_ID IN (
                                            SELECT txn_id
                                            FROM TBL_BOOKING
                                            WHERE txn_msg = 'success'
                                            AND Is_cancelled <> 1
                                            AND agentid NOT In (select distinct (Agentid) as Agentid from tbl_agent_culture)
                                            AND PackgID = {pkg_id}
                                            AND {query}
                                        )
                                    )

                                    SELECT 
                                        TN.Pkgid,
                                        PD.PKG_TITLE,
                                        TN.T_FNAME,
                                        TN.T_LNAME,
                                        TN.CREATED_BY,
                                        TN.TRAV_ID,
                                        TN.TourDate
                                    FROM TBL_TRAVELLER_NAME TN
                                    INNER JOIN TBL_PKG_DETAILS PD ON TN.Pkgid = PD.PKG_ID
                                    INNER JOIN RecentPayments RP ON TN.TRAV_ID = RP.PAXID
                                    WHERE TN.Status1 = 'Active'
                                    AND TN.PaxDepositAmount BETWEEN 1 AND 200
                                    AND CONVERT(DATE, TN.TourDate, 103) >= CONVERT(DATE, GETDATE())
                                    AND TN.PKG_QUERY_ID IN (SELECT QueryID FROM ValidBookings)
                                    ORDER BY CAST(TN.Pkgid AS INT)
                                    """
                            )
                data = cursor.fetchall()
                headers = [col[0] for col in cursor.description]

            df = pd.DataFrame(data, columns=headers)

            return JsonResponse(df.to_dict(orient='records'), safe=False)
            

        
        except Exception as e:
            print("Error:", e)
            return JsonResponse({"error": str(e)}, status=500)
      
        
    def list_of_distributor_agents(request):
        print("api/get-distributor-agentid")
        try:
            with connection.cursor() as cursor:
                cursor.execute(f"""SELECT DISTINCT A.Name, trim(upper(A.AgentID)) as AgentID
                                    FROM TBL_Agent A
                                    JOIN TBL_Agent B ON trim(upper(B.DistributerID)) = trim(upper(A.AgentID))
                                    WHERE A.Status = 1
                                        AND A.AgentID is NOT NULL
                                        AND A.AgentID != 'NULL'
                                        AND A.AgentID != ''
                                        AND A.AgentID NOT IN (select distinct (Agentid) as Agentid from tbl_agent_culture)
                                    order by A.Name
                        """)

                data = cursor.fetchall()
                header = [desc[0] for desc in cursor.description]
            df = pd.DataFrame(data, columns=header)
            print(df)

            return JsonResponse(df.to_dict(orient='records'), safe=False)

        except Exception as e:
            print("Error:", e)
            return JsonResponse({"error": str(e)}, status=500)



    def total_sub_agents(request):
        print('api/total-sub-agents')
        try:
            distributor_ID = request.GET.get('agentid')
            # distributor_ID = 'CHAGT000001558'

            with connection.cursor() as cursor:
                cursor.execute(f"""select Trim(Name) as Name, Trim(upper(AgentID)) as AgentID from tbl_agent
                                    WHERE Status = 1
                                        AND AgentID is NOT NULL
                                        AND AgentID != 'NULL'
                                        AND AgentID != ''
                                        AND DistributerID = '{distributor_ID}'
									order by Name
                               """)
                
                data = cursor.fetchall()
                headers = [desc[0] for desc in cursor.description]
            sub_agent_details = pd.DataFrame(data, columns=headers)

            sub_agent_details['Name'] = sub_agent_details['Name'].str.title()
            number_of_sub_agent = int(sub_agent_details['AgentID'].count())


            result = [{
                    "total_number_of_sub_agents": number_of_sub_agent,
                    "data": sub_agent_details.to_dict(orient='records')  # Convert this to list-of-dicts
                }]

            return JsonResponse(result, safe=False)
        
        except Exception as e:
            print("Error:", e)
            return JsonResponse({"error": str(e)}, status=500)



    def sub_agent_with_booking(request):
        print("api/sub-agents-booking")

        #distributor_ID = 'CHAGT000001558'
        distributor_ID = request.GET.get('agentid')
        try:
            

            with connection.cursor() as cursor:
                cursor.execute(f"""
                            select trim(upper(AgentID)) as AgentID, trim(Name) as Name, trim(upper(Emailid)) as Emailid, trim(upper(UserName)) as UserName
                            from tbl_agent
                            WHERE Status = 1
                                AND AgentID is NOT NULL
                                AND AgentID != 'NULL'
                                AND AgentID != ''
                                AND DistributerID = '{distributor_ID}'
                               """)
                sub_agent_data = cursor.fetchall()
                sub_agent_header = [row[0] for row in cursor.description]
                
            
                df_sub_agent = pd.DataFrame(sub_agent_data, columns=sub_agent_header)
                df_sub_agent['Name'] = df_sub_agent['Name'].str.title()

                total_sub_agents =  df_sub_agent['AgentID'].nunique()
                print(total_sub_agents)

                
                
                cursor.execute(f"""select distinct(trim(upper(agentid))) as agentid
                                        from TBL_BOOKING
                                        WHERE txn_msg = 'success' 
                                            AND Is_cancelled <> 1       
                                            AND CreatedDate is NOT NULL
                                            AND tourdate is NOT NULL
                                            AND agentId is NOT NULL
                                            AND PackgID Is NOT NULL
                                            AND CreatedDate != ''
                                            AND tourdate != ''
                                            AND agentId != ''
                                            AND PackgID != ''
                                            AND agentId in (select trim(upper(Agentid)) as Agentid from tbl_agent
                                                        WHERE Status = 1
                                                            AND AgentID is NOT NULL
                                                            AND AgentID != 'NULL'
                                                            AND AgentID != ''
                                                            AND DistributerID = '{distributor_ID}'
                                                )
                            """)
                

                sub_agent_with_booking_data = cursor.fetchall()
                sub_agent_with_booking_header = [row[0] for row in cursor.description]

                sub_agent_with_booking_df = pd.DataFrame(sub_agent_with_booking_data, columns=sub_agent_with_booking_header)

                sub_agent_with_booking = sub_agent_with_booking_df['agentid'].nunique()
                print(sub_agent_with_booking)


                sub_agent_without_booking = total_sub_agents - sub_agent_with_booking
                print(sub_agent_without_booking)


              
                cursor.execute(f""" select T_FNAME, T_LNAME, TRAV_ID, trim(upper(CREATED_BY)) as CREATED_BY, TourDate, Pkgid,
                                    (select Name from TBL_Agent where TN.CREATED_BY=AgentID) as Name
                                    from TBL_TRAVELLER_NAME TN
                                    where status1 = 'Active'
                                    and PaxDepositAmount > 99
                                    AND CREATED_BY IN (select distinct(agentid)
                                                        from TBL_BOOKING
                                                        WHERE txn_msg = 'success' 
                                                            AND Is_cancelled <> 1       
                                                            AND CreatedDate is NOT NULL
                                                            AND tourdate is NOT NULL
                                                            AND agentId is NOT NULL
                                                            AND PackgID Is NOT NULL
                                                            AND CreatedDate != ''
                                                            AND tourdate != ''
                                                            AND agentId != ''
                                                            AND PackgID != ''
                                                            AND agentId in (select trim(upper(Agentid)) as Agentid from tbl_agent
                                                                        WHERE Status = 1
                                                                            AND AgentID is NOT NULL
                                                                            AND AgentID != 'NULL'
                                                                            AND AgentID != ''
                                                                            AND DistributerID = '{distributor_ID}'))
                                                        """)
                
                data = cursor.fetchall()
                headers = [desc[0] for desc in cursor.description]
                sub_agent_guest_df = pd.DataFrame(data, columns=headers)



                cursor.execute(f"""select distinct(upper(trim(EMAIL_ID))) as EMAIL_ID  
                                    from TBL_MESSAGE
                                    Where CREATED_DATE IS NOT NULL 
                                    AND AGENT_ID NOT IN (select distinct (Agentid) as Agentid from tbl_agent_culture)
                                    AND len(MSG_DETAILS) > 10
                             """)

                df_query_data = cursor.fetchall()
                df_query_header = [row[0] for row in cursor.description]

                df_query = pd.DataFrame(df_query_data, columns=df_query_header)


                cursor.execute("""select Distinct(trim(upper(AGENTID))) as AGENTID 
                                from TBL_LOGIN
                                where Loginby is NULL 
                                AND LOGINDATE is NOT NULL
                                AND AGENTID is NOT NULL
                                AND AGENTID != ''
                                AND AgentID NOT In (select distinct (Agentid) as Agentid from tbl_agent_culture)               
                            """)
                
                df_login_data =  cursor.fetchall()
                df_login_header = [row[0] for row in cursor.description]
            
                df_login = pd.DataFrame(df_login_data, columns=df_login_header)

        
                




            GuestCount_list_of_sub_agents_with_booking = sub_agent_guest_df.groupby(['CREATED_BY','Name'])['TRAV_ID'].count().sort_values(ascending=False).reset_index(name='No_of_Guests')


            print("-----ALL Agents-----")
            print(df_sub_agent)
            print("-------------------")
            print("-----Agents with Booking-----")
            print(sub_agent_with_booking_df)

            GuestCount_list_of_sub_agents_without_booking = df_sub_agent[~df_sub_agent['AgentID'].isin(sub_agent_with_booking_df['agentid'])].reset_index(drop=True)

            

                
            agent_df_with_query = df_sub_agent[df_sub_agent['Emailid'].isin(df_query['EMAIL_ID'])]
            print("------------------------agent_df_with_query-----------------------------------")
            print(agent_df_with_query)
            print("--------------------------sub_agent_with_booking_df---------------------------------")
            print(sub_agent_with_booking_df)

            agents_with_query_but_no_booking = agent_df_with_query.loc[~agent_df_with_query['AgentID'].isin(sub_agent_with_booking_df['agentid']), ['AgentID', 'Name', 'Emailid']].reset_index(drop=True)

            print("--------------------------agents_with_query_but_no_booking---------------------------------")
            print(agents_with_query_but_no_booking)


            agent_without_login = df_sub_agent.loc[~df_sub_agent['UserName'].isin(df_login['AGENTID']), ['AgentID', 'Name', 'Emailid']].reset_index(drop=True)


            result = [{
                "total_number_of_sub_agents": total_sub_agents,
                "number_of_sub_agents_with_booking" : sub_agent_with_booking,
                "number_of_sub_agents_without_booking" : sub_agent_without_booking,
                "number_of_sub_agents_with_query_but_no_booking" : agents_with_query_but_no_booking.shape[0],
                "number_of_sub_agents_without_login" : agent_without_login.shape[0],
                "GuestCount_list_of_sub_agents_with_booking" : GuestCount_list_of_sub_agents_with_booking.to_dict(orient='records'),
                "GuestCount_list_of_sub_agents_without_booking" : GuestCount_list_of_sub_agents_without_booking.to_dict(orient='records'),
                "list_of_sub_agents_with_query_but_no_booking" : agents_with_query_but_no_booking.to_dict(orient='records'),
                "list_of_sub_agents_without_login" : agent_without_login.to_dict(orient='records')

            }]

            return JsonResponse(result, safe=False)
        

        except Exception as e:
            print("Error:", e)
            return JsonResponse({"error": str(e)}, status=500)





    def get_agentid_list(request):
        print("api/get-agentid-list")
        try:
            with connection.cursor() as cursor:
                cursor.execute(f""" select AgentID, trim(Name) as Name 
                                FROM TBL_Agent
                                where status = 1
                                AND Name is NOT NULL
                                AND Name != ''
                                AND AgentID is NOT NULL
                                AND AgentID != ''
                                AND AgentID NOT In (select distinct (Agentid) as Agentid from tbl_agent_culture)   
                """)
                agent_data = cursor.fetchall()
                agent_header = [row[0] for row in cursor.description]

            df_agent = pd.DataFrame(agent_data, columns=agent_header)


            df_agent['Name'] = df_agent['Name'].str.title()
                

            return JsonResponse(df_agent.to_dict(orient='records'), safe=False)

        except Exception as e:
            print("Error:", e)
            return JsonResponse({"error": str(e)}, status=500)




    def get_agentid_list_2(request):
        print("api/get-agentid-list-2")
        data= request.GET.get("id","")
        try:
            with connection.cursor() as cursor:
                cursor.execute(f""" select top 20 AgentID, trim(Name) as Name 
                                    FROM TBL_Agent
                                    where status = 1
                                        AND Name is NOT NULL
                                        AND Name != ''
                                        AND AgentID is NOT NULL
                                        AND AgentID != ''
                                        AND AgentID NOT In (select distinct (Agentid) as Agentid from tbl_agent_culture)
                                        AND (AgentID LIKE '%{data}%' or Name LIKE '%{data}%')
                                    order by name, AgentID
                """)
                executed_query = connection.queries[-1]['sql'] if connection.queries else "No query executed"
                print(executed_query)

                agent_data = cursor.fetchall()
                agent_header = [row[0] for row in cursor.description]

            df_agent = pd.DataFrame(agent_data, columns=agent_header)


            df_agent['Name'] = df_agent['Name'].str.title()
                

            return JsonResponse(df_agent.to_dict(orient='records'), safe=False)

        except Exception as e:
            print("Error:", e)
            return JsonResponse({"error": str(e)}, status=500)





    def agent_wallet_summary(request):
        print("api/wallet-summary")

        agentid = request.GET.get("agentid")
        #agentid = 'CHAGT000003780'  #'CHAGT0001000022962' 
        

        try:
            with connection.cursor() as cursor:
                cursor.execute("""
                                SELECT 
                                    TRIM(UPPER(w.Txn_ID)) AS Txn_ID,
                                    w.Amount,
                                    w.PaymentMode,
                                    w.CreatedDate,
                                    w.PaymentMessage,
                                    trim(upper(b.PackageName)) as PackageName,
                                    CONVERT(DATE, b.TourDate, 103) AS TourDate
                                FROM Tbl_Wallet w
                                OUTER APPLY (
                                    SELECT TOP 1 
                                        PackageName, 
                                        TourDate
                                    FROM TBL_BOOKING b
                                    WHERE 
                                        b.Txn_ID = w.Txn_ID
                                        AND w.PaymentMode = 'Debit'
                                ) b
                                WHERE 
                                    w.AgentID = %s
                                    AND w.BookingStatus = 'Success'
                                    AND w.Amount > 0
                                    AND w.Txn_ID IS NOT NULL
                                    AND w.Txn_ID != ''
                                ORDER BY w.CreatedDate;
                               """,[agentid])
            
            
                executed_query = connection.queries[-1]['sql'] if connection.queries else "No query executed"
                print(executed_query)   


                wallet_data = cursor.fetchall()
                wallet_header = [row[0] for row in cursor.description]
            
            df_wallet_data = pd.DataFrame(wallet_data, columns=wallet_header)

            

            print(df_wallet_data)
            
            if df_wallet_data.empty:
                print("⚠️ No wallet transactions found.")
            else:
                # df_wallet_data['PaymentMode'] = df_wallet_data['PaymentMode'].str.title()
                # df_wallet_data['PaymentMode'] = df_wallet_data['PaymentMode'].astype(str)
                df_wallet_data['PaymentMode'] = df_wallet_data['PaymentMode'].astype(str).str.strip().str.title()
                #df_wallet_data['PaymentMessage'] = df_wallet_data['PaymentMessage'].astype(str).str.strip().str.title()
                
                df_wallet_data['Amount'] = df_wallet_data['Amount'].astype(float)
                df_wallet_data['CreatedDate'] = pd.to_datetime(df_wallet_data['CreatedDate'])

          

                balance = 0
                balances = []
                messages = []
                statuses = []

                for _, row in df_wallet_data.iterrows():
                    mode = row['PaymentMode'].strip().lower()
                    amount = row['Amount']
                    message = row['PaymentMessage']
                    status = ""
                    package_name = row['PackageName']
                    tour_date = row['TourDate']

                    

                    # Update balance
                    if mode == 'credit':
                        balance += amount
                    elif mode == 'debit':
                        balance -= amount
                    balances.append(balance)

                    # Fix PaymentMessage if needed
                    if mode == 'credit' and (pd.isna(message) or str(message).strip() == ""):
                        message = "Amount Added by Agent"

                    elif mode == 'debit':
                        if message and (message.lower().startswith('300') or 'choffer' in message.lower()) and amount == 300:
                            message = "Joining Offer Amount Expire"
                            
                        else:
                            message = f'Amount is Used for Package: "{package_name}" of Tour Date: {tour_date}'
                            status = "Amount Debit"

                    
        
                    messages.append(message)
                    statuses.append(status)
                    
                

                # Assign both columns at once
                df_wallet_data['Balance'] = balances
                df_wallet_data['PaymentMessage'] = messages
                df_wallet_data['More_Info'] = statuses

                # Final output formatting
                df_result = df_wallet_data.sort_values(by='CreatedDate', ascending=False).reset_index(drop=True)
                df_result = df_result.rename(columns={'CreatedDate': 'Txn_Date'}) 

                df_wallet_data = df_result[['Txn_Date', 'Txn_ID', 'PaymentMode', 'Amount', 'Balance', 'PaymentMessage', 'More_Info']]

                print(df_wallet_data)


                df_wallet_data['Txn_Date'] = pd.to_datetime(df_wallet_data['Txn_Date'], errors='coerce')
                df_wallet_data['Txn_Date'] = df_wallet_data['Txn_Date'].dt.strftime('%Y-%m-%d %H:%M:%S')  # keep full datetime

                final_rows = []
                seen_dates = set()

                for i, row in df_wallet_data.iterrows():
                    txn_datetime = pd.to_datetime(row['Txn_Date'])  # safely parse again
                    txn_date = txn_datetime.date()

                    # If Closing Balance for this date not yet added
                    if txn_date not in seen_dates:
                        closing_row = {
                            'Txn_Date': txn_date.strftime('%Y-%m-%d'),  # only the date
                            'Txn_ID': 'Closing Balance',
                            'PaymentMode': 'None',
                            'Amount': 'None',
                            'Balance': row['Balance'],
                            'PaymentMessage': '',
                            'More_Info' : 'None'
                        }
                        final_rows.append(closing_row)
                        seen_dates.add(txn_date)

                    # Add original transaction row
                    final_rows.append({
                        'Txn_Date': txn_datetime.strftime('%Y-%m-%d %H:%M:%S'),
                        'Txn_ID': row['Txn_ID'],
                        'PaymentMode': row['PaymentMode'],
                        'Amount': row['Amount'],
                        'Balance': row['Balance'],
                        'PaymentMessage': row['PaymentMessage'],
                        'More_Info' : row['More_Info']
                    })

                # Final DataFrame
                df_result = pd.DataFrame(final_rows)

                df_result.reset_index(inplace=True)

                
                print(df_result)

                return JsonResponse(df_result.to_dict(orient='records'), safe=False)

        except Exception as e:
            print("Error:", e)
            return JsonResponse({"error": str(e)}, status=500)
    


    def guest_payment_by_wallet(request):
        print("api/guest-payment-by-wallet")
        
        txn_id = request.GET.get("txnid")
        try:
            with connection.cursor() as cursor:
                cursor.execute("""
                                select 
                                    TN.TRAV_ID, 
                                    TRIM(TN.T_FNAME) as T_FNAME, 
	                                TRIM(TN.T_LNAME) as T_LNAME,  
                                    TRIM(UPPER(TPD.PKG_TITLE)) AS PKG_TITLE,
                                    TRIM(upper(TN.CREATED_BY)) as AgentID, 
                                    convert(date,TN.TourDate,103) as TourDate,  
                                    TN.Pkgid, 
                                    TP.PaxDepositAmount
                                FROM 
                                    TBL_TRAVELLER_PAYMENT TP 
                                JOIN 
                                    TBL_TRAVELLER_NAME TN ON TP.PAXID = TN.TRAV_ID
                                JOIN 
                                    TBL_PKG_DETAILS TPD ON TN.Pkgid = TPD.PKG_ID
                                where TP.TXN_ID = %s
                               """,[txn_id])
                df_data = cursor.fetchall()
                df_header = [col[0] for col in cursor.description]

            df_guest_details = pd.DataFrame(df_data, columns=df_header)

            df_guest_details['T_FNAME'] = df_guest_details['T_FNAME'].str.title()
            df_guest_details['T_LNAME'] = df_guest_details['T_LNAME'].str.title()
            df_guest_details['PKG_TITLE'] = df_guest_details['PKG_TITLE'].str.title()
            df_guest_details['PaxDepositAmount'] = df_guest_details['PaxDepositAmount'].map(lambda x: f"{x:.2f}")

            print(df_guest_details)

            return JsonResponse(df_guest_details.to_dict(orient='records'), safe=False)
        
        except Exception as e:
            print("Error:", e)
            return JsonResponse({"error": str(e)}, status=500)


    def flyer_region(request):
        print("api/flyer-search")
        
        selected_region = request.GET.get("region")
        selected_pkgcountry = request.GET.get("country")


        try:
            with connection.cursor() as cursor:
                cursor.execute(f"""
                            select FA.AgentID, FA.agent_country, FA.agent_region, FA.agent_city, FA.pkgId, FA.pkgCountry, TPD.PKG_TITLE
                            from tbl_flyer_analytics FA
								JOIN TBL_PKG_DETAILS TPD ON FA.pkgId = TPD.PKG_ID
                            where FA.AgentID IS NOT NULL
                                AND FA.agent_country IS NOT NULL
                                AND FA.agent_region IS NOT NULL
                                AND FA.agent_city IS NOT NULL
                                AND FA.pkgId IS NOT NULL
                                AND FA.pkgCountry IS NOT NULL
                                AND FA.type = 'flyer'
                                AND FA.AgentID != ''
                                AND FA.agent_country NOT IN ('', '---')
                                AND FA.agent_region NOT IN ('', '---')
                                AND FA.agent_city NOT IN ('', '---')
                                AND FA.pkgId NOT IN ('', '---')
                                AND FA.pkgCountry NOT IN ('', '---')
                """)
                flyer_data = cursor.fetchall()
                flyer_header = [col[0] for col in cursor.description]
            df_flyer = pd.DataFrame(flyer_data, columns=flyer_header)


            # df_flyer = df_flyer[['AgentID','agent_country', 'agent_region', 'agent_city', 'pkgId', 'pkgCountry' ]]


            df_flyer['AgentID'] = df_flyer['AgentID'].str.upper().str.strip()
            df_flyer['agent_country'] = df_flyer['agent_country'].str.upper().str.strip()
            df_flyer['agent_region'] = df_flyer['agent_region'].str.upper().str.strip()
            df_flyer['agent_city'] = df_flyer['agent_city'].str.upper().str.strip()
            df_flyer['pkgCountry'] = df_flyer['pkgCountry'].str.upper().str.strip()
            df_flyer['PKG_TITLE'] = df_flyer['PKG_TITLE'].str.title().str.strip()



            if selected_region:
                df_flyer = df_flyer[df_flyer['agent_region'] == selected_region].reset_index()

                df_region = df_flyer.groupby('pkgCountry')['AgentID'].count().sort_values(ascending=False).reset_index(name='Flyer_Count')

                print("-----Selected Region-----")
                print(df_region)

                if selected_pkgcountry:
                    df_flyer = df_flyer[df_flyer['pkgCountry'] == selected_pkgcountry].reset_index()


                    df_pkg_country = df_flyer.groupby('PKG_TITLE')['AgentID'].count().sort_values(ascending=False).reset_index(name='Flyer_Count')


                    print("-----Selected Country-----")
                    print(df_pkg_country)

                    return JsonResponse(df_pkg_country.to_dict(orient='records'), safe=False) 
                

                return JsonResponse(df_region.to_dict(orient='records'), safe=False) 
            

            print("-----Main DF-----")
            print(df_flyer)
            
            return JsonResponse(df_flyer['agent_region'].dropna().unique().tolist(), safe=False)
         



        except Exception as e:
            print("Error:", e)
            return JsonResponse({"error": str(e)}, status=500)



    def daily_mailsend_report(request):
        print('api/daily-mailsend-report')

        params = []
        date_filter = ""
        date_filter_auto = ""

        start_date = request.GET.get("startdate")
        end_date = request.GET.get("enddate")
        single_date = request.GET.get("singledate")

        print("Start:", start_date, "End:", end_date, "Single:", single_date)

        if end_date:
            end_date = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1)
            end_date = end_date.strftime('%Y-%m-%d')

        if single_date:
            date_filter = "AND CONVERT(date, b.Created_date, 103) = %s"
            date_filter_auto = "AND CONVERT(date, Created_date, 103) = %s"
            params.append(single_date)

        elif start_date and end_date:
            date_filter = "AND CONVERT(date, b.Created_date, 103) BETWEEN %s AND %s"
            date_filter_auto = "AND CONVERT(date, Created_date, 103) BETWEEN %s AND %s"
            params.extend([start_date, end_date])

        try:
            with connection.cursor() as cursor:
                # ✅ Successful Mails
                cursor.execute(f"""
                    SELECT 
                        a.TemplateDisplayName, 
                        a.template_name, 
                        COUNT(b.Emailid) AS send_mail
                    FROM tbl_automatedmail_details a 
                    LEFT JOIN AutoMailSendDeatilsNJS b ON a.template_name = b.mail_type
                        {date_filter}
                        AND b.Emailid IS NOT NULL AND b.Emailid != ''
                    WHERE a.template_name != 'messagealertmail'
                    GROUP BY a.TemplateDisplayName, a.template_name
                    ORDER BY a.TemplateDisplayName
                """, params)
                df_sent = pd.DataFrame(cursor.fetchall(), columns=[col[0] for col in cursor.description])

                # ✅ Failed Mails
                cursor.execute(f"""
                    SELECT 
                        a.TemplateDisplayName, 
                        a.template_name, 
                        COUNT(b.Emailid) AS number_of_failed_mail
                    FROM tbl_automatedmail_details a 
                    LEFT JOIN AutoMailSendDeatilsOnError b ON a.template_name = b.mail_type
                        {date_filter}
                        AND b.Emailid IS NOT NULL AND b.Emailid != ''
                    WHERE a.template_name != 'messagealertmail'
                    GROUP BY a.TemplateDisplayName, a.template_name
                    ORDER BY a.TemplateDisplayName
                """, params)
                df_failed = pd.DataFrame(cursor.fetchall(), columns=[col[0] for col in cursor.description])

                # ✅ Mail Not Read
                cursor.execute(f"""
                    SELECT mail_type, COUNT(*) AS No_Of_Mails_Not_Read
                    FROM AutoMailSendDeatilsNJS
                    WHERE IsActive = 1 AND Created_date IS NOT NULL
                        {date_filter_auto}
                        AND mail_type != 'messagealertmail'
                        AND CONCAT(TRIM(LOWER(Emailid)), CONVERT(date, Created_date, 103), TRIM(LOWER(mail_type))) NOT IN (
                            SELECT CONCAT(TRIM(LOWER(RID)), CONVERT(date, MailSendDate, 103), TRIM(LOWER(type)))
                            FROM TBL_EMAIL_READNJS
                            WHERE IsActive = 1 AND Created_date IS NOT NULL AND MailSendDate IS NOT NULL
                        )
                    GROUP BY mail_type
                """, params)
                df_not_read = pd.DataFrame(cursor.fetchall(), columns=['mail_type', 'No_Of_Mails_Not_Read'])

                # ✅ Total Sent
                cursor.execute(f"""
                    SELECT mail_type, COUNT(*) AS Total_Send_Mail
                    FROM AutoMailSendDeatilsNJS
                    WHERE IsActive = 1 AND Created_date IS NOT NULL
                        {date_filter_auto}
                        AND mail_type != 'messagealertmail'
                    GROUP BY mail_type
                """, params)
                df_total = pd.DataFrame(cursor.fetchall(), columns=['mail_type', 'Total_Send_Mail'])

            # ✅ Compute Read Count
            df_read = pd.merge(df_total, df_not_read, on='mail_type', how='left').fillna(0)
            df_read['No_Of_Mails_Not_Read'] = df_read['No_Of_Mails_Not_Read'].astype(int)
            df_read['No_of_Mail_Read'] = df_read['Total_Send_Mail'] - df_read['No_Of_Mails_Not_Read']
            df_read = df_read[['mail_type', 'No_of_Mail_Read']]

            # ✅ Merge All
            df_final = pd.merge(df_sent, df_failed, on='template_name', how='left').fillna(0)
            df_final = pd.merge(df_final, df_read, left_on='template_name', right_on='mail_type', how='left').fillna(0)
            df_final.drop(columns=['mail_type'], inplace=True)

            df_final = df_final[['TemplateDisplayName_x', 'template_name', 'send_mail','number_of_failed_mail', 'No_of_Mail_Read']]

            df_final = df_final.rename(columns={'TemplateDisplayName_x': 'TemplateDisplayName'})


            return JsonResponse(df_final.to_dict(orient='records'), safe=False)

        except Exception as e:
            print("Error:", e)
            return JsonResponse({"error": str(e)}, status=500)





    def wallet_transaction_details(request):
        print("api/wallet-transaction-details")

        start_date = request.GET.get("startdate")
        end_date = request.GET.get("enddate")
        single_date = request.GET.get("singledate")
        show_data = request.GET.get('data')

        params = []
        date_filter = ""

        
        if end_date:
            end_date = datetime.strptime(end_date, "%Y-%m-%d")
            end_date = end_date + timedelta(days=1)

        if single_date:
            date_filter = "AND CONVERT(date, tw.CreatedDate, 103) = %s"
            params.append(single_date)
        elif start_date and end_date:
            date_filter = "AND CONVERT(date, tw.CreatedDate, 103) BETWEEN %s AND %s"
            params.extend([start_date, end_date])

        print(single_date)
        print(start_date)
        print(end_date)

        print(date_filter)
        print(params)
        print(show_data)

        try:
            with connection.cursor() as cursor:
                # Total Credit
                cursor.execute(f"""
                    SELECT COUNT(DISTINCT tw.AgentID) AS No_of_Agents, 
                        ISNULL(SUM(CAST(tw.Amount AS FLOAT)), 0) AS Total_Credit_Amount
                    FROM Tbl_Wallet tw
                    WHERE tw.PaymentMode = 'Credit'
                        AND tw.AgentID NOT in (select distinct (Agentid) as Agentid from tbl_agent_culture)
                        {date_filter}
                """, params)
                df_full_credit = pd.DataFrame(cursor.fetchall(), columns=[col[0] for col in cursor.description])

                # print(df_full_credit)

                # Total Debit
                cursor.execute(f"""
                    SELECT COUNT(DISTINCT tw.AgentID) AS No_of_Agents, 
                        ISNULL(SUM(CAST(tw.Amount AS FLOAT)), 0) AS Total_Debit_Amount
                    FROM Tbl_Wallet tw
                    WHERE tw.PaymentMode = 'Debit'
                        AND tw.AgentID NOT in (select distinct (Agentid) as Agentid from tbl_agent_culture)
                    {date_filter}
                """, params)
                df_full_debit = pd.DataFrame(cursor.fetchall(), columns=[col[0] for col in cursor.description])

                # print(df_full_debit)

                # Actual Credit
                cursor.execute(f"""
                    SELECT COUNT(DISTINCT tw.AgentID) AS No_of_Agents, 
                        ISNULL(SUM(CAST(tw.Amount AS FLOAT)), 0) AS Actual_Credit_Amount
                    FROM Tbl_Wallet tw
                    WHERE tw.PaymentMode = 'Credit'
                    AND (tw.PaymentMessage IS NULL OR (tw.PaymentMessage NOT LIKE '%%$300%%' AND tw.PaymentMessage NOT LIKE 'CHOFFER'))
                        AND tw.AgentID NOT in (select distinct (Agentid) as Agentid from tbl_agent_culture)
                    {date_filter}
                """, params)

                df_actual_credit = pd.DataFrame(cursor.fetchall(), columns=[col[0] for col in cursor.description])

                # print(df_actual_credit)

                # Actual Debit
                cursor.execute(f"""
                    SELECT COUNT(DISTINCT tw.AgentID) AS No_of_Agents, 
                        ISNULL(SUM(CAST(tw.Amount AS FLOAT)), 0) AS Actual_Debit_Amount
                    FROM Tbl_Wallet tw
                    WHERE tw.PaymentMode = 'Debit'
                    AND (tw.PaymentMessage IS NULL OR (tw.PaymentMessage NOT LIKE '300%%' AND tw.PaymentMessage NOT LIKE 'CHOFFER'))
                        AND tw.AgentID NOT in (select distinct (Agentid) as Agentid from tbl_agent_culture)
                    {date_filter}
                """, params)
                df_actual_debit = pd.DataFrame(cursor.fetchall(), columns=[col[0] for col in cursor.description])

            result = {
                'Total_Credit_Details': df_full_credit.to_dict(orient='records'),
                'Total_Debit_Details': df_full_debit.to_dict(orient='records'),
                'Actual_Credit_Details': df_actual_credit.to_dict(orient='records'),
                'Actual_Debit_Details': df_actual_debit.to_dict(orient='records'),
            }

            print(result)

            if show_data:
                label_map = {
                    "Total_Credit": {
                        "mode": "Credit",
                        "extra": ""
                    },
                    "Total_Debit": {
                        "mode": "Debit",
                        "extra": ""
                    },
                    "Actual_Credit": {
                        "mode": "Credit",
                        "extra": f"AND (tw.PaymentMessage IS NULL OR (tw.PaymentMessage NOT LIKE '%%$300%%' AND tw.PaymentMessage NOT LIKE 'CHOFFER'))"
                    },
                    "Actual_Debit": {
                        "mode": "Debit",
                        "extra": f"AND (tw.PaymentMessage IS NULL OR (tw.PaymentMessage NOT LIKE '300%%' AND tw.PaymentMessage NOT LIKE 'CHOFFER'))"
                    }
                }

                if show_data in label_map:
                    with connection.cursor() as cursor:
                        entry = label_map[show_data]
                        cursor.execute(f"""
                            SELECT TRIM(UPPER(tw.AgentID)) AS AgentID, 
                                ta.Name, 
                                ISNULL(SUM(CAST(tw.Amount AS FLOAT)), 0) AS Amount
                            FROM Tbl_Wallet tw
                            JOIN tbl_agent ta ON tw.AgentID = ta.AgentID
                            WHERE tw.PaymentMode = %s
                            AND tw.AgentID NOT in (select distinct (Agentid) as Agentid from tbl_agent_culture)
                            {entry['extra']}
                            {date_filter}
                            GROUP BY tw.AgentID, ta.Name
                        """, [entry["mode"], *params])

                        executed_query = connection.queries[-1]['sql'] if connection.queries else "No query executed"
                        print(executed_query)   


                        data = cursor.fetchall()
                        headers = [col[0] for col in cursor.description]
                        df_detail = pd.DataFrame(data, columns=headers)

                    df_detail['Name'] = df_detail['Name'].str.title()
                    print(df_detail)
                    return JsonResponse(df_detail.to_dict(orient='records'), safe=False)

            
            print(result)
            return JsonResponse(result, safe=False)

        except Exception as e:
            print("Error:", e)
            return JsonResponse({"error": str(e)}, status=500)


            
    def customize_sales_report(request):
        print('api/customize-sales-report')

        msg_type = request.GET.get("msg_type")
        start_date = request.GET.get("startdate")
        end_date = request.GET.get("enddate")
        selected_country = request.GET.get("country")
        selected_salesid = request.GET.get("salesid") 

        # selected_salesid = 'CH0148'

        print("Message Type : ", msg_type)

        end_date = datetime.strptime(end_date, "%Y-%m-%d")

        if end_date:
            end_date = end_date + timedelta(days=1)


        msg_type_filter = f"AND MSG_TYPE = '{msg_type}'" if msg_type else ""
        assign_date_filter = f"AssignDate BETWEEN '{start_date}' AND '{end_date}'" if start_date and end_date else ""
        msg_date_filter = f"CONVERT(date, CREATED_DATE, 103) BETWEEN '{start_date}' AND '{end_date}'" if start_date and end_date else ""
        booking_date_filter= f"HAVING MIN(CreatedDate) BETWEEN '{start_date}' AND '{end_date}'" if start_date and end_date else ""
        msg_country_filter = f"AND Country = '{selected_country}'" if selected_country else ""
        booking_country_filter = f" AND PD.country = '{selected_country}'" if selected_country else ""

        try:
            with connection.cursor() as cursor:
                df_country_filter = pd.DataFrame()
                if msg_date_filter or msg_type_filter:
                    cursor.execute(f"""
                        SELECT DISTINCT trim(Country) as Country
                        FROM TBL_MESSAGE
                        WHERE {msg_date_filter}
                        {msg_type_filter}
                        order by country
                    """)

                    executed_query = connection.queries[-1]['sql'] if connection.queries else "No query executed"
                    print(executed_query)  
                
                    
                    df_country_filter = pd.DataFrame(cursor.fetchall(), columns=[row[0] for row in cursor.description]).dropna()
                    print(df_country_filter)

                    df_country_filter['value'] = df_country_filter['Country']

                cursor.execute(f"""select count(*) from TBL_MESSAGE
                        where {msg_date_filter}
                        --AND AGENT_ID not in (select distinct (Agentid) as Agentid from tbl_agent_culture)
                        """)
                total_query_count = cursor.fetchone()[0]


                cursor.execute(f"""
                    SELECT MSG_ID, TRIM(UPPER(Agent_id)) AS Agent_ID, CREATED_DATE, TRIM(UPPER(AssignTo)) AS AssignTo
                    FROM TBL_MESSAGE
                    WHERE {assign_date_filter}
                    {msg_type_filter}
                    {msg_country_filter}
                    --AND AGENT_ID not in (select distinct (Agentid) as Agentid from tbl_agent_culture)
                    ORDER BY MSG_ID DESC
                """)

                executed_query = connection.queries[-1]['sql'] if connection.queries else "No query executed"
                print(executed_query)  


                df_query_data = pd.DataFrame(cursor.fetchall(), columns=[row[0] for row in cursor.description])

                # total_query_count = df_query_data.shape[0]

                df_query_data = df_query_data[df_query_data['AssignTo'].notna() & (df_query_data['AssignTo'] != '')] 

                print("df_query_data")
                print(df_query_data)

                if df_query_data.empty:
                    print("Blank")
                    return JsonResponse("Currently not a single query is assign to Anyone", safe=False)

               
                query_count_by_staff = (
                    df_query_data.groupby('AssignTo')['MSG_ID']
                    .count()
                    .reset_index()
                    .rename(columns={"MSG_ID": "No_of_Query_Assign", "AssignTo": "UserID"})
                )
                query_count_by_staff['Query_Percentage'] = (
                    query_count_by_staff['No_of_Query_Assign'] / total_query_count * 100
                )

                

                staff_ids = query_count_by_staff['UserID'].astype(str).str.upper().tolist()
                staff_placeholders = ', '.join(['%s'] * len(staff_ids))
                cursor.execute(f"""
                    SELECT TRIM(UPPER(UserID)) AS UserID, Trim(StaffName) as StaffName, TRIM(UPPER(Empid)) AS SalesID
                    FROM tblstaff
                    WHERE UserID IN ({staff_placeholders})
                """, staff_ids)

                staff_data = pd.DataFrame(cursor.fetchall(), columns=[row[0] for row in cursor.description])

                query_count_by_staff = pd.merge(query_count_by_staff, staff_data , left_on='UserID', right_on='UserID', how='left')

                print(query_count_by_staff)
            
                print("Working")

                cursor.execute(f"""
                               SELECT 
                            TB.*, TRIM(UPPER((select salesid from TBL_PKG_DETAILS where pkg_id = PackgID and status =1))) AS SalesID
                            FROM (
                                    SELECT 
                                    PackgID,
                                    MIN(CreatedDate) AS FirstCreatedDate
                                    FROM TBL_BOOKING
                                    WHERE txn_msg = 'success'
                                        AND YEAR(CONVERT(date, CreatedDate, 103)) >= 2022
                                        AND Is_cancelled <> 1
                                        AND tourdate IS NOT NULL
                                        AND agentId IS NOT NULL
                                        AND PackgID IS NOT NULL
                                        AND CreatedDate != ''
                                        AND tourdate != ''
                                        AND agentId != ''
                                        AND PackgID != ''
                                        and PackgID in(select PD.pkg_id from TBL_PKG_DETAILS PD  WHERE PD.SalesID IS NOT NULL AND PD.SalesID <> '' and Status=1 {booking_country_filter})
                                    GROUP BY PackgID 
                                    {booking_date_filter}
                            ) as TB
                            """)
                executed_query = connection.queries[-1]['sql'] if connection.queries else "No query executed"
                print(executed_query)


                df_booking_by_staff = pd.DataFrame(cursor.fetchall(), columns=[row[0] for row in cursor.description])
                total_booking_count = df_booking_by_staff.shape[0]

                print(df_booking_by_staff)



                if total_booking_count == 0:
                    print("No Data")
                    merged_df = query_count_by_staff.copy()
                    merged_df = merged_df.fillna(0)

                    emp_ids = merged_df['UserID'].astype(str).str.upper().tolist()
                    placeholders = ', '.join(['%s'] * len(emp_ids))
                    cursor.execute(f"""
                        SELECT TRIM(UPPER(UserID)) AS UserID, TRIM(UPPER(Empid)) AS Empid
                        FROM tblstaff
                        WHERE UserID IN ({placeholders})
                    """, emp_ids)

                    df_staff_data = pd.DataFrame(cursor.fetchall(), columns=[row[0] for row in cursor.description])

                    merged_df = pd.merge(merged_df, df_staff_data , left_on='UserID', right_on='UserID', how='left')
                    
                    print("Inside")
                    
                    merged_df = merged_df.rename(columns={'Empid':'SalesID' })
                    print(merged_df)

                    merged_df['Booking_Count'] = 0
                    merged_df['Booking_Percentage'] = 0


                else:
                    booking_count_by_staff = (
                        df_booking_by_staff.groupby('SalesID')['PackgID']
                        .count()
                        .reset_index()
                        .rename(columns={"PackgID": "Booking_Count"})
                    )

                    print("Booking")
                    print(booking_count_by_staff)
                    print("Query")
                    print(query_count_by_staff)


                    emp_ids = booking_count_by_staff['SalesID'].astype(str).str.upper().tolist()
                    placeholders = ', '.join(['%s'] * len(emp_ids))
                    cursor.execute(f"""
                        SELECT TRIM(UPPER(UserID)) AS UserID, TRIM(UPPER(Empid)) AS Empid
                        FROM tblstaff
                        WHERE EmpId IN ({placeholders})
                    """, emp_ids)

                    df_staff_data = pd.DataFrame(cursor.fetchall(), columns=[row[0] for row in cursor.description])

                    merged_df = pd.merge(df_staff_data, booking_count_by_staff, left_on='Empid', right_on='SalesID', how='left')

                    print(merged_df)
                    merged_df = merged_df.drop(columns='Empid')

                    # merged_df['StaffName'] = merged_df['StaffName'].str.title()

                   

                    merged_df['Booking_Percentage'] = (
                        merged_df['Booking_Count'] / total_booking_count * 100
                    )

                    merged_df = pd.merge(merged_df, query_count_by_staff, on='UserID', how='outer').fillna(0)

                    print("checling")
                    print(merged_df)

                    merged_df = merged_df.drop(columns=['UserID','SalesID_x'])
                    merged_df = merged_df.rename(columns={'SalesID_y':'SalesID'})
                    
    
                    print("checling 2")
                    print(merged_df)



                    merged_df[['Booking_Count', 'No_of_Query_Assign']] = merged_df[['Booking_Count', 'No_of_Query_Assign']].astype(int)
                    merged_df[['Booking_Percentage', 'Query_Percentage']] = merged_df[['Booking_Percentage', 'Query_Percentage']].round(2)

                    merged_df = merged_df[merged_df['SalesID'] != 0]

                print("--------------MERGED DF----------------")
                print(merged_df)

                if selected_salesid:
                    print(selected_salesid)
                    df_booking_by_single_staff = df_booking_by_staff[df_booking_by_staff['SalesID']== selected_salesid]

                    print(df_booking_by_single_staff)
                    print(df_booking_by_single_staff['PackgID'])


                    if df_booking_by_single_staff is None or df_booking_by_single_staff.empty:
                        print("Nooo")
                        return JsonResponse([], safe=False)

                    df_staff_final_booking = []

                    for _, row in df_booking_by_single_staff.iterrows():
                        pkg_id = row['PackgID']

                        with connection.cursor() as cursor:
                            cursor.execute("""
                                    select PD.PKG_ID, PD.PKG_TITLE, PD.AGENTID, Min(convert(date,TB.tourdate,103)) as Tourdate
                                    from TBL_PKG_DETAILS PD
                                        LEFT JOIN TBL_BOOKING TB
                                    ON PD.PKG_ID = TB.PackgID
                                    where PD.PKG_ID = %s
                                    Group by PD.PKG_ID, PD.PKG_TITLE, PD.AGENTID
                                    """,[pkg_id])

                            executed_query = connection.queries[-1]['sql'] if connection.queries else "No query executed"
                            print(executed_query)

                            final_data = pd.DataFrame(cursor.fetchall(), columns=[row[0] for row in cursor.description])
                            df_staff_final_booking.append(final_data)


                    if df_staff_final_booking:
                        df_staff_final_booking = pd.concat(df_staff_final_booking, ignore_index=True)
                    else:
                        df_staff_final_booking = pd.DataFrame()

                    print(df_staff_final_booking)


                    df_staff_final_booking['PKG_ID'] = df_staff_final_booking['PKG_ID'].astype(str)
                    df_booking_by_single_staff['PackgID'] = df_booking_by_single_staff['PackgID'].astype(str)

                    df_staff_final_booking = pd.merge(df_staff_final_booking, df_booking_by_single_staff, left_on='PKG_ID', right_on='PackgID', how='right')   

                    df_staff_final_booking = df_staff_final_booking[['PKG_ID', 'AGENTID', 'Tourdate', 'PKG_TITLE', 'FirstCreatedDate']]

                    df_staff_final_booking['Tourdate'] = pd.to_datetime(df_staff_final_booking['Tourdate'], format='%d/%m/%Y', errors='coerce')
                    df_staff_final_booking['Tourdate'] = df_staff_final_booking['Tourdate'].dt.strftime('%Y-%m-%d')

                    df_staff_final_booking['FirstCreatedDate'] = pd.to_datetime(df_staff_final_booking['FirstCreatedDate'], errors='coerce')
                    df_staff_final_booking['FirstCreatedDate'] = df_staff_final_booking['FirstCreatedDate'].dt.strftime('%Y-%m-%d %H:%M:%S')


                    print(df_staff_final_booking) 

                    return JsonResponse(df_staff_final_booking.to_dict(orient='records') if not df_staff_final_booking.empty else [], safe=False)


                result = {
                    'Total_Queries' : total_query_count,
                    'Total_Booking' : total_booking_count,
                    'Country_for_Filter': df_country_filter.to_dict(orient='records') if not df_country_filter.empty else [],
                    'Data_to_Show': merged_df.to_dict(orient='records') if isinstance(merged_df, pd.DataFrame) else []
                }
                

                return JsonResponse(result, safe=False)

        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)







    def tour_full_report(request):
        print("api/tour-full-report")

        year = request.GET.get("year")

        if year:
            query = f"and year(convert(date,Tourdate,103)) = {year} "
        else:
            query = "and year(convert(date,Tourdate,103)) > 2022"

        print("Year:", year)
        print("Query:", query)

        try:
            with connection.cursor() as cursor:
                cursor.execute(f"""
                               select 
                                    distinct year(convert(date,tourdate,103)) as Tour_Year
                                from 
                                    TBL_TRAVELLER_NAME t
                                where 
                                    year(convert(date,Tourdate,103)) > year(convert(date,getdate(),103))-3
                                    and Status1='active'
                                    and CREATED_BY not in('CHAGT000100005862','CHAGT000003780','test','CHAGT0001000012263','CHAGT0001000018656') 
                                    and pkgid in(select pkg_id from TBL_PKG_DETAILS where agentid not in (select distinct (Agentid) as Agentid from tbl_agent_culture) and status=1)
                                    and CREATED_BY not in(select distinct (Agentid) as Agentid from tbl_agent_culture)
                                    and PKG_QUERY_ID in(
                                        select QueryID 
                                        from 
                                            tbl_booking 
                                        where 
                                            PackgID=t.Pkgid 
                                            and QueryID=t.PKG_QUERY_ID
                                            and txn_msg='success' 
                                            AND Is_cancelled<>1
                                            )
                                group by 
                                    CREATED_BY,pkgid,tourdate having sum(PaxPaidAmount)>10
                                order by year(convert(date,tourdate,103))
                                    """)
                
                unique_year_df = pd.DataFrame(cursor.fetchall(), columns=[row[0] for row in cursor.description])

                unique_year_list = unique_year_df.iloc[:, 0].dropna().tolist()


                cursor.execute(f"""
                                select 
                                    CREATED_BY as Agent_ID,
                                    (select trim(PKG_TITLE) from TBL_PKG_DETAILS where pkg_id=t.Pkgid) PKG_TITLE,
                                    pkgid,
                                    convert(date,tourdate,103) as tourdate,
                                    count(*) noofguest,
                                    (select count(*) from TBL_TRAVELLER_NAME where CREATED_BY=t.CREATED_BY and TourDate=t.TourDate and pkgid=t.Pkgid and PaxDepositAmount>50 and Status1='active' and PKG_QUERY_ID IN (select QueryID from TBL_BOOKING where txn_msg = 'success' and Is_cancelled <> 1))  paidguest,
                                    (select count(*) from TBL_TRAVELLER_NAME where CREATED_BY=t.CREATED_BY and TourDate=t.TourDate and pkgid=t.Pkgid and isnull(PaxDepositAmount,0)=0 and Status1='active' and PKG_QUERY_ID IN (select QueryID from TBL_BOOKING where txn_msg = 'success' and Is_cancelled <> 1)) nonpaidguest,
                                    (select count(*) from TBL_TRAVELLER_NAME where CREATED_BY=t.CREATED_BY and TourDate=t.TourDate and RoomType='Single' and pkgid=t.Pkgid and Status1='active' and PKG_QUERY_ID IN (select QueryID from TBL_BOOKING where txn_msg = 'success' and Is_cancelled <> 1)) sgloccupancy,
                                    (select count(*) from TBL_TRAVELLER_NAME where CREATED_BY=t.CREATED_BY and TourDate=t.TourDate and RoomType='Double' and pkgid=t.Pkgid  and Status1='active' and PKG_QUERY_ID IN (select QueryID from TBL_BOOKING where txn_msg = 'success' and Is_cancelled <> 1)) dbloccupancy,
                                    (select count(*) from TBL_TRAVELLER_NAME where CREATED_BY=t.CREATED_BY and TourDate=t.TourDate and RoomType='Triple' and pkgid=t.Pkgid  and Status1='active' and PKG_QUERY_ID IN (select QueryID from TBL_BOOKING where txn_msg = 'success' and Is_cancelled <> 1)) tploccupancy,
                                    sum(PKGORIGINALCOST) totalamount,
                                    sum(Commision) Commission,
                                    sum(Makrup) Markup,
                                    (select sum(cast(amount as float)) from TBL_ADD_REMOVE_ADD_ON k where k.agentid=t.CREATED_BY and k.TourDate=t.TourDate and k.pkg_id=t.PKGID and k.Status=1 AND TravID IN(SELECT TRAV_ID FROM TBL_TRAVELLER_NAME WHERE TRAV_ID=K.TravID AND Status1='ACTIVE') and QueryID IN (select QueryID from TBL_BOOKING where txn_msg = 'success' and Is_cancelled <> 1) group by agentid,TourDate,pkg_id ) addoncost,
                                    sum(PaxDepositAmount) paidamount,
                                    (select top 1 tourid 
                                        from 
                                            tbl_booking
                                        where 
                                            agentid=t.CREATED_BY 
                                            and TourDate=t.TourDate 
                                            and PackgID=t.Pkgid) tourid 
                                from 
                                    TBL_TRAVELLER_NAME t
                                where 
                                    Status1='active'
                                    {query}
                                    and CREATED_BY not in('CHAGT000100005862','CHAGT000003780','test','CHAGT0001000012263','CHAGT0001000018656') 
                                    and pkgid in(select pkg_id from TBL_PKG_DETAILS where agentid not in (select distinct (Agentid) as Agentid from tbl_agent_culture) and status=1)
                                    and CREATED_BY not in(select distinct (Agentid) as Agentid from tbl_agent_culture)
                                    and PKG_QUERY_ID in(
                                        select QueryID 
                                        from 
                                            tbl_booking 
                                        where 
                                            PackgID=t.Pkgid 
                                            and QueryID=t.PKG_QUERY_ID
                                            and txn_msg='success' 
                                            AND Is_cancelled<>1
                                            )
                                group by 
                                    CREATED_BY,pkgid,tourdate having sum(PaxPaidAmount)>10
                            """)
                data = pd.DataFrame(cursor.fetchall(), columns=[row[0] for row in cursor.description])

                print(data)

                data['addoncost'] = data['addoncost'].fillna(0)
                data['tourid'] = data['tourid'].fillna('')
                data['totalamount'] = data['totalamount'].fillna(0)
                data['Markup'] = data['Markup'].fillna(0)
                data['Commission'] = data['Commission'].fillna(0)
                data['paidamount'] = data['paidamount'].fillna(0)


                data['tourdate'] = pd.to_datetime(data['tourdate'], errors='coerce')

                data = data[data['tourdate'].notna()]
                
                data['touryear'] = data['tourdate'].dt.year.astype(int)


                data['pkgid'] = data['pkgid'].astype(int)
                data['Commission'] = data['Commission'].astype(int) 
                data['totalamount'] = data['totalamount'].astype(int) 
                data['Markup'] = data['Markup'].astype(int)
                data['paidamount'] = data['paidamount'].astype(int)
                # data['tourdate'] = data['tourdate'].dt.strftime('%Y-%m-%d')

                # data['tourdate'] = data['tourdate'].fillna('No_Date')

            
                result = {
                    "unique_year": unique_year_list,
                    "data": data.to_dict(orient='records')
                }

                return JsonResponse(result, safe=False)

        except Exception as e:
            print("Error:", e)
            return JsonResponse({"error": str(e)}, status=500)




        


    def tour_hotel_details(request):
        print("api/tour-hotel-details")

        year = request.GET.get("year")

        if year:
            query = f"and year(convert(date,Tourdate,103)) = {year}"
        else:
            query = "and year(convert(date,Tourdate,103)) > 2022"

        print("Year:", year)
        print("Query:", query)




        try:
            with connection.cursor() as cursor:
                cursor.execute(f"""
                               select 
                                    distinct year(convert(date,tourdate,103)) as Tour_Year
                                from 
                                    TBL_TRAVELLER_NAME t
                                where 
                                    year(convert(date,Tourdate,103)) > year(convert(date,getdate(),103))-3
                                    and Status1='active'
                                    and CREATED_BY not in('CHAGT000100005862','CHAGT000003780','test','CHAGT0001000012263','CHAGT0001000018656') 
                                    and pkgid in(select pkg_id from TBL_PKG_DETAILS where agentid not in (select distinct (Agentid) as Agentid from tbl_agent_culture) and status=1)
                                    and CREATED_BY not in(select distinct (Agentid) as Agentid from tbl_agent_culture)
                                    and PKG_QUERY_ID in(
                                        select QueryID 
                                        from 
                                            tbl_booking 
                                        where 
                                            PackgID=t.Pkgid 
                                            and QueryID=t.PKG_QUERY_ID
                                            and txn_msg='success' 
                                            AND Is_cancelled<>1
                                            )
                                group by 
                                    CREATED_BY,pkgid,tourdate having sum(PaxPaidAmount)>10
                                order by year(convert(date,tourdate,103))
                                    """)
                
                unique_year_df = pd.DataFrame(cursor.fetchall(), columns=[row[0] for row in cursor.description])

                unique_year_list = unique_year_df.iloc[:, 0].dropna().tolist()



                cursor.execute(f"""
                            select 
                                h.CREATED_BY as Agent_ID,
                                h.PKG_TITLE,
                                h.tourid,
                                h.tourdate,
                                (select HTL_NAME from TBL_HTL_DETAILS where HTL_ID=t.PKG_HTL_ID) as htlname,
                                t.NIGHTS, 
                                cast(dateadd(d,[dbo].[get_hotel_seq_chkin](t.pkg_id,t.PKG_HTL_ID,t.SerialNo),convert(date,h.tourdate,103)) as date) as chkin, 
                                cast(dateadd(d,[dbo].[get_hotel_seq](t.pkg_id,t.PKG_HTL_ID,t.SerialNo),convert(date,h.tourdate,103)) as date) as chkout,
                                h.noofguest,
                                h.dbloccupancy,
                                h.sgloccupancy,
                                h.tploccupancy
                            from TBL_PKG_HOTEL t  
                                left outer join (select 
                                            CREATED_BY ,
                                            (select trim(PKG_TITLE) from TBL_PKG_DETAILS  where pkg_id=t.Pkgid) PKG_TITLE,
                                            pkgid,
                                            convert(date,tourdate,103) as tourdate,
                                            count(*) noofguest,
                                            (select count(*) from TBL_TRAVELLER_NAME where CREATED_BY=t.CREATED_BY and TourDate=t.TourDate and pkgid=t.Pkgid and PaxDepositAmount>50 and Status1='active' and PKG_QUERY_ID IN (select QueryID from TBL_BOOKING where txn_msg = 'success' and Is_cancelled <> 1))  paidguest,
                                            (select count(*) from TBL_TRAVELLER_NAME where CREATED_BY=t.CREATED_BY and TourDate=t.TourDate and pkgid=t.Pkgid and isnull(PaxDepositAmount,0)=0 and Status1='active' and PKG_QUERY_ID IN (select QueryID from TBL_BOOKING where txn_msg = 'success' and Is_cancelled <> 1)) nonpaidguest,
                                            (select count(*) from TBL_TRAVELLER_NAME where CREATED_BY=t.CREATED_BY and TourDate=t.TourDate and RoomType='Single' and pkgid=t.Pkgid and Status1='active' and PKG_QUERY_ID IN (select QueryID from TBL_BOOKING where txn_msg = 'success' and Is_cancelled <> 1)) sgloccupancy,
                                            (select count(*) from TBL_TRAVELLER_NAME where CREATED_BY=t.CREATED_BY and TourDate=t.TourDate and RoomType='Double' and pkgid=t.Pkgid  and Status1='active' and PKG_QUERY_ID IN (select QueryID from TBL_BOOKING where txn_msg = 'success' and Is_cancelled <> 1)) dbloccupancy,
                                            (select count(*) from TBL_TRAVELLER_NAME where CREATED_BY=t.CREATED_BY and TourDate=t.TourDate and RoomType='Triple' and pkgid=t.Pkgid  and Status1='active' and PKG_QUERY_ID IN (select QueryID from TBL_BOOKING where txn_msg = 'success' and Is_cancelled <> 1)) tploccupancy,
                                            sum(PKGORIGINALCOST) totalamount,
                                            sum(Commision) Commission,
                                            sum(Makrup) Markup,
                                            (select sum(cast(amount as float)) from TBL_ADD_REMOVE_ADD_ON k where k.agentid=t.CREATED_BY and k.TourDate=t.TourDate and k.pkg_id=t.PKGID and k.Status=1 AND TravID IN(SELECT TRAV_ID FROM TBL_TRAVELLER_NAME WHERE TRAV_ID=K.TravID AND Status1='ACTIVE') and QueryID IN (select QueryID from TBL_BOOKING where txn_msg = 'success' and Is_cancelled <> 1) group by agentid,TourDate,pkg_id ) addoncost,
                                            sum(PaxDepositAmount) paidamount,
                                            (select top 1 tourid 
                                                from 
                                                    tbl_booking
                                                where 
                                                    agentid=t.CREATED_BY 
                                                    and TourDate=t.TourDate 
                                                    and PackgID=t.Pkgid) tourid 
                                        from 
                                            TBL_TRAVELLER_NAME t
                                        where 
                                            convert(date,Tourdate,103) > convert(date,getdate(),103)
                                            and Status1='active'
                                            and CREATED_BY not in('CHAGT000100005862','CHAGT000003780','test','CHAGT0001000012263','CHAGT0001000018656') 
                                            and pkgid in(select pkg_id from TBL_PKG_DETAILS where agentid not in (select distinct (Agentid) as Agentid from tbl_agent_culture) and status=1)
                                            and CREATED_BY not in(select distinct (Agentid) as Agentid from tbl_agent_culture)
                                            and PKG_QUERY_ID in(
                                                select QueryID 
                                                from 
                                                    tbl_booking 
                                                where 
                                                    PackgID=t.Pkgid 
                                                    and QueryID=t.PKG_QUERY_ID
                                                    and txn_msg='success' 
                                                    AND Is_cancelled<>1
                                                    )
                                        group by 
                                            CREATED_BY, pkgid,tourdate having sum(PaxPaidAmount)>10) h 
                            on t.PKG_ID=h.pkgid 
                            where t.PKG_ID is not null 
                                and h.pkgid is not null 
                                and t.Status=1 
                                and h.CREATED_BY not in (select distinct (Agentid) as Agentid from tbl_agent_culture)
                            order by t.PKG_ID,t.SerialNo
                               """)
                
                data = pd.DataFrame(cursor.fetchall(), columns=[row[0] for row in cursor.description])
                print(data)

                data['tourid'] = data['tourid'].fillna('')

                data['noofguest'] = data['noofguest'].fillna(0)
                data['dbloccupancy'] = data['dbloccupancy'].fillna(0)
                data['sgloccupancy'] = data['sgloccupancy'].fillna(0)
                data['tploccupancy'] = data['tploccupancy'].fillna(0)


                data['tourdate'] = pd.to_datetime(data['tourdate'], errors='coerce')

                data = data[data['tourdate'].notna()]
                data['touryear'] = data['tourdate'].dt.year.astype(int)
                # data['tourdate'] = data['tourdate'].dt.strftime('%Y-%m-%d')

                # data['tourdate'] = data['tourdate'].fillna('No_Date') 
                
                data['NIGHTS'] = data['NIGHTS'].astype(int)
                data['noofguest'] = data['noofguest'].astype(int)
                data['dbloccupancy'] = data['dbloccupancy'].astype(int)
                data['sgloccupancy'] = data['sgloccupancy'].astype(int)
                data['tploccupancy'] = data['tploccupancy'].astype(int)




                result = {
                    "unique_year": unique_year_list,
                    "data": data.to_dict(orient='records')
                }

                return JsonResponse(result, safe=False)



        except Exception as e:
            print("Error:", e)
            return JsonResponse({"error": str(e)}, status=500)



    # def request_data(request):
    #     print("Test")

    #     try:
    #         with connection.cursor() as cursor:
    #             cursor.execute(f""" SELECT CONCAT(UPPER(trim(agentId)), PackgID, convert(date,tourdate,103)) AS UniqueID, 
    #                     trim(upper(agentId)) as AgentID
    #                 from TBL_BOOKING
    #                 where txn_msg = 'success'
    #                     AND Is_cancelled <> 1       
    #                     AND CreatedDate is NOT NULL
    #                     AND tourdate is NOT NULL
    #                     AND agentId is NOT NULL
    #                     AND PackgID Is NOT NULL
    #                     AND CreatedDate != ''
    #                     AND tourdate != ''
    #                     AND agentId != ''
    #                     AND PackgID != ''
    #                     AND agentId NOT In (select distinct (Agentid) as Agentid from tbl_agent_culture)
    #                            """)
    #             df_booking = pd.DataFrame(cursor.fetchall(), columns=[row[0] for row in cursor.description])

    #         print(df_booking)
            
    #         df_booking = df_booking.drop_duplicates(subset=['UniqueID']).reset_index()
    #         df_booking = df_booking.groupby(['AgentID']).size().reset_index(name='No_of_Booking')

    #         df_booking = (df_booking[df_booking['No_of_Booking'] > 1]).reset_index()
    #         print(df_booking)

    #         agent_ids = df_booking['AgentID'].tolist()

    #         if agent_ids:
    #             placeholders = ",".join(["%s"] * len(agent_ids)) 
    #             query = f"""
    #                 SELECT Name, Emailid, AgentID
    #                 FROM tbl_agent
    #                 WHERE status = 1
    #                 AND AgentID IN ({placeholders})
    #             """

    #             with connection.cursor() as cursor:
    #                 cursor.execute(query, agent_ids)
    #                 df_agent = pd.DataFrame(cursor.fetchall(), columns=[row[0] for row in cursor.description])

    #         print(df_agent)

    #         final_df = pd.merge(df_agent, df_booking, on='AgentID', how='left')

    #         print(final_df)
    #         final_df = final_df.drop(columns=['index'])

    #         final_df = final_df.sort_values(by='No_of_Booking', ascending=False).reset_index(drop=True)
    #         print(final_df)


    #         ## "For Downloading the File in code_terminal"
    #         # csv_buffer = StringIO()
    #         # data.to_csv(csv_buffer, index=False)  ## Changed 'data' with Dataframe
    #         # csv_data = csv_buffer.getvalue()

    #         # # Create response for direct download
    #         # response = HttpResponse(csv_data, content_type='text/csv')
    #         # response['Content-Disposition'] = 'attachment; filename="downloaded_file.csv"'

    #         # globals()['response'] = response

    #         # print("Download Complete")  

    #         return JsonResponse("Done", safe=False)

    #     except Exception as e:
    #         print("Error:", e)
    #         return JsonResponse({"error": str(e)}, status=500)

        
